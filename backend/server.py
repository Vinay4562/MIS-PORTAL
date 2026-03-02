from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, ValidationError
from typing import List, Optional, Dict, Any, Tuple
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
import io
import smtplib
import calendar
import re
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import random
import string
import secrets
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def format_date(date_str):
    if not date_str:
        return "-"
    try:
        parts = date_str.split("-")
        if len(parts) == 3:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
    except:
        pass
    return date_str

def format_time(time_str):
    if not time_str:
        return ""
    s = str(time_str).strip()
    if s in ("-", "N/S"):
        return s
    upper = s.upper()
    match_12h = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?\s*([AP]M)$", upper)
    if match_12h:
        hour = int(match_12h.group(1))
        minute = match_12h.group(2)
        period = match_12h.group(4)
        if period == "PM" and hour != 12:
            hour += 12
        if period == "AM" and hour == 12:
            hour = 0
        return f"{hour:02d}:{minute}"
    if ":" in s:
        parts = s.split(":")
        if len(parts) >= 2 and parts[0].isdigit():
            minute_part = parts[1][:2]
            if minute_part.isdigit():
                hour = int(parts[0])
                return f"{hour:02d}:{minute_part.zfill(2)}"
    return s


def format_duration_hhmm(value):
    if value is None or value == "":
        return ""
    try:
        minutes = float(value)
    except Exception:
        return str(value)
    total = int(round(minutes))
    if total <= 0:
        return ""
    days = total // (24 * 60)
    remaining = total % (24 * 60)
    hours = remaining // 60
    mins = remaining % 60
    if days == 0:
        return f"{hours:02d}:{mins:02d}"
    if days == 1:
        return f"1Day {hours:02d}:{mins:02d} Hours"
    return f"{days}Days {hours:02d}:{mins:02d} Hours"

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import socket
import ssl

class IPv4SMTP(smtplib.SMTP):
    """SMTP client that forces IPv4 connection and tries all resolved addresses."""
    def _get_socket(self, host, port, timeout):
        print(f"DEBUG: IPv4SMTP connecting to {host}:{port}")
        
        # Get all IPv4 addresses
        try:
            addr_info = socket.getaddrinfo(host, port, socket.AF_INET, socket.SOCK_STREAM)
            print(f"DEBUG: Resolved {host} to {len(addr_info)} addresses: {[r[4] for r in addr_info]}")
        except socket.gaierror as e:
            print(f"DEBUG: IPv4 resolution failed: {e}")
            raise

        # Try each address until one works (Happy Eyeballs for IPv4)
        last_err = None
        for i, res in enumerate(addr_info):
            af, socktype, proto, canonname, sa = res
            print(f"DEBUG: Connecting to IP {i+1}/{len(addr_info)}: {sa}")
            try:
                sock = socket.socket(af, socktype, proto)
                try:
                    sock.settimeout(timeout)
                    if self.source_address:
                        sock.bind(self.source_address)
                    sock.connect(sa)
                    print(f"DEBUG: Connected to {sa}")
                    return sock # Success
                except OSError as e:
                    print(f"DEBUG: Connection to {sa} failed: {e}")
                    last_err = e
                    sock.close()
            except OSError as e:
                print(f"DEBUG: Socket creation failed: {e}")
                last_err = e
                pass
        
        # If we get here, all attempts failed
        if last_err:
            raise last_err
        raise OSError("No IPv4 addresses found or all connection attempts failed")

class IPv4SMTP_SSL(smtplib.SMTP_SSL):
    """SMTP_SSL client that forces IPv4 connection and tries all resolved addresses."""
    def _get_socket(self, host, port, timeout):
        print(f"DEBUG: IPv4SMTP_SSL connecting to {host}:{port}")
        
        try:
            addr_info = socket.getaddrinfo(host, port, socket.AF_INET, socket.SOCK_STREAM)
            print(f"DEBUG: Resolved {host} to {len(addr_info)} addresses: {[r[4] for r in addr_info]}")
        except socket.gaierror as e:
            print(f"DEBUG: IPv4 resolution failed: {e}")
            raise

        last_err = None
        for i, res in enumerate(addr_info):
            af, socktype, proto, canonname, sa = res
            print(f"DEBUG: Connecting to IP {i+1}/{len(addr_info)}: {sa}")
            try:
                sock = socket.socket(af, socktype, proto)
                try:
                    sock.settimeout(timeout)
                    if self.source_address:
                        sock.bind(self.source_address)
                    sock.connect(sa)
                    print(f"DEBUG: Connected to {sa}, wrapping SSL...")
                    
                    # Connection successful, now wrap SSL
                    try:
                        # Use 'host' argument explicitly for SNI
                        ssl_sock = self.context.wrap_socket(sock, server_hostname=host)
                        print(f"DEBUG: SSL Handshake success")
                        return ssl_sock
                    except Exception as ssl_err:
                        print(f"DEBUG: SSL Wrap failed: {ssl_err}")
                        sock.close()
                        raise ssl_err
                        
                except Exception as e:
                    print(f"DEBUG: Connection to {sa} failed: {e}")
                    last_err = e
                    sock.close()
            except Exception as e:
                last_err = e
                pass
                
        if last_err:
            raise last_err
        raise OSError("No IPv4 addresses found or all connection attempts failed")

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
import base64

def _send_email_core(to_email: str, subject: str, message: MIMEMultipart):
    """Core function to handle email sending using Gmail API."""
    sender_email = os.environ.get("SMTP_EMAIL")
    
    # Gmail API Credentials
    client_id = os.environ.get("GOOGLE_CLIENT_ID")
    client_secret = os.environ.get("GOOGLE_CLIENT_SECRET")
    refresh_token = os.environ.get("GOOGLE_REFRESH_TOKEN")

    if not all([sender_email, client_id, client_secret, refresh_token]):
        # Fallback to SMTP if Gmail API creds are missing (for backward compatibility/local dev)
        print("Warning: Gmail API credentials missing. Falling back to SMTP...")
        _send_email_smtp_fallback(to_email, subject, message)
        return

    try:
        # Construct Credentials object
        creds = Credentials(
            None, # access_token (will be refreshed)
            refresh_token=refresh_token,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=client_id,
            client_secret=client_secret
        )

        # Build Gmail Service
        service = build('gmail', 'v1', credentials=creds)

        # Create raw message
        # message is already a MIMEMultipart object
        # We need to encode it to base64url
        raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
        body = {'raw': raw_message}

        # Send email
        print(f"Sending email via Gmail API to {to_email}...")
        message = service.users().messages().send(userId="me", body=body).execute()
        print(f"Email sent successfully! Message Id: {message['id']}")
        return

    except Exception as e:
        print(f"Gmail API Error: {e}")
        # Try fallback if API fails? Or just raise?
        # If API fails (e.g. auth error), SMTP might not work either if blocked.
        # But let's try fallback just in case we are on a system where SMTP works but API failed.
        print("Attempting fallback to SMTP...")
        try:
            _send_email_smtp_fallback(to_email, subject, message)
        except Exception as smtp_err:
             print(f"Fallback SMTP also failed: {smtp_err}")
             raise e # Raise the original API error as it's likely the primary config

def _send_email_smtp_fallback(to_email: str, subject: str, message: MIMEMultipart):
    """Legacy SMTP sending logic as fallback."""
    sender_email = os.environ.get("SMTP_EMAIL")
    sender_password = os.environ.get("SMTP_PASSWORD")
    
    # Default to port 587 (STARTTLS)
    smtp_server = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
    env_port = int(os.environ.get("SMTP_PORT", 587))
    
    if not sender_email or not sender_password:
        raise ValueError("SMTP configuration (email/password) missing in environment variables")

    # Define attempts: [primary_config, fallback_config]
    attempts = []
    
    # Attempt 1: Configured port
    attempts.append((smtp_server, env_port))
    
    # Attempt 2: The alternative port (if 587 -> 465, if 465 -> 587)
    if env_port == 587:
        attempts.append((smtp_server, 465))
    elif env_port == 465:
        attempts.append((smtp_server, 587))
    else:
        # If user configured weird port (e.g. 25), try both standard ones as fallback
        attempts.append((smtp_server, 587))
        attempts.append((smtp_server, 465))

    last_error = None

    for i, (server_host, port) in enumerate(attempts):
        print(f"SMTP Attempt {i+1}/{len(attempts)}: Connecting to {server_host}:{port} for {to_email}")
        try:
            if port == 465:
                # Implicit SSL - Use IPv4 forced class
                with IPv4SMTP_SSL(server_host, port, timeout=60) as server:
                    server.login(sender_email, sender_password)
                    server.sendmail(sender_email, to_email, message.as_string())
            else:
                # Explicit SSL (STARTTLS) - usually port 587 - Use IPv4 forced class
                with IPv4SMTP(server_host, port, timeout=60) as server:
                    server.starttls()
                    server.login(sender_email, sender_password)
                    server.sendmail(sender_email, to_email, message.as_string())
            
            print(f"Email sent successfully to {to_email} using port {port}")
            return # Success!
            
        except Exception as e:
            print(f"SMTP Attempt {i+1} failed ({server_host}:{port}): {e}")
            last_error = e
            # Continue to next attempt
    
    # If we exit the loop, all attempts failed
    print(f"All SMTP attempts failed. Last error: {last_error}")
    if last_error:
        raise last_error
    else:
        raise HTTPException(status_code=500, detail="Failed to send email (unknown error)")
            
def send_otp_email(user_email: str, otp: str, reason: str = "reset"):
    sender_email = os.environ.get("SMTP_EMAIL")
    # Admin email is same as sender in this requirement
    admin_email = sender_email 
    
    message = MIMEMultipart("alternative")
    message["From"] = sender_email
    message["To"] = admin_email

    if reason == "signup":
        message["Subject"] = "New Account Signup OTP Request"
        action_text = "new account signup"
    else:
        message["Subject"] = "Password Reset OTP Request"
        action_text = "password reset"

    text = f"""\
    A {action_text} was requested for user: {user_email}.
    Your OTP is: {otp}
    """
    html = f"""\
    <html>
      <body>
        <h2>{message["Subject"]}</h2>
        <p>A {action_text} was requested for user: <b>{user_email}</b>.</p>
        <p>Your OTP is: <b style="font-size: 24px;">{otp}</b></p>
        <p>Please approve this request by sharing the OTP with the user if appropriate.</p>
      </body>
    </html>
    """

    part1 = MIMEText(text, "plain")
    part2 = MIMEText(html, "html")
    message.attach(part1)
    message.attach(part2)

    try:
        _send_email_core(admin_email, message["Subject"], message)
    except Exception as e:
        print(f"Error sending OTP email: {e}")
        raise HTTPException(status_code=500, detail="Failed to send OTP email")

def send_reports_email(recipient_email: str, attachments: list, subject: str):
    try:
        sender_email = os.environ.get("SMTP_EMAIL")
        
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = recipient_email
        message["Subject"] = subject

        body = "Please find the attached reports for the selected period."
        message.attach(MIMEText(body, "plain"))

        for filename, content in attachments:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(content)
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {filename}",
            )
            message.attach(part)

        _send_email_core(recipient_email, message["Subject"], message)
        
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"Error sending reports email: {e}")
        print(f"Traceback: {error_msg}")
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()

@app.on_event("startup")
async def startup_db_client():
    # Create indexes for performance
    try:
        # Users
        await db.users.create_index("email", unique=True)
        await db.users.create_index("id", unique=True)
        
        # Feeders
        await db.feeders.create_index("id", unique=True)
        
        # Entries
        await db.entries.create_index("id", unique=True)
        await db.entries.create_index([("feeder_id", 1), ("date", 1)])
        await db.entries.create_index("date")
        
        # Max Min Feeders
        await db.max_min_feeders.create_index("id", unique=True)
        
        # Max Min Entries
        await db.max_min_entries.create_index("id", unique=True)
        await db.max_min_entries.create_index([("feeder_id", 1), ("date", 1)])
        
        print("Database indexes created successfully")
    except Exception as e:
        print(f"Error creating indexes: {e}")


origins = [
    "http://localhost:3000",
    "http://localhost:3001",
    "https://mis-portal-liard.vercel.app",
    "https://mis-portal-production.up.railway.app",
    "https://mis-portal.vercel.app"
]

# Ensure we allow Vercel and Railway subdomains dynamically
# and allow credentials to be sent (cookies, auth headers)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=origins, # Specific origins
    allow_origin_regex=r"https://.*\.vercel\.app|https://.*\.railway\.app", # Regex for dynamic subdomains
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

api_router = APIRouter(prefix="/api")

@api_router.get("/ping")
async def ping():
    return {"status": "ok"}

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

SECRET_KEY = os.environ['JWT_SECRET_KEY']
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7

class UserRegister(BaseModel):
    email: str
    password: str
    full_name: str

class UserLogin(BaseModel):
    email: str
    password: str

class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    email: str
    otp: str
    new_password: str

class AdminOtpVerifyRequest(BaseModel):
    email: str
    otp: str

class AdminResetPasswordRequest(BaseModel):
    email: str
    reset_token: str
    new_password: str

class AdminChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

class SignupVerifyRequest(BaseModel):
    email: str
    otp: str

class EmailReportRequest(BaseModel):
    email: str
    year: int
    month: int
    report_ids: Optional[List[str]] = None


class AdminMaxMinExportColumn(BaseModel):
    key: str
    label: str
    field: str


class AdminMaxMinExportViewPayload(BaseModel):
    columns: List[AdminMaxMinExportColumn]
    rows: List[dict]
    meta: Optional[dict] = None

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    full_name: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserInDB(User):
    hashed_password: str
    admin_hashed_password: Optional[str] = ""

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class Feeder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    end1_name: str
    end2_name: str
    end1_import_mf: float
    end1_export_mf: float
    end2_import_mf: float
    end2_export_mf: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DailyEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    feeder_id: str
    date: str
    end1_import_initial: float
    end1_import_final: float
    end1_export_initial: float
    end1_export_final: float
    end2_import_initial: float
    end2_import_final: float
    end2_export_initial: float
    end2_export_final: float
    end1_import_consumption: float = 0
    end1_export_consumption: float = 0
    end2_import_consumption: float = 0
    end2_export_consumption: float = 0
    loss_percent: float = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DailyEntryCreate(BaseModel):
    feeder_id: str
    date: str
    end1_import_final: float
    end1_export_final: float
    end2_import_final: float
    end2_export_final: float

class DailyEntryUpdate(BaseModel):
    end1_import_initial: Optional[float] = None
    end1_import_final: Optional[float] = None
    end1_export_initial: Optional[float] = None
    end1_export_final: Optional[float] = None
    end2_import_initial: Optional[float] = None
    end2_import_final: Optional[float] = None
    end2_export_initial: Optional[float] = None
    end2_export_final: Optional[float] = None

class EnergySheet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    meters: Optional[List['EnergyMeter']] = None

class EnergyMeter(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sheet_id: str
    name: str
    mf: float
    unit: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EnergyReading(BaseModel):
    model_config = ConfigDict(extra="ignore")
    meter_id: str
    initial: float
    final: float
    consumption: float

class EnergyEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sheet_id: str
    date: str
    readings: List[EnergyReading]
    total_consumption: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EnergyReadingInput(BaseModel):
    meter_id: str
    final: float

class EnergyEntryInput(BaseModel):
    sheet_id: str
    date: str
    readings: List[EnergyReadingInput]

class MaxMinFeeder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: str  # bus_station, feeder_400kv, feeder_220kv, ict_feeder
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MaxMinEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    feeder_id: str
    date: str
    data: dict
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MaxMinEntryCreate(BaseModel):
    feeder_id: str
    date: str
    data: dict

class InterruptionEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    feeder_id: str
    date: str
    data: dict
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class InterruptionEntryUpdate(BaseModel):
    date: Optional[str] = None
    data: Optional[Dict[str, Any]] = None


class InterruptionEntryCreate(BaseModel):
    date: str
    data: Dict[str, Any]

class InterruptionsImportPayload(BaseModel):
    feeder_id: str
    entries: List[dict]

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

ADMIN_EMAILS = os.environ.get("ADMIN_EMAILS", "")
ADMIN_EMAIL_SET = {e.strip().lower() for e in ADMIN_EMAILS.split(",") if e.strip()}


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        try:
            return User(**user)
        except ValidationError as e:
            print(f"User validation error for {user_id}: {e}")
            # Attempt to return user with minimal fields if validation fails
            return User(email=user.get("email", ""), id=user_id, full_name=user.get("full_name", ""))
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")


async def get_current_admin(current_user: User = Depends(get_current_user)):
    email = (current_user.email or "").lower()
    if email not in ADMIN_EMAIL_SET:
        raise HTTPException(status_code=403, detail="Not authorised for admin access")
    return current_user


@api_router.get("/admin/me", response_model=User)
async def get_admin_me(current_admin: User = Depends(get_current_admin)):
    return current_admin

@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserRegister):
    existing_user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed_password = get_password_hash(user_data.password)
    user_obj = UserInDB(
        email=user_data.email,
        full_name=user_data.full_name,
        hashed_password=hashed_password
    )
    
    doc = user_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    user_response = User(id=user_obj.id, email=user_obj.email, full_name=user_obj.full_name, created_at=user_obj.created_at)
    access_token = create_access_token(data={"sub": user_obj.id})
    
    return Token(access_token=access_token, token_type="bearer", user=user_response)

@api_router.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin):
    user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(user_data.password, user['hashed_password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    user_response = User(**user)
    access_token = create_access_token(data={"sub": user['id']})
    
    return Token(access_token=access_token, token_type="bearer", user=user_response)


@api_router.post("/admin/auth/login", response_model=Token)
async def admin_login(user_data: UserLogin):
    user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    email = (user.get("email") or "").lower()
    if email not in ADMIN_EMAIL_SET:
        raise HTTPException(status_code=403, detail="Not authorised for admin access")

    admin_hash = user.get("admin_hashed_password")
    if admin_hash:
        if not verify_password(user_data.password, admin_hash):
            raise HTTPException(status_code=401, detail="Invalid credentials")
    else:
        if not verify_password(user_data.password, user.get("hashed_password", "")):
            raise HTTPException(status_code=401, detail="Invalid credentials")

    user_response = User(**user)
    access_token = create_access_token(data={"sub": user['id']})

    return Token(access_token=access_token, token_type="bearer", user=user_response)

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    user = await db.users.find_one({"email": request.email})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    otp = ''.join(random.choices(string.digits, k=6))
    # Save OTP with expiry (15 mins)
    await db.password_resets.update_one(
        {"email": request.email},
        {"$set": {"otp": otp, "expires_at": datetime.now(timezone.utc) + timedelta(minutes=15)}},
        upsert=True
    )
    
    # Send to admin
    send_otp_email(request.email, otp)
    
    return {"message": "OTP sent to admin for approval"}

@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    record = await db.password_resets.find_one({"email": request.email})
    if not record:
        raise HTTPException(status_code=400, detail="Invalid request")
        
    if record["otp"] != request.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP")
        
    # Handle both naive and aware datetime for expires_at
    expires_at = record["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
        
    if expires_at < datetime.now(timezone.utc):
         raise HTTPException(status_code=400, detail="OTP expired")
         
    # Update password
    hashed_password = get_password_hash(request.new_password)
    await db.users.update_one(
        {"email": request.email},
        {"$set": {"hashed_password": hashed_password}}
    )
    
    # Delete OTP
    await db.password_resets.delete_one({"email": request.email})
    
    return {"message": "Password reset successful"}


@api_router.post("/admin/auth/forgot-password")
async def admin_forgot_password(request: ForgotPasswordRequest):
    email = request.email.lower()
    if email not in ADMIN_EMAIL_SET:
        return {"message": "If this admin account exists, an OTP has been sent"}
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        return {"message": "If this admin account exists, an OTP has been sent"}
    now = datetime.now(timezone.utc)
    record = await db.admin_password_resets.find_one({"email": email})
    if record:
        last_sent = record.get("last_sent_at")
        if isinstance(last_sent, str):
            try:
                last_sent = datetime.fromisoformat(last_sent)
            except Exception:
                last_sent = None
        if last_sent:
            if last_sent.tzinfo is None:
                last_sent = last_sent.replace(tzinfo=timezone.utc)
            delta = now - last_sent
            if delta.total_seconds() < 60:
                raise HTTPException(status_code=429, detail="OTP already sent recently. Please wait before requesting again")
            send_count = int(record.get("send_count", 0))
            first_sent = record.get("created_at") or last_sent
            if isinstance(first_sent, str):
                try:
                    first_sent = datetime.fromisoformat(first_sent)
                except Exception:
                    first_sent = now
            if first_sent.tzinfo is None:
                first_sent = first_sent.replace(tzinfo=timezone.utc)
            if (now - first_sent).total_seconds() < 3600 and send_count >= 5:
                raise HTTPException(status_code=429, detail="Too many OTP requests. Please try again later")
    otp = "".join(random.choices(string.digits, k=6))
    otp_hash = get_password_hash(otp)
    payload = {
        "email": email,
        "otp_hash": otp_hash,
        "expires_at": now + timedelta(minutes=10),
        "created_at": record.get("created_at") if record else now,
        "last_sent_at": now,
        "send_count": int(record.get("send_count", 0)) + 1 if record else 1,
        "attempts": 0,
        "otp_verified": False,
        "reset_token": None,
        "reset_token_expires_at": None,
    }
    if isinstance(payload["created_at"], str):
        try:
            payload["created_at"] = datetime.fromisoformat(payload["created_at"])
        except Exception:
            payload["created_at"] = now
    await db.admin_password_resets.update_one(
        {"email": email},
        {"$set": payload},
        upsert=True,
    )
    try:
        smtp_sender = os.environ.get("SMTP_EMAIL")
        sender_email = smtp_sender or os.environ.get("SMTP_USER")
        message = MIMEMultipart("alternative")
        message["From"] = sender_email or ""
        message["To"] = email
        message["Subject"] = "Admin Panel Password Reset OTP"
        text = f"An admin panel password reset was requested for {email}. Your OTP is: {otp}"
        html = f"""
        <html>
          <body>
            <h2>Admin Panel Password Reset</h2>
            <p>An admin panel password reset was requested for <b>{email}</b>.</p>
            <p>Your OTP is: <b style="font-size: 24px;">{otp}</b></p>
            <p>This OTP will expire in 10 minutes.</p>
          </body>
        </html>
        """
        part1 = MIMEText(text, "plain")
        part2 = MIMEText(html, "html")
        message.attach(part1)
        message.attach(part2)
        if sender_email:
            _send_email_core(email, message["Subject"], message)
        else:
            print("SMTP_EMAIL/SMTP_USER not configured; cannot send admin OTP email")
    except Exception as e:
        print(f"Error sending admin password reset OTP email: {e}")
    return {"message": "If this admin account exists, an OTP has been sent"}


@api_router.post("/admin/auth/verify-otp")
async def admin_verify_otp(request: AdminOtpVerifyRequest):
    email = request.email.lower()
    if email not in ADMIN_EMAIL_SET:
        raise HTTPException(status_code=400, detail="Invalid OTP or email")
    record = await db.admin_password_resets.find_one({"email": email})
    if not record:
        raise HTTPException(status_code=400, detail="Invalid OTP or email")
    now = datetime.now(timezone.utc)
    expires_at = record.get("expires_at")
    if isinstance(expires_at, str):
        try:
            expires_at = datetime.fromisoformat(expires_at)
        except Exception:
            expires_at = None
    if expires_at:
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        if expires_at < now:
            raise HTTPException(status_code=400, detail="OTP expired")
    attempts = int(record.get("attempts", 0))
    if attempts >= 5:
        raise HTTPException(status_code=400, detail="Too many incorrect attempts. Request a new OTP")
    otp_hash = record.get("otp_hash") or ""
    if not verify_password(request.otp, otp_hash):
        await db.admin_password_resets.update_one(
            {"email": email},
            {"$set": {"attempts": attempts + 1}},
        )
        raise HTTPException(status_code=400, detail="Invalid OTP")
    reset_token = secrets.token_urlsafe(32)
    await db.admin_password_resets.update_one(
        {"email": email},
        {
            "$set": {
                "otp_verified": True,
                "reset_token": reset_token,
                "reset_token_expires_at": now + timedelta(minutes=15),
            }
        },
    )
    return {"reset_token": reset_token}


@api_router.post("/admin/auth/reset-password")
async def admin_reset_password(request: AdminResetPasswordRequest):
    email = request.email.lower()
    if email not in ADMIN_EMAIL_SET:
        raise HTTPException(status_code=400, detail="Invalid reset token or email")
    record = await db.admin_password_resets.find_one({"email": email})
    if not record:
        raise HTTPException(status_code=400, detail="Invalid reset token or email")
    if not record.get("otp_verified"):
        raise HTTPException(status_code=400, detail="OTP verification required")
    stored_token = record.get("reset_token") or ""
    if not stored_token or stored_token != request.reset_token:
        raise HTTPException(status_code=400, detail="Invalid reset token or email")
    now = datetime.now(timezone.utc)
    token_expires = record.get("reset_token_expires_at")
    if isinstance(token_expires, str):
        try:
            token_expires = datetime.fromisoformat(token_expires)
        except Exception:
            token_expires = None
    if token_expires:
        if token_expires.tzinfo is None:
            token_expires = token_expires.replace(tzinfo=timezone.utc)
        if token_expires < now:
            raise HTTPException(status_code=400, detail="Reset token expired")
    if len(request.new_password or "") < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters long")
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=400, detail="User not found")
    hashed_password = get_password_hash(request.new_password)
    await db.users.update_one(
        {"email": email},
        {"$set": {"admin_hashed_password": hashed_password}},
    )
    await db.admin_password_resets.delete_one({"email": email})
    return {"message": "Admin password reset successful"}


@api_router.post("/admin/auth/change-password")
async def admin_change_password(request: AdminChangePasswordRequest, current_admin: User = Depends(get_current_admin)):
    if len(request.new_password or "") < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters long")
    user = await db.users.find_one({"id": current_admin.id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=400, detail="User not found")
    current_admin_hash = user.get("admin_hashed_password") or user.get("hashed_password", "")
    if not verify_password(request.current_password, current_admin_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    hashed_password = get_password_hash(request.new_password)
    await db.users.update_one(
        {"id": current_admin.id},
        {"$set": {"admin_hashed_password": hashed_password}},
    )
    await db.admin_password_resets.delete_one({"email": current_admin.email.lower()})
    return {"message": "Password updated successfully"}

@api_router.post("/auth/signup-request")
async def signup_request(user_data: UserRegister):
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    otp = ''.join(random.choices(string.digits, k=6))
    hashed_password = get_password_hash(user_data.password)
    
    request_data = {
        "email": user_data.email,
        "full_name": user_data.full_name,
        "hashed_password": hashed_password,
        "otp": otp,
        "expires_at": datetime.now(timezone.utc) + timedelta(minutes=15),
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.signup_requests.update_one(
        {"email": user_data.email},
        {"$set": request_data},
        upsert=True
    )
    
    send_otp_email(user_data.email, otp, reason="signup")
    
    return {"message": "OTP sent to admin for approval"}

@api_router.post("/auth/signup-verify", response_model=Token)
async def signup_verify(request: SignupVerifyRequest):
    record = await db.signup_requests.find_one({"email": request.email})
    if not record:
        raise HTTPException(status_code=400, detail="Invalid request or expired")
        
    if record["otp"] != request.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP")
        
    expires_at = record["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
        
    if expires_at < datetime.now(timezone.utc):
         raise HTTPException(status_code=400, detail="OTP expired")
    
    # Create user
    user_obj = UserInDB(
        email=record["email"],
        full_name=record["full_name"],
        hashed_password=record["hashed_password"]
    )
    
    doc = user_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    # Delete request
    await db.signup_requests.delete_one({"email": request.email})
    
    user_response = User(id=user_obj.id, email=user_obj.email, full_name=user_obj.full_name, created_at=user_obj.created_at)
    access_token = create_access_token(data={"sub": user_obj.id})
    
    return Token(access_token=access_token, token_type="bearer", user=user_response)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

@api_router.get("/feeders", response_model=List[Feeder])
async def get_feeders(current_user: User = Depends(get_current_user)):
    feeders = await db.feeders.find({}, {"_id": 0}).to_list(100)
    for feeder in feeders:
        if isinstance(feeder.get('created_at'), str):
            feeder['created_at'] = datetime.fromisoformat(feeder['created_at'])
    return feeders

@api_router.get("/feeders/{feeder_id}", response_model=Feeder)
async def get_feeder(feeder_id: str, current_user: User = Depends(get_current_user)):
    feeder = await db.feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    if isinstance(feeder.get('created_at'), str):
        feeder['created_at'] = datetime.fromisoformat(feeder['created_at'])
    return Feeder(**feeder)

@api_router.post("/entries", response_model=DailyEntry)
async def create_entry(entry_data: DailyEntryCreate, current_user: User = Depends(get_current_user)):
    feeder = await db.feeders.find_one({"id": entry_data.feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    
    existing_entry = await db.entries.find_one({"feeder_id": entry_data.feeder_id, "date": entry_data.date}, {"_id": 0})
    if existing_entry:
        raise HTTPException(status_code=400, detail="Entry already exists for this date")
    
    date_obj = datetime.strptime(entry_data.date, "%Y-%m-%d")
    prev_date = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
    prev_entry = await db.entries.find_one({"feeder_id": entry_data.feeder_id, "date": prev_date}, {"_id": 0})
    
    if prev_entry:
        end1_import_initial = prev_entry['end1_import_final']
        end1_export_initial = prev_entry['end1_export_final']
        end2_import_initial = prev_entry['end2_import_final']
        end2_export_initial = prev_entry['end2_export_final']
    else:
        end1_import_initial = 0
        end1_export_initial = 0
        end2_import_initial = 0
        end2_export_initial = 0
    
    end1_import_consumption = (entry_data.end1_import_final - end1_import_initial) * feeder['end1_import_mf']
    end1_export_consumption = (entry_data.end1_export_final - end1_export_initial) * feeder['end1_export_mf']
    end2_import_consumption = (entry_data.end2_import_final - end2_import_initial) * feeder['end2_import_mf']
    end2_export_consumption = (entry_data.end2_export_final - end2_export_initial) * feeder['end2_export_mf']
    
    total_import = end1_import_consumption + end2_import_consumption
    if total_import == 0:
        loss_percent = 0
    else:
        loss_percent = ((end1_import_consumption - end1_export_consumption + end2_import_consumption - end2_export_consumption) / total_import) * 100
    
    entry_obj = DailyEntry(
        feeder_id=entry_data.feeder_id,
        date=entry_data.date,
        end1_import_initial=end1_import_initial,
        end1_import_final=entry_data.end1_import_final,
        end1_export_initial=end1_export_initial,
        end1_export_final=entry_data.end1_export_final,
        end2_import_initial=end2_import_initial,
        end2_import_final=entry_data.end2_import_final,
        end2_export_initial=end2_export_initial,
        end2_export_final=entry_data.end2_export_final,
        end1_import_consumption=end1_import_consumption,
        end1_export_consumption=end1_export_consumption,
        end2_import_consumption=end2_import_consumption,
        end2_export_consumption=end2_export_consumption,
        loss_percent=loss_percent
    )
    
    doc = entry_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.entries.insert_one(doc)
    
    return entry_obj

@api_router.get("/entries", response_model=List[DailyEntry])
async def get_entries(
    feeder_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if feeder_id:
        query['feeder_id'] = feeder_id
    if year and month:
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query['date'] = {"$gte": start_date, "$lt": end_date}
    
    entries = await db.entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    for entry in entries:
        if isinstance(entry.get('created_at'), str):
            entry['created_at'] = datetime.fromisoformat(entry['created_at'])
        if isinstance(entry.get('updated_at'), str):
            entry['updated_at'] = datetime.fromisoformat(entry['updated_at'])
    return entries

@api_router.get("/interruptions/entries/{feeder_id}", response_model=List[InterruptionEntry])
async def get_interruption_entries(
    feeder_id: str,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    query = {"feeder_id": feeder_id}
    if year and month:
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query["date"] = {"$gte": start_date, "$lt": end_date}
    entries = await db.interruption_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    for entry in entries:
        if isinstance(entry.get("created_at"), str):
            entry["created_at"] = datetime.fromisoformat(entry["created_at"])
        if isinstance(entry.get("updated_at"), str):
            entry["updated_at"] = datetime.fromisoformat(entry["updated_at"])
    return [InterruptionEntry(**entry) for entry in entries]


@api_router.post("/interruptions/entries/{feeder_id}", response_model=InterruptionEntry)
async def create_interruption_entry(
    feeder_id: str,
    payload: InterruptionEntryCreate,
    current_user: User = Depends(get_current_user),
):
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    data = payload.data or {}
    date_str = payload.date
    existing = None
    start_time = data.get("start_time")
    if start_time:
        existing = await db.interruption_entries.find_one(
            {"feeder_id": feeder_id, "date": date_str, "data.start_time": start_time},
            {"_id": 0},
        )
    if existing:
        raise HTTPException(status_code=400, detail="Interruption entry already exists for this start time")
    entry_obj = InterruptionEntry(feeder_id=feeder_id, date=date_str, data=data).model_dump()
    created_at = entry_obj.get("created_at")
    updated_at = entry_obj.get("updated_at")
    entry_obj["created_at"] = created_at.isoformat() if isinstance(created_at, datetime) else created_at
    entry_obj["updated_at"] = updated_at.isoformat() if isinstance(updated_at, datetime) else updated_at
    await db.interruption_entries.insert_one(entry_obj)
    if isinstance(entry_obj.get("created_at"), str):
        entry_obj["created_at"] = datetime.fromisoformat(entry_obj["created_at"])
    if isinstance(entry_obj.get("updated_at"), str):
        entry_obj["updated_at"] = datetime.fromisoformat(entry_obj["updated_at"])
    return InterruptionEntry(**entry_obj)


@api_router.put("/interruptions/entries/{entry_id}", response_model=InterruptionEntry)
async def update_interruption_entry(
    entry_id: str,
    update_data: InterruptionEntryUpdate,
    current_user: User = Depends(get_current_user),
):
    entry = await db.interruption_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Interruption entry not found")
    update_dict = update_data.model_dump(exclude_unset=True)
    if "date" in update_dict and update_dict["date"] is not None:
        entry["date"] = update_dict["date"]
    if "data" in update_dict and update_dict["data"] is not None:
        current_data = entry.get("data") or {}
        current_data.update(update_dict["data"])
        entry["data"] = current_data
    entry["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.interruption_entries.update_one({"id": entry_id}, {"$set": entry})
    if isinstance(entry.get("created_at"), str):
        entry["created_at"] = datetime.fromisoformat(entry["created_at"])
    if isinstance(entry.get("updated_at"), str):
        entry["updated_at"] = datetime.fromisoformat(entry["updated_at"])
    return InterruptionEntry(**entry)


@api_router.delete("/interruptions/entries/{entry_id}")
async def delete_interruption_entry(
    entry_id: str,
    current_user: User = Depends(get_current_user),
):
    result = await db.interruption_entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Interruption entry not found")
    return {"message": "Interruption entry deleted successfully"}

def _parse_whatsapp_messages(content: str):
    lines = content.splitlines()
    messages = []
    current = None
    pattern1 = re.compile(r'^(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4}),?\s+(\d{1,2}):(\d{2})(?::\d{2})?(?:\s*([AP]M|[ap]m))?\s*-\s*(.*)$')
    pattern2 = re.compile(r'^\[(\d{1,2})[\/\-.](\d{1,2})[\/\-.](\d{2,4}),\s+(\d{1,2}):(\d{2})(?::\d{2})?(?:\s*([AP]M|[ap]m))?\]\s*(.*)$')
    for line in lines:
        m = pattern1.match(line) or pattern2.match(line)
        if m:
            day = int(m.group(1))
            month = int(m.group(2))
            year = int(m.group(3))
            if year < 100:
                year += 2000
            hour = int(m.group(4))
            minute = int(m.group(5))
            ampm = m.group(6)
            if ampm:
                up = ampm.upper()
                if up == "PM" and hour != 12:
                    hour += 12
                if up == "AM" and hour == 12:
                    hour = 0
            try:
                ts = datetime(year, month, day, hour, minute)
            except ValueError:
                continue
            text = m.group(7).strip()
            text = text.replace("*", "")
            sender_split = text.split(":", 1)
            if len(sender_split) == 2:
                text = sender_split[1].strip()
            current = {"timestamp": ts, "text": text}
            messages.append(current)
        else:
            if current is not None:
                extra = line.strip()
                if extra:
                    extra = extra.replace("*", "")
                    current["text"] += " " + extra
    return messages

def _classify_interruption_message(text: str):
    s = text.lower()
    ignore_keywords = [
        "image omitted",
        "video omitted",
        "excel",
        ".xlsx",
        "data upload",
        "data uploaded",
        "material received",
    ]
    s_clean = s
    for k in ignore_keywords:
        s_clean = s_clean.replace(k, "")
    if not s_clean.strip():
        return None
    if "lc status:" in s_clean:
        return None
    if "stood ok" in s_clean or "line stood ok" in s_clean:
        return "restore"
    if "a/r success" in s_clean:
        return "outage"
    outage_keywords = [
        "tripped",
        "trip",
        "outage",
        "interruption",
        "failed",
        "shutdown",
        "shut down",
        "hand tripped",
        "hand-tripped",
        "protection operated",
        "busbar protection optd",
        "breakdown declared",
        "blast occurred",
        "a/r kept in off condition",
        "under breakdown condition",
        "lc issued",
        "lc applied",
        "nbfc issued",
        "under lc",
        "line clear issued",
        "line clear for",
        "taken out",
        "taken for shutdown",
        "shutdown type: planned",
        "shutdown type: emergency",
    ]
    restore_keywords = [
        "charged",
        "charged at",
        "normalized",
        "normalised",
        "revived",
        "restored",
        "back",
        "resumed",
        "energised",
        "energized",
        "nbfc returned",
        "lc returned",
        "line clear returned",
        "taken into service",
        "taken in to service",
        "put in to service",
        "put into service",
        "breaker taken into service",
        "line stood ok",
        "stood ok",
    ]
    has_outage = any(k in s_clean for k in outage_keywords)
    has_restore = any(k in s_clean for k in restore_keywords)
    if has_outage and not has_restore:
        return "outage"
    if has_restore and not has_outage:
        return "restore"
    return None


def _extract_time_from_text(text: str, base_dt: datetime) -> datetime:
    s = text.lower()
    m_dt = re.search(
        r"(\d{1,2})-(\d{1,2})-(\d{2,4}).*?(\d{1,2})[:\.](\d{2})\s*(am|pm|hrs|hr|hours)?",
        s,
    )
    if m_dt:
        day = int(m_dt.group(1))
        month = int(m_dt.group(2))
        year_raw = m_dt.group(3)
        if len(year_raw) == 2:
            year = 2000 + int(year_raw)
        else:
            year = int(year_raw)
        hour = int(m_dt.group(4))
        minute = int(m_dt.group(5))
        suffix = (m_dt.group(6) or "").lower()
        if suffix in ["am", "pm"]:
            if suffix == "pm" and hour != 12:
                hour += 12
            if suffix == "am" and hour == 12:
                hour = 0
        try:
            return base_dt.replace(
                year=year, month=month, day=day, hour=hour, minute=minute, second=0, microsecond=0
            )
        except ValueError:
            pass
    matches = list(re.finditer(r"(\d{1,2})[:\.](\d{2})\s*(am|pm|hrs|hr|hours)?", s))
    if not matches:
        return base_dt
    with_suffix = [m for m in matches if (m.group(3) or "").strip()]
    m = with_suffix[-1] if with_suffix else matches[-1]
    hour = int(m.group(1))
    minute = int(m.group(2))
    suffix = m.group(3) or ""
    suffix = suffix.lower()
    if suffix in ["am", "pm"]:
        if suffix == "pm" and hour != 12:
            hour += 12
        if suffix == "am" and hour == 12:
            hour = 0
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return base_dt
    return base_dt.replace(hour=hour, minute=minute, second=0, microsecond=0)


def _split_cause_and_relay(text: str) -> Tuple[str, str]:
    base = text or ""
    lower = base.lower()
    if "bus reactor" in lower and "hand tripped" in lower:
        idx_bus = lower.find("125 mvar bus reactor")
        if idx_bus == -1:
            idx_bus = lower.find("bus reactor")
        if idx_bus != -1:
            base = base[idx_bus:]
            lower = base.lower()
    if "a/r success" in lower and "with following indications" in lower:
        idx_follow = lower.find("with following indications")
        relay = base[idx_follow + len("with following indications") :].strip(" -,:")
        cause_segment = base[:idx_follow]
        idx_ar = lower.find("a/r success")
        if idx_ar != -1:
            cause_segment = base[idx_ar:idx_follow]
        cause = re.sub(
            r"at\s+\d{1,2}[:\.]\d{2}\s*(?:hrs?|hours?)?",
            "",
            cause_segment,
            flags=re.IGNORECASE,
        )
        cause = cause.strip(" .,-:")
        return cause, relay
    if "lc issued" in lower:
        idx_lc = lower.find("lc issued")
        segment = base[idx_lc:]
        lower_seg = segment.lower()
        idx_for = lower_seg.find(" for ")
        if idx_for != -1:
            cause_segment = segment[:idx_for]
            relay = segment[idx_for:].lstrip(" -,:.")
        else:
            cause_segment = segment
            relay = ""
        cause = re.sub(
            r"\bat\s+\d{1,2}[:\.]\d{2}\s*(?:hrs?|hours?)?\b",
            "",
            cause_segment,
            flags=re.IGNORECASE,
        )
        cause = cause.strip(" .,-:")
        return cause, relay.strip()
    idx = lower.find(" for ")
    if idx != -1:
        cause_segment = base[:idx]
        cause = re.sub(
            r"\bat\s+\d{1,2}[:\.]\d{2}\s*(?:hrs?|hours?)?\b",
            "",
            cause_segment,
            flags=re.IGNORECASE,
        )
        cause = cause.strip(" .,-:")
        relay = base[idx:].lstrip(" -,:.")
        return cause, relay
    return base, ""


def _extract_interruption_metadata(
    outage_text: str, restore_text: Optional[str]
) -> Dict[str, Any]:
    base = outage_text or ""
    restore = restore_text or ""
    cause, relay = _split_cause_and_relay(base)
    lower_both = (base + " " + restore).lower()

    m_bay = re.search(r"(\d+-\d+)\s*bay\s+under\s+lc", base, flags=re.IGNORECASE)
    if m_bay:
        bay_id = m_bay.group(1)
        cause = f"LC Issued on {bay_id} Bay"
        if "4-14-52" in lower_both and "replacement" in lower_both:
            relay = "for replacement of B ph limb of breaker 4-14-52"
        elif "replacement" in lower_both and not relay:
            relay = f"for replacement of breaker {bay_id}-52"

    breakdown = "YES" if (
        "breakdown declared" in lower_both
        or "under breakdown condition" in lower_both
    ) else "NO"

    fault_identified = ""
    fault_location = ""

    m = re.search(r"(\d+(?:\.\d+)?)\s*km", base, flags=re.IGNORECASE)
    if not m:
        m = re.search(
            r"distance\s*=?\s*(\d+(?:\.\d+)?)\s*km", base, flags=re.IGNORECASE
        )
    if m:
        km_val = m.group(1)
        fault_location = f"{km_val}km"

    fault_keywords = [
        "fault identified",
        "flashover",
        "flash over",
        "conductor",
        "insulator",
        "oil leakage",
        "pole damage",
    ]
    if any(k in lower_both for k in fault_keywords):
        for sentence in re.split(r"[\.!?]+", base + " " + restore):
            if any(k in sentence.lower() for k in fault_keywords):
                fault_identified = sentence.strip()
                break

    if not fault_identified:
        fault_identified = "-"
    if not fault_location:
        fault_location = "-"

    remark_phrases = [
        ("charged as per ld instructions", "Charged as per LD instructions"),
        ("informed to ld", "Informed to LD"),
        ("busbar charged through tie", "Busbar charged through tie"),
        ("pir replacement done", "PIR replacement done"),
        ("oil leakage arrested", "Oil leakage arrested"),
        ("line stood ok", "Line stood OK"),
        ("stood ok", "Stood OK"),
        ("taken into service", "Taken into service"),
        ("taken in to service", "Taken into service"),
        ("put in to service", "Put into service"),
        ("put into service", "Put into service"),
    ]
    remarks_parts: List[str] = []
    for needle, label in remark_phrases:
        if needle in lower_both and label not in remarks_parts:
            remarks_parts.append(label)
    remarks = ". ".join(remarks_parts) if remarks_parts else ""
    action_taken = remarks

    return {
        "cause_of_interruption": cause,
        "relay_indications_lc_work": relay,
        "breakdown_declared": breakdown,
        "fault_identified_during_patrolling": fault_identified,
        "fault_location": fault_location,
        "remarks": remarks,
        "action_taken": action_taken,
    }

def _build_feeder_aliases(feeder_name: str):
    name_lower = feeder_name.lower()
    aliases = set()
    aliases.add(name_lower)
    tokens = name_lower.split()
    if tokens:
        first = tokens[0]
        if first not in ("400kv", "220kv"):
            aliases.add(first)
            if "-" in first:
                no_dash_space = first.replace("-", " ")
                no_dash_compact = first.replace("-", "")
                aliases.add(no_dash_space)
                aliases.add(no_dash_compact)
    if len(tokens) > 1:
        aliases.add(" ".join(tokens[1:]))
    clean = name_lower.replace("(", " ").replace(")", " ")
    parts = clean.split()
    if parts:
        aliases.add(" ".join(parts))
    suffix = None
    if "-1" in name_lower:
        suffix = "-1"
    elif "-2" in name_lower:
        suffix = "-2"
    if "nizamabad" in name_lower:
        aliases.add(name_lower.replace("nizamabad", "nizambad"))
        if suffix:
            aliases.add(f"nizambad{suffix}")
    if "maheshwaram" in name_lower:
        aliases.add(name_lower.replace("maheshwaram", "maheswaram"))
        if suffix:
            aliases.add(f"maheswaram{suffix}")
    if "sadasivapet" in name_lower:
        aliases.add(name_lower.replace("sadasivapet", "sadashivapet"))
        if suffix:
            aliases.add(f"sadashivapet{suffix}")
    if "thandur" in name_lower:
        aliases.add(name_lower.replace("thandur", "tandur"))
        aliases.add("tandur")
    if "tandur" in name_lower:
        aliases.add(name_lower.replace("tandur", "thandur"))
        aliases.add("thandur")
    return list(aliases)


async def _build_interruptions_from_chat_for_all_feeders(
    content: str,
    year: Optional[int],
    month: Optional[int],
):
    feeders = await db.max_min_feeders.find(
        {"type": {"$in": ["feeder_400kv", "feeder_220kv", "ict_feeder", "reactor_feeder", "bay_feeder"]}},
        {"_id": 0},
    ).to_list(1000)
    messages = _parse_whatsapp_messages(content)
    aliases_map = {}
    bus_reactor_feeder_id = None
    for f in feeders:
        aliases_map[f["id"]] = _build_feeder_aliases(f["name"])
        name_lower = f["name"].lower()
        if "bus reactor" in name_lower and f.get("type") == "reactor_feeder":
            bus_reactor_feeder_id = f["id"]
    events_by_feeder: Dict[str, List[Dict[str, Any]]] = {}
    for msg in messages:
        text = msg["text"]
        if not text:
            continue
        lower = text.lower()
        matched_feeder_id = None
        if bus_reactor_feeder_id and "bus reactor" in lower:
            matched_feeder_id = bus_reactor_feeder_id
        else:
            for feeder in feeders:
                fid = feeder["id"]
                aliases = aliases_map.get(fid) or []
                if any(a in lower for a in aliases):
                    matched_feeder_id = fid
                    break
        if not matched_feeder_id:
            continue
        kind = _classify_interruption_message(text)
        if not kind:
            continue
        ts = _extract_time_from_text(text, msg["timestamp"])
        if matched_feeder_id not in events_by_feeder:
            events_by_feeder[matched_feeder_id] = []
        events_by_feeder[matched_feeder_id].append(
            {"timestamp": ts, "kind": kind, "text": text}
        )
    preview: List[Dict[str, Any]] = []
    for feeder in feeders:
        fid = feeder["id"]
        events = events_by_feeder.get(fid, [])
        if not events:
            continue
        if year:
            events = [e for e in events if e["timestamp"].year == year]
            if not events:
                continue
        events.sort(key=lambda e: e["timestamp"])
        open_outages: List[Dict[str, Any]] = []
        pairs: List[Dict[str, Any]] = []
        for ev in events:
            if ev["kind"] == "outage":
                open_outages.append(ev)
            elif ev["kind"] == "restore":
                candidates = [o for o in open_outages if o["timestamp"] < ev["timestamp"]]
                if not candidates:
                    continue
                restore_date = ev["timestamp"].date()
                same_day = [o for o in candidates if o["timestamp"].date() == restore_date]
                if same_day:
                    outage = same_day[-1]
                else:
                    outage = candidates[-1]
                open_outages.remove(outage)
                start_ts = outage["timestamp"]
                end_ts = ev["timestamp"]
                if month and start_ts.month != month:
                    continue
                if end_ts <= start_ts:
                    continue
                duration_minutes = (end_ts - start_ts).total_seconds() / 60.0
                date_str = start_ts.date().isoformat()
                end_date_str = end_ts.date().isoformat()
                start_time = start_ts.strftime("%H:%M")
                end_time = end_ts.strftime("%H:%M")
                meta = _extract_interruption_metadata(outage["text"], ev["text"])
                pairs.append(
                    {
                        "feeder_id": fid,
                        "feeder_name": feeder["name"],
                        "date": date_str,
                        "end_date": end_date_str,
                        "start_time": start_time,
                        "end_time": end_time,
                        "duration_minutes": round(duration_minutes, 2),
                        "description": outage["text"],
                        **meta,
                    }
                )
        for p in pairs:
            exists = await db.interruption_entries.find_one(
                {
                    "feeder_id": p["feeder_id"],
                    "date": p["date"],
                    "data.start_time": p["start_time"],
                },
                {"_id": 0},
            )
            item = dict(p)
            item["exists"] = bool(exists)
            preview.append(item)
    return preview

@api_router.post("/interruptions/preview-import/{feeder_id}")
async def preview_interruptions_import(
    feeder_id: str,
    year: Optional[int] = None,
    month: Optional[int] = None,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    if feeder.get("type") not in ["feeder_400kv", "feeder_220kv", "ict_feeder", "reactor_feeder", "bay_feeder"]:
        raise HTTPException(status_code=400, detail="Interruptions supported only for 400KV, 220KV, ICT, Reactor and Bay feeders")
    filename = file.filename or ""
    if not filename.lower().endswith(".txt"):
        raise HTTPException(status_code=400, detail="Only WhatsApp .txt exports are supported")
    raw = await file.read()
    try:
        content = raw.decode("utf-8", errors="ignore")
    except Exception:
        raise HTTPException(status_code=400, detail="Failed to decode file as UTF-8 text")
    messages = _parse_whatsapp_messages(content)
    aliases = _build_feeder_aliases(feeder["name"])
    events = []
    for msg in messages:
        text = msg["text"]
        if not text:
            continue
        lower = text.lower()
        if not any(a in lower for a in aliases):
            continue
        kind = _classify_interruption_message(text)
        if not kind:
            continue
        events.append(
            {
                "timestamp": msg["timestamp"],
                "kind": kind,
                "text": text,
            }
        )
    events.sort(key=lambda e: e["timestamp"])
    open_outages = []
    pairs = []
    for ev in events:
        if ev["kind"] == "outage":
            open_outages.append(ev)
        elif ev["kind"] == "restore":
            candidates = [o for o in open_outages if o["timestamp"] < ev["timestamp"]]
            if not candidates:
                continue
            restore_date = ev["timestamp"].date()
            same_day = [o for o in candidates if o["timestamp"].date() == restore_date]
            if same_day:
                outage = same_day[-1]
            else:
                outage = candidates[-1]
            open_outages.remove(outage)
            start_ts = _extract_time_from_text(outage["text"], outage["timestamp"])
            end_ts = _extract_time_from_text(ev["text"], ev["timestamp"])
            if year and month:
                if not ((start_ts.year == year and start_ts.month == month) or (end_ts.year == year and end_ts.month == month)):
                    continue
            if end_ts <= start_ts:
                continue
            duration_minutes = (end_ts - start_ts).total_seconds() / 60.0
            date_str = start_ts.date().isoformat()
            end_date_str = end_ts.date().isoformat()
            start_time = start_ts.strftime("%H:%M")
            end_time = end_ts.strftime("%H:%M")
            meta = _extract_interruption_metadata(outage["text"], ev["text"])
            pairs.append(
                {
                    "date": date_str,
                    "end_date": end_date_str,
                    "start_time": start_time,
                    "end_time": end_time,
                    "duration_minutes": round(duration_minutes, 2),
                    "description": outage["text"],
                    **meta,
                }
            )
    preview = []
    for p in pairs:
        exists = await db.interruption_entries.find_one(
            {
                "feeder_id": feeder_id,
                "date": p["date"],
                "data.start_time": p["start_time"],
            },
            {"_id": 0},
        )
        item = dict(p)
        item["exists"] = bool(exists)
        preview.append(item)
    return preview

@api_router.post("/interruptions/import-entries")
async def import_interruption_entries(payload: InterruptionsImportPayload, current_user: User = Depends(get_current_user)):
    feeder = await db.max_min_feeders.find_one({"id": payload.feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    if feeder.get("type") not in ["feeder_400kv", "feeder_220kv", "ict_feeder", "reactor_feeder", "bay_feeder"]:
        raise HTTPException(status_code=400, detail="Interruptions supported only for 400KV, 220KV, ICT, Reactor and Bay feeders")
    entries_sorted = sorted(payload.entries, key=lambda x: (x.get("date") or "", x.get("start_time") or ""))
    imported = 0
    for e in entries_sorted:
        date_str = e.get("date")
        start_time = e.get("start_time")
        if not date_str or not start_time:
            continue
        existing = await db.interruption_entries.find_one(
            {"feeder_id": payload.feeder_id, "date": date_str, "data.start_time": start_time},
            {"_id": 0},
        )
        if existing:
            continue
        data = {
            "start_time": start_time,
            "end_time": e.get("end_time"),
            "end_date": e.get("end_date") or date_str,
            "duration_minutes": e.get("duration_minutes"),
            "description": e.get("description"),
            "cause_of_interruption": e.get("cause_of_interruption"),
            "relay_indications_lc_work": e.get("relay_indications_lc_work"),
            "breakdown_declared": e.get("breakdown_declared"),
            "fault_identified_during_patrolling": e.get("fault_identified_during_patrolling"),
            "fault_location": e.get("fault_location"),
            "remarks": e.get("remarks"),
            "action_taken": e.get("action_taken"),
        }
        entry_obj = InterruptionEntry(feeder_id=payload.feeder_id, date=date_str, data=data).model_dump()
        created_at = entry_obj.get("created_at")
        updated_at = entry_obj.get("updated_at")
        entry_obj["created_at"] = created_at.isoformat() if isinstance(created_at, datetime) else created_at
        entry_obj["updated_at"] = updated_at.isoformat() if isinstance(updated_at, datetime) else updated_at
        await db.interruption_entries.insert_one(entry_obj)
        imported += 1
    return {"imported": imported}


@api_router.post("/interruptions/preview-import-all")
async def preview_interruptions_import_all(
    year: Optional[int] = None,
    month: Optional[int] = None,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    filename = file.filename or ""
    if not filename.lower().endswith(".txt"):
        raise HTTPException(status_code=400, detail="Only WhatsApp .txt exports are supported")
    raw = await file.read()
    try:
        content = raw.decode("utf-8", errors="ignore")
    except Exception:
        raise HTTPException(status_code=400, detail="Failed to decode file as UTF-8 text")
    preview = await _build_interruptions_from_chat_for_all_feeders(content, year, month)
    return preview


class InterruptionsBulkImportPayload(BaseModel):
    entries: List[Dict[str, Any]]


@api_router.post("/interruptions/import-entries-all")
async def import_interruption_entries_all(
    payload: InterruptionsBulkImportPayload,
    current_user: User = Depends(get_current_user),
):
    entries_sorted = sorted(
        payload.entries,
        key=lambda x: (
            x.get("feeder_id") or "",
            x.get("date") or "",
            x.get("start_time") or "",
        ),
    )
    imported = 0
    for e in entries_sorted:
        feeder_id = e.get("feeder_id")
        if not feeder_id:
            continue
        feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
        if not feeder:
            continue
        if feeder.get("type") not in ["feeder_400kv", "feeder_220kv", "ict_feeder", "reactor_feeder", "bay_feeder"]:
            continue
        date_str = e.get("date")
        start_time = e.get("start_time")
        if not date_str or not start_time:
            continue
        existing = await db.interruption_entries.find_one(
            {"feeder_id": feeder_id, "date": date_str, "data.start_time": start_time},
            {"_id": 0},
        )
        if existing:
            continue
        data = {
            "start_time": start_time,
            "end_time": e.get("end_time"),
            "end_date": e.get("end_date") or date_str,
            "duration_minutes": e.get("duration_minutes"),
            "description": e.get("description"),
            "cause_of_interruption": e.get("cause_of_interruption"),
            "relay_indications_lc_work": e.get("relay_indications_lc_work"),
            "breakdown_declared": e.get("breakdown_declared"),
            "fault_identified_during_patrolling": e.get("fault_identified_during_patrolling"),
            "fault_location": e.get("fault_location"),
            "remarks": e.get("remarks"),
            "action_taken": e.get("action_taken"),
        }
        entry_obj = InterruptionEntry(feeder_id=feeder_id, date=date_str, data=data).model_dump()
        created_at = entry_obj.get("created_at")
        updated_at = entry_obj.get("updated_at")
        entry_obj["created_at"] = created_at.isoformat() if isinstance(created_at, datetime) else created_at
        entry_obj["updated_at"] = updated_at.isoformat() if isinstance(updated_at, datetime) else updated_at
        await db.interruption_entries.insert_one(entry_obj)
        imported += 1
    return {"imported": imported}


@api_router.post("/interruptions/normalize-existing")
async def normalize_existing_interruptions(
    current_user: User = Depends(get_current_user),
):
    cursor = db.interruption_entries.find({}, {"_id": 1, "date": 1, "data": 1})
    updated = 0
    async for doc in cursor:
        data = doc.get("data") or {}
        description = data.get("description") or ""
        changed = False
        if description:
            cause, relay = _split_cause_and_relay(description)
            if cause and data.get("cause_of_interruption") != cause:
                data["cause_of_interruption"] = cause
                changed = True
            if relay and data.get("relay_indications_lc_work") != relay:
                data["relay_indications_lc_work"] = relay
                changed = True
        date_str = doc.get("date")
        start_time = data.get("start_time")
        duration = data.get("duration_minutes")
        if date_str and start_time and duration is not None:
            try:
                year, month, day = [int(p) for p in date_str.split("-")]
                base = datetime(year, month, day)
                parts = str(start_time).split(":")
                if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
                    hour = int(parts[0])
                    minute = int(parts[1])
                    base = base.replace(hour=hour, minute=minute)
                    minutes = float(duration)
                    end_dt = base + timedelta(minutes=minutes)
                    end_date_str = end_dt.date().isoformat()
                    if data.get("end_date") != end_date_str:
                        data["end_date"] = end_date_str
                        changed = True
            except Exception:
                pass
        if changed:
            await db.interruption_entries.update_one(
                {"_id": doc["_id"]},
                {"$set": {"data": data}},
            )
            updated += 1
    return {"updated": updated}

@api_router.post("/max-min/preview-import/{feeder_id}")
async def preview_max_min_import(
    feeder_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
    def tokens(h):
        return h.lower().replace("_", " ").replace("-", " ").replace(".", " ").replace("(", " ").replace(")", " ").split()
    def find_col(required):
        for idx, h in enumerate(headers):
            t = tokens(h)
            if all(r in t for r in required):
                return idx + 1
        return None
    preview = []
    ict_ids = []
    if feeder['type'] == 'bus_station':
        ict_feeders = await db.max_min_feeders.find({"type": "ict_feeder"}, {"id": 1}).to_list(None)
        ict_ids = [f['id'] for f in ict_feeders]

    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=1).value
        if not date_cell:
            continue
        if isinstance(date_cell, datetime):
            date_str = date_cell.date().isoformat()
        else:
            val_str = str(date_cell).strip()
            fmts = ["%Y-%m-%d", "%d-%m-%Y", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"]
            parsed = False
            for fmt in fmts:
                try:
                    date_str = datetime.strptime(val_str, fmt).date().isoformat()
                    parsed = True
                    break
                except:
                    continue
            if not parsed:
                continue
        data = {}
        if feeder['type'] == 'bus_station':
            v400_max_col = find_col({"max", "bus", "voltage", "400kv"}) or find_col({"max", "400kv"})
            v220_max_col = find_col({"max", "bus", "voltage", "220kv"}) or find_col({"max", "220kv"})
            v400_min_col = find_col({"min", "bus", "voltage", "400kv"}) or find_col({"min", "400kv"})
            v220_min_col = find_col({"min", "bus", "voltage", "220kv"}) or find_col({"min", "220kv"})
            max_time_col = find_col({"max", "time"})
            min_time_col = find_col({"min", "time"})
            load_max_mw_col = find_col({"station", "load", "max", "mw"}) or find_col({"max", "mw"})
            load_mvar_col = find_col({"station", "load", "mvar"}) or find_col({"mvar"})
            load_time_col = find_col({"station", "load", "time"})

            def getf(c):
                try:
                    v = ws.cell(row=row, column=c).value if c else None
                    return float(v) if v is not None and str(v).strip() != "" else None
                except:
                    return None
            def gets(c):
                v = ws.cell(row=row, column=c).value if c else None
                return str(v).strip() if v is not None else None
            max_time = gets(max_time_col)
            min_time = gets(min_time_col)
            
            station_time = gets(load_time_col)
            station_max_mw = getf(load_max_mw_col)
            station_mvar = getf(load_mvar_col)
            
            if ict_ids:
                ict_entries = await db.max_min_entries.find(
                    {"feeder_id": {"$in": ict_ids}, "date": date_str},
                    {"data.max": 1}
                ).to_list(None)
                
                times = []
                total_mw = 0.0
                total_mvar = 0.0
                has_ict_data = False
                
                for e in ict_entries:
                    d = e.get('data', {}).get('max', {})
                    t = d.get('time')
                    mw = d.get('mw')
                    mvar = d.get('mvar')
                    
                    if t: times.append(t)
                    if mw is not None: 
                        try: total_mw += float(mw)
                        except: pass
                        has_ict_data = True
                    if mvar is not None:
                        try: total_mvar += float(mvar)
                        except: pass
                
                if has_ict_data:
                    station_max_mw = total_mw
                    station_mvar = total_mvar
                
                times = [t for t in times if t]
                if times:
                    from collections import Counter
                    station_time = Counter(times).most_common(1)[0][0]

            data = {
                "max_bus_voltage_400kv": {"value": getf(v400_max_col), "time": max_time},
                "max_bus_voltage_220kv": {"value": getf(v220_max_col), "time": max_time},
                "min_bus_voltage_400kv": {"value": getf(v400_min_col), "time": min_time},
                "min_bus_voltage_220kv": {"value": getf(v220_min_col), "time": min_time},
                "station_load": {"max_mw": station_max_mw, "mvar": station_mvar, "time": station_time}
            }
            if all(
                x is None for x in [
                    data["max_bus_voltage_400kv"]["value"],
                    data["max_bus_voltage_220kv"]["value"],
                    data["min_bus_voltage_400kv"]["value"],
                    data["min_bus_voltage_220kv"]["value"],
                    data["station_load"]["max_mw"],
                    data["station_load"]["mvar"]
                ]
            ):
                continue
        elif feeder['type'] == 'ict_feeder':
            max_amps_col = find_col({"max", "amps"})
            max_mw_col = find_col({"max", "mw"})
            max_mvar_col = find_col({"max", "mvar"})
            max_time_col = find_col({"max", "time"})
            min_amps_col = find_col({"min", "amps"})
            min_mw_col = find_col({"min", "mw"})
            min_mvar_col = find_col({"min", "mvar"})
            min_time_col = find_col({"min", "time"})
            avg_amps_col = find_col({"avg", "amps"})
            avg_mw_col = find_col({"avg", "mw"})
            def getf(c):
                try:
                    v = ws.cell(row=row, column=c).value if c else None
                    return float(v) if v is not None and str(v).strip() != "" else None
                except:
                    return None
            def gets(c):
                v = ws.cell(row=row, column=c).value if c else None
                return str(v).strip() if v is not None else None
            
            # Calculate Avg if missing
            max_amps_val = getf(max_amps_col)
            min_amps_val = getf(min_amps_col)
            avg_amps_val = getf(avg_amps_col)
            
            if avg_amps_val is None and max_amps_val is not None and min_amps_val is not None:
                avg_amps_val = (max_amps_val + min_amps_val) / 2
                
            max_mw_val = getf(max_mw_col)
            min_mw_val = getf(min_mw_col)
            avg_mw_val = getf(avg_mw_col)
            
            if avg_mw_val is None and max_mw_val is not None and min_mw_val is not None:
                avg_mw_val = (max_mw_val + min_mw_val) / 2

            data = {
                "max": {"amps": max_amps_val, "mw": max_mw_val, "mvar": getf(max_mvar_col), "time": gets(max_time_col)},
                "min": {"amps": min_amps_val, "mw": min_mw_val, "mvar": getf(min_mvar_col), "time": gets(min_time_col)},
                "avg": {"amps": avg_amps_val, "mw": avg_mw_val}
            }
            if all(x is None for x in [data["max"]["amps"], data["max"]["mw"], data["min"]["amps"], data["min"]["mw"], data["avg"]["amps"], data["avg"]["mw"]]):
                continue
        else:
            max_amps_col = find_col({"max", "amps"})
            max_mw_col = find_col({"max", "mw"})
            max_time_col = find_col({"max", "time"})
            min_amps_col = find_col({"min", "amps"})
            min_mw_col = find_col({"min", "mw"})
            min_time_col = find_col({"min", "time"})
            avg_amps_col = find_col({"avg", "amps"})
            avg_mw_col = find_col({"avg", "mw"})
            def getf(c):
                try:
                    v = ws.cell(row=row, column=c).value if c else None
                    return float(v) if v is not None and str(v).strip() != "" else None
                except:
                    return None
            def gets(c):
                v = ws.cell(row=row, column=c).value if c else None
                return str(v).strip() if v is not None else None
            
            # Calculate Avg if missing
            max_amps_val = getf(max_amps_col)
            min_amps_val = getf(min_amps_col)
            avg_amps_val = getf(avg_amps_col)
            
            if avg_amps_val is None and max_amps_val is not None and min_amps_val is not None:
                avg_amps_val = (max_amps_val + min_amps_val) / 2
                
            max_mw_val = getf(max_mw_col)
            min_mw_val = getf(min_mw_col)
            avg_mw_val = getf(avg_mw_col)
            
            if avg_mw_val is None and max_mw_val is not None and min_mw_val is not None:
                avg_mw_val = (max_mw_val + min_mw_val) / 2

            data = {
                "max": {"amps": max_amps_val, "mw": max_mw_val, "time": gets(max_time_col)},
                "min": {"amps": min_amps_val, "mw": min_mw_val, "time": gets(min_time_col)},
                "avg": {"amps": avg_amps_val, "mw": avg_mw_val}
            }
            if all(x is None for x in [data["max"]["amps"], data["max"]["mw"], data["min"]["amps"], data["min"]["mw"], data["avg"]["amps"], data["avg"]["mw"]]):
                continue
        exists = await db.max_min_entries.find_one({"feeder_id": feeder_id, "date": date_str})
        preview.append({"date": date_str, "data": data, "exists": bool(exists)})
    return preview

class MaxMinImportPayload(BaseModel):
    feeder_id: str
    entries: List[dict]

@api_router.post("/max-min/import-entries")
async def import_max_min_entries(payload: MaxMinImportPayload, current_user: User = Depends(get_current_user)):
    feeder_id = payload.feeder_id
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    entries_sorted = sorted(payload.entries, key=lambda x: x['date'])
    imported = 0
    for e in entries_sorted:
        date_str = e['date']
        existing = await db.max_min_entries.find_one({"feeder_id": feeder_id, "date": date_str})
        if existing:
            await db.max_min_entries.update_one(
                {"_id": existing["_id"]},
                {"$set": {
                    "data": e.get("data", {}),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            imported += 1
            continue
        entry_obj = MaxMinEntry(feeder_id=feeder_id, date=date_str, data=e.get("data", {}))
        doc = entry_obj.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.max_min_entries.insert_one(doc)
        imported += 1
    return {"imported": imported}

@api_router.put("/entries/{entry_id}", response_model=DailyEntry)
async def update_entry(entry_id: str, update_data: DailyEntryUpdate, current_user: User = Depends(get_current_user)):
    entry = await db.entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    feeder = await db.feeders.find_one({"id": entry['feeder_id']}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    for key, value in update_dict.items():
        entry[key] = value
    
    end1_import_consumption = (entry['end1_import_final'] - entry['end1_import_initial']) * feeder['end1_import_mf']
    end1_export_consumption = (entry['end1_export_final'] - entry['end1_export_initial']) * feeder['end1_export_mf']
    end2_import_consumption = (entry['end2_import_final'] - entry['end2_import_initial']) * feeder['end2_import_mf']
    end2_export_consumption = (entry['end2_export_final'] - entry['end2_export_initial']) * feeder['end2_export_mf']
    
    total_import = end1_import_consumption + end2_import_consumption
    if total_import == 0:
        loss_percent = 0
    else:
        loss_percent = ((end1_import_consumption - end1_export_consumption + end2_import_consumption - end2_export_consumption) / total_import) * 100
    
    entry['end1_import_consumption'] = end1_import_consumption
    entry['end1_export_consumption'] = end1_export_consumption
    entry['end2_import_consumption'] = end2_import_consumption
    entry['end2_export_consumption'] = end2_export_consumption
    entry['loss_percent'] = loss_percent
    entry['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.entries.update_one({"id": entry_id}, {"$set": entry})
    
    # Update next day's initial values if exists
    date_obj = datetime.strptime(entry['date'], "%Y-%m-%d")
    next_date = (date_obj + timedelta(days=1)).strftime("%Y-%m-%d")
    next_entry = await db.entries.find_one({"feeder_id": entry['feeder_id'], "date": next_date})
    
    if next_entry:
        next_entry['end1_import_initial'] = entry['end1_import_final']
        next_entry['end1_export_initial'] = entry['end1_export_final']
        next_entry['end2_import_initial'] = entry['end2_import_final']
        next_entry['end2_export_initial'] = entry['end2_export_final']
        
        # Recalculate next entry consumption
        next_entry['end1_import_consumption'] = (next_entry['end1_import_final'] - next_entry['end1_import_initial']) * feeder['end1_import_mf']
        next_entry['end1_export_consumption'] = (next_entry['end1_export_final'] - next_entry['end1_export_initial']) * feeder['end1_export_mf']
        next_entry['end2_import_consumption'] = (next_entry['end2_import_final'] - next_entry['end2_import_initial']) * feeder['end2_import_mf']
        next_entry['end2_export_consumption'] = (next_entry['end2_export_final'] - next_entry['end2_export_initial']) * feeder['end2_export_mf']
        
        total_import = next_entry['end1_import_consumption'] + next_entry['end2_import_consumption']
        if total_import == 0:
            next_entry['loss_percent'] = 0
        else:
            next_entry['loss_percent'] = ((next_entry['end1_import_consumption'] - next_entry['end1_export_consumption'] + next_entry['end2_import_consumption'] - next_entry['end2_export_consumption']) / total_import) * 100
            
        next_entry['updated_at'] = datetime.now(timezone.utc).isoformat()
        await db.entries.replace_one({"_id": next_entry['_id']}, next_entry)
    
    if isinstance(entry.get('created_at'), str):
        entry['created_at'] = datetime.fromisoformat(entry['created_at'])
    if isinstance(entry.get('updated_at'), str):
        entry['updated_at'] = datetime.fromisoformat(entry['updated_at'])
    
    return DailyEntry(**entry)

@api_router.delete("/entries/{entry_id}")
async def delete_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    result = await db.entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted successfully"}

# Helper to parse float safely
def get_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

# Helper to calculate stats for a period
def calculate_period_stats(period_entries, feeder_type):
    stats = {}
    
    if feeder_type == 'bus_station':
        # Initialize stats with defaults
        stats = {
            'max_400kv': -float('inf'), 'max_400kv_date': '-', 'max_400kv_time': '-',
            'min_400kv': float('inf'), 'min_400kv_date': '-', 'min_400kv_time': '-',
            'max_220kv': -float('inf'), 'max_220kv_date': '-', 'max_220kv_time': '-',
            'min_220kv': float('inf'), 'min_220kv_date': '-', 'min_220kv_time': '-',
            'max_load': -float('inf'), 'max_load_date': '-', 'max_load_time': '-', 'max_load_mvar': '-'
        }

        # 1. Find Max/Min Values
        for e in period_entries:
            d = e.get('data') or {}
            
            # 400KV
            v = get_float((d.get('max_bus_voltage_400kv') or {}).get('value'))
            if v is not None and v > stats['max_400kv']: stats['max_400kv'] = v
            v = get_float((d.get('min_bus_voltage_400kv') or {}).get('value'))
            if v is not None and v < stats['min_400kv']: stats['min_400kv'] = v
            
            # 220KV
            v = get_float((d.get('max_bus_voltage_220kv') or {}).get('value'))
            if v is not None and v > stats['max_220kv']: stats['max_220kv'] = v
            v = get_float((d.get('min_bus_voltage_220kv') or {}).get('value'))
            if v is not None and v < stats['min_220kv']: stats['min_220kv'] = v

            # Load (Independent)
            v = get_float((d.get('station_load') or {}).get('max_mw'))
            if v is not None and v > stats['max_load']:
                stats['max_load'] = v
                stats['max_load_date'] = e['date']
                stats['max_load_time'] = (d.get('station_load') or {}).get('time', '-')
                stats['max_load_mvar'] = (d.get('station_load') or {}).get('mvar', '-')

        # 2. Collect Candidates
        cands_max_400 = []
        cands_min_400 = []
        cands_max_220 = []
        cands_min_220 = []

        for e in period_entries:
            d = e.get('data') or {}
            date = e['date']
            
            # Max 400
            v = get_float((d.get('max_bus_voltage_400kv') or {}).get('value'))
            if v == stats['max_400kv'] and v is not None:
                cands_max_400.append({'date': date, 'time': str((d.get('max_bus_voltage_400kv') or {}).get('time', '')).strip()})
            
            # Min 400
            v = get_float((d.get('min_bus_voltage_400kv') or {}).get('value'))
            if v == stats['min_400kv'] and v is not None:
                cands_min_400.append({'date': date, 'time': str((d.get('min_bus_voltage_400kv') or {}).get('time', '')).strip()})
            
            # Max 220
            v = get_float((d.get('max_bus_voltage_220kv') or {}).get('value'))
            if v == stats['max_220kv'] and v is not None:
                cands_max_220.append({'date': date, 'time': str((d.get('max_bus_voltage_220kv') or {}).get('time', '')).strip()})
            
            # Min 220
            v = get_float((d.get('min_bus_voltage_220kv') or {}).get('value'))
            if v == stats['min_220kv'] and v is not None:
                cands_min_220.append({'date': date, 'time': str((d.get('min_bus_voltage_220kv') or {}).get('time', '')).strip()})

        # 3. Find Best Matches (Common Time Priority)
        def find_best(list_a, list_b):
            # Try to find intersection
            for a in list_a:
                for b in list_b:
                    if a['date'] == b['date'] and a['time'] == b['time']:
                        return a, b
            # No intersection, return first available
            res_a = list_a[0] if list_a else {'date': '-', 'time': '-'}
            res_b = list_b[0] if list_b else {'date': '-', 'time': '-'}
            return res_a, res_b

        final_max_400, final_max_220 = find_best(cands_max_400, cands_max_220)
        final_min_400, final_min_220 = find_best(cands_min_400, cands_min_220)

        stats['max_400kv_date'] = final_max_400['date']
        stats['max_400kv_time'] = final_max_400['time']
        stats['max_220kv_date'] = final_max_220['date']
        stats['max_220kv_time'] = final_max_220['time']

        stats['min_400kv_date'] = final_min_400['date']
        stats['min_400kv_time'] = final_min_400['time']
        stats['min_220kv_date'] = final_min_220['date']
        stats['min_220kv_time'] = final_min_220['time']
        
        # Cleanup infinities
        if stats['max_400kv'] == -float('inf'): stats['max_400kv'] = '-'
        if stats['min_400kv'] == float('inf'): stats['min_400kv'] = '-'
        if stats['max_220kv'] == -float('inf'): stats['max_220kv'] = '-'
        if stats['min_220kv'] == float('inf'): stats['min_220kv'] = '-'
        if stats['max_load'] == -float('inf'): stats['max_load'] = '-'
        
    else:
        # Feeder / ICT
        # Initialize stats
        stats['max_amps'] = -float('inf'); stats['max_amps_date'] = '-'; stats['max_amps_time'] = '-'
        stats['min_amps'] = float('inf'); stats['min_amps_date'] = '-'; stats['min_amps_time'] = '-'
        stats['max_mw'] = -float('inf'); stats['max_mw_date'] = '-'; stats['max_mw_time'] = '-'
        stats['min_mw'] = float('inf'); stats['min_mw_date'] = '-'; stats['min_mw_time'] = '-'
        
        total_amps = 0; count_amps = 0
        total_mw = 0; count_mw = 0
        
        max_mw_entry = None
        min_mw_entry = None
        
        for e in period_entries:
            d = e.get('data') or {}
            date = e['date']
            
            # Collect Averages (Amps)
            max_amps = get_float((d.get('max') or {}).get('amps'))
            min_amps = get_float((d.get('min') or {}).get('amps'))
            avg_amps = get_float((d.get('avg') or {}).get('amps'))
            
            # Auto-calc avg if missing
            if avg_amps is None and max_amps is not None and min_amps is not None:
                avg_amps = (max_amps + min_amps) / 2
            
            if avg_amps is not None:
                total_amps += avg_amps
                count_amps += 1
                
            # Collect Averages (MW) and Find Max/Min MW
            max_mw = get_float((d.get('max') or {}).get('mw'))
            min_mw = get_float((d.get('min') or {}).get('mw'))
            avg_mw = get_float((d.get('avg') or {}).get('mw'))
            
            # Auto-calc avg if missing
            if avg_mw is None and max_mw is not None and min_mw is not None:
                avg_mw = (max_mw + min_mw) / 2
                
            if avg_mw is not None:
                total_mw += avg_mw
                count_mw += 1
            
            # Find Max MW Entry
            if max_mw is not None and max_mw > stats['max_mw']:
                stats['max_mw'] = max_mw
                max_mw_entry = e
            
            # Find Min MW Entry
            if min_mw is not None and min_mw < stats['min_mw']:
                stats['min_mw'] = min_mw
                min_mw_entry = e
        
        # Apply Max MW Logic: Get corresponding Amps/Time from Max MW entry
        if max_mw_entry:
            d = max_mw_entry.get('data') or {}
            stats['max_mw_date'] = max_mw_entry['date']
            stats['max_mw_time'] = (d.get('max') or {}).get('time', '-')
            
            stats['max_amps'] = get_float((d.get('max') or {}).get('amps'))
            stats['max_amps_date'] = max_mw_entry['date']
            stats['max_amps_time'] = (d.get('max') or {}).get('time', '-')
        
        # Apply Min MW Logic: Get corresponding Amps/Time from Min MW entry
        if min_mw_entry:
            d = min_mw_entry.get('data') or {}
            stats['min_mw_date'] = min_mw_entry['date']
            stats['min_mw_time'] = (d.get('min') or {}).get('time', '-')
            
            stats['min_amps'] = get_float((d.get('min') or {}).get('amps'))
            stats['min_amps_date'] = min_mw_entry['date']
            stats['min_amps_time'] = (d.get('min') or {}).get('time', '-')
        
        # Cleanup and averages
        if stats['max_amps'] is None or stats['max_amps'] == -float('inf'): stats['max_amps'] = '-'
        if stats['min_amps'] is None or stats['min_amps'] == float('inf'): stats['min_amps'] = '-'
        stats['avg_amps'] = round(total_amps / count_amps, 2) if count_amps > 0 else '-'
        
        if stats['max_mw'] == -float('inf'): stats['max_mw'] = '-'
        if stats['min_mw'] == float('inf'): stats['min_mw'] = '-'
        stats['avg_mw'] = round(total_mw / count_mw, 2) if count_mw > 0 else '-'
        
    return stats

@api_router.get("/export/{feeder_id}/{year}/{month}")
async def export_feeder_data(feeder_id: str, year: int, month: int, current_user: User = Depends(get_current_user)):
    feeder = await db.feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
    
    entries = await db.entries.find(
        {"feeder_id": feeder_id, "date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = feeder['name'][:31]
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = [
        "Date",
        f"{feeder['end1_name']} Import Initial",
        f"{feeder['end1_name']} Import Final",
        f"{feeder['end1_name']} Import MF",
        f"{feeder['end1_name']} Import Consumption",
        f"{feeder['end1_name']} Export Initial",
        f"{feeder['end1_name']} Export Final",
        f"{feeder['end1_name']} Export MF",
        f"{feeder['end1_name']} Export Consumption",
        f"{feeder['end2_name']} Import Initial",
        f"{feeder['end2_name']} Import Final",
        f"{feeder['end2_name']} Import MF",
        f"{feeder['end2_name']} Import Consumption",
        f"{feeder['end2_name']} Export Initial",
        f"{feeder['end2_name']} Export Final",
        f"{feeder['end2_name']} Export MF",
        f"{feeder['end2_name']} Export Consumption",
        "% Loss"
    ]
    
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for entry in entries:
        ws.append([
            format_date(entry['date']),
            entry['end1_import_initial'],
            entry['end1_import_final'],
            feeder['end1_import_mf'],
            entry['end1_import_consumption'],
            entry['end1_export_initial'],
            entry['end1_export_final'],
            feeder['end1_export_mf'],
            entry['end1_export_consumption'],
            entry['end2_import_initial'],
            entry['end2_import_final'],
            feeder['end2_import_mf'],
            entry['end2_import_consumption'],
            entry['end2_export_initial'],
            entry['end2_export_final'],
            feeder['end2_export_mf'],
            entry['end2_export_consumption'],
            entry['loss_percent']
        ])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{feeder['name']}_{year}_{month:02d}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    
@api_router.post("/preview-import/{feeder_id}")
async def preview_import(
    feeder_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    feeder = await db.feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
    def find_col(end_name: str, kind: str):
        target_tokens = end_name.strip().lower().replace("_", " ").replace("-", " ").replace(".", " ").split()
        for idx, h in enumerate(headers):
            h_low = str(h).strip().lower()
            normalized = h_low.replace("_", " ").replace("-", " ").replace(".", " ")
            tokens = normalized.split()
            if all(t in tokens for t in target_tokens) and kind in tokens and "final" in tokens:
                return idx + 1
        return None
    def find_generic_col(end_label: str, kind: str):
        end_label = end_label.strip().lower()
        for idx, h in enumerate(headers):
            h_low = str(h).strip().lower()
            normalized = h_low.replace("_", " ").replace("-", " ").replace(".", " ")
            tokens = normalized.split()
            if end_label in tokens and kind in tokens and "final" in tokens:
                return idx + 1
        return None
    end1_imp_col = find_col(feeder['end1_name'], "import") or find_generic_col("end1", "import")
    end1_exp_col = find_col(feeder['end1_name'], "export") or find_generic_col("end1", "export")
    end2_imp_col = find_col(feeder['end2_name'], "import") or find_generic_col("end2", "import")
    end2_exp_col = find_col(feeder['end2_name'], "export") or find_generic_col("end2", "export")
    preview = []
    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=1).value
        if not date_cell:
            continue
        if isinstance(date_cell, datetime):
            date_str = date_cell.date().isoformat()
        else:
            val_str = str(date_cell).strip()
            formats = [
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%d-%b-%y",
                "%d-%b-%Y",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%d.%m.%Y"
            ]
            parsed = False
            for fmt in formats:
                try:
                    date_str = datetime.strptime(val_str, fmt).date().isoformat()
                    parsed = True
                    break
                except:
                    continue
            if not parsed:
                continue
        def get_float(v):
            try:
                return float(v) if v is not None and str(v).strip() != "" else None
            except:
                return None
        e1i = get_float(ws.cell(row=row, column=end1_imp_col).value) if end1_imp_col else None
        e1e = get_float(ws.cell(row=row, column=end1_exp_col).value) if end1_exp_col else None
        e2i = get_float(ws.cell(row=row, column=end2_imp_col).value) if end2_imp_col else None
        e2e = get_float(ws.cell(row=row, column=end2_exp_col).value) if end2_exp_col else None
        if e1i is None and e1e is None and e2i is None and e2e is None:
            continue
        exists = await db.entries.find_one({"feeder_id": feeder_id, "date": date_str})
        preview.append({
            "date": date_str,
            "end1_import_final": e1i,
            "end1_export_final": e1e,
            "end2_import_final": e2i,
            "end2_export_final": e2e,
            "exists": bool(exists)
        })
    return preview

class LineLossesImportPayload(BaseModel):
    feeder_id: str
    entries: List[dict]

@api_router.post("/import-entries")
async def import_entries(payload: LineLossesImportPayload, current_user: User = Depends(get_current_user)):
    feeder = await db.feeders.find_one({"id": payload.feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    entries_sorted = sorted(payload.entries, key=lambda x: x['date'])
    imported = 0
    for e in entries_sorted:
        date_str = e['date']
        existing = await db.entries.find_one({"feeder_id": payload.feeder_id, "date": date_str})
        if existing:
            continue
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        prev_date = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
        prev_entry = await db.entries.find_one({"feeder_id": payload.feeder_id, "date": prev_date}, {"_id": 0})
        end1_import_initial = prev_entry['end1_import_final'] if prev_entry else 0
        end1_export_initial = prev_entry['end1_export_final'] if prev_entry else 0
        end2_import_initial = prev_entry['end2_import_final'] if prev_entry else 0
        end2_export_initial = prev_entry['end2_export_final'] if prev_entry else 0
        end1_import_final = float(e.get('end1_import_final', 0) or 0)
        end1_export_final = float(e.get('end1_export_final', 0) or 0)
        end2_import_final = float(e.get('end2_import_final', 0) or 0)
        end2_export_final = float(e.get('end2_export_final', 0) or 0)
        end1_import_consumption = (end1_import_final - end1_import_initial) * feeder['end1_import_mf']
        end1_export_consumption = (end1_export_final - end1_export_initial) * feeder['end1_export_mf']
        end2_import_consumption = (end2_import_final - end2_import_initial) * feeder['end2_import_mf']
        end2_export_consumption = (end2_export_final - end2_export_initial) * feeder['end2_export_mf']
        total_import = end1_import_consumption + end2_import_consumption
        loss_percent = 0 if total_import == 0 else ((end1_import_consumption - end1_export_consumption + end2_import_consumption - end2_export_consumption) / total_import) * 100
        entry_obj = DailyEntry(
            feeder_id=payload.feeder_id,
            date=date_str,
            end1_import_initial=end1_import_initial,
            end1_import_final=end1_import_final,
            end1_export_initial=end1_export_initial,
            end1_export_final=end1_export_final,
            end2_import_initial=end2_import_initial,
            end2_import_final=end2_import_final,
            end2_export_initial=end2_export_initial,
            end2_export_final=end2_export_final,
            end1_import_consumption=end1_import_consumption,
            end1_export_consumption=end1_export_consumption,
            end2_import_consumption=end2_import_consumption,
            end2_export_consumption=end2_export_consumption,
            loss_percent=loss_percent
        ).model_dump()
        entry_obj['created_at'] = entry_obj['created_at'].isoformat() if isinstance(entry_obj['created_at'], datetime) else entry_obj['created_at']
        entry_obj['updated_at'] = entry_obj['updated_at'].isoformat() if isinstance(entry_obj['updated_at'], datetime) else entry_obj['updated_at']
        await db.entries.insert_one(entry_obj)
        imported += 1
    return {"imported": imported}

@api_router.post("/energy/preview-import/{sheet_id}")
async def preview_energy_import(
    sheet_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    sheet = await db.energy_sheets.find_one({"id": sheet_id}, {"_id": 0})
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    meters = await db.energy_meters.find({"sheet_id": sheet_id}, {"_id": 0}).to_list(100)
    if not meters:
        raise HTTPException(status_code=400, detail="No meters configured for this sheet")
    
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")
    
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    
    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
    
    def find_final_col(meter_name: str):
        target = meter_name.strip().lower()
        for idx, h in enumerate(headers):
            h_low = h.strip().lower()
            
            # 1. Exact phrase match
            if h_low == f"{target} final":
                return idx + 1
                
            # 2. Token-based match (safer for short names like "IV")
            # Normalize separators to spaces
            normalized = h_low.replace("_", " ").replace("-", " ").replace(".", " ")
            tokens = normalized.split()
            
            if target in tokens and "final" in tokens:
                return idx + 1
                
        return None
    
    meter_final_cols = {}
    for m in meters:
        col = find_final_col(m['name'])
        if col:
            meter_final_cols[m['id']] = col
    
    preview = []
    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=1).value
        if not date_cell:
            continue
        if isinstance(date_cell, datetime):
            date_str = date_cell.date().isoformat()
        else:
            val_str = str(date_cell).strip()
            # Try multiple common date formats
            formats = [
                "%Y-%m-%d",      # 2026-01-01
                "%d-%m-%Y",      # 01-01-2026
                "%d-%b-%y",      # 01-Jan-26
                "%d-%b-%Y",      # 01-Jan-2026
                "%d/%m/%Y",      # 01/01/2026
                "%m/%d/%Y",      # 01/01/2026 (US)
                "%d.%m.%Y"       # 01.01.2026
            ]
            parsed = False
            for fmt in formats:
                try:
                    date_str = datetime.strptime(val_str, fmt).date().isoformat()
                    parsed = True
                    break
                except:
                    continue
            
            if not parsed:
                continue
        
        readings = []
        for meter_id, col in meter_final_cols.items():
            val = ws.cell(row=row, column=col).value
            try:
                final = float(val) if val is not None else None
            except:
                final = None
            if final is not None:
                readings.append({"meter_id": meter_id, "final": final})
        
        if not readings:
            continue
        
        exists = await db.energy_entries.find_one({"sheet_id": sheet_id, "date": date_str})
        preview.append({
            "date": date_str,
            "readings": readings,
            "exists": bool(exists)
        })
    
    return preview

class EnergyImportPayload(BaseModel):
    sheet_id: str
    entries: List[dict]

@api_router.post("/energy/import-entries")
async def import_energy_entries(payload: EnergyImportPayload, current_user: User = Depends(get_current_user)):
    sheet_id = payload.sheet_id
    meters = await db.energy_meters.find({"sheet_id": sheet_id}, {"_id": 0}).to_list(100)
    meter_map = {m['id']: m for m in meters}
    
    entries_sorted = sorted(payload.entries, key=lambda x: x['date'])
    imported = 0
    for e in entries_sorted:
        date_str = e['date']
        existing = await db.energy_entries.find_one({"sheet_id": sheet_id, "date": date_str})
        if existing:
            continue
        
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        prev_date = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
        prev_entry = await db.energy_entries.find_one({"sheet_id": sheet_id, "date": prev_date}, {"_id": 0})
        prev_finals = {}
        if prev_entry:
            for r in prev_entry.get('readings', []):
                prev_finals[r['meter_id']] = r['final']
        
        readings = []
        total_consumption = 0.0
        for r in e.get('readings', []):
            meter = meter_map.get(r['meter_id'])
            if not meter:
                continue
            initial = prev_finals.get(r['meter_id'], 0.0)
            final = float(r['final'])
            consumption = (final - initial) * meter['mf']
            readings.append({
                "meter_id": r['meter_id'],
                "initial": initial,
                "final": final,
                "consumption": consumption
            })
            total_consumption += consumption
        
        entry_doc = EnergyEntry(
            sheet_id=sheet_id,
            date=date_str,
            readings=[EnergyReading(**rr) for rr in readings],
            total_consumption=total_consumption
        ).model_dump()
        
        if isinstance(entry_doc.get('created_at'), datetime):
            entry_doc['created_at'] = entry_doc['created_at'].isoformat()
        if isinstance(entry_doc.get('updated_at'), datetime):
            entry_doc['updated_at'] = entry_doc['updated_at'].isoformat()
        
        await db.energy_entries.insert_one(entry_doc)
        imported += 1
    
    return {"imported": imported}


@api_router.post("/admin/bulk-import/line-losses/excel/{feeder_id}")
async def admin_bulk_import_line_losses_excel(
    feeder_id: str,
    year: int | None = None,
    month: int | None = None,
    overwrite: bool = False,
    file: UploadFile = File(...),
    current_admin: User = Depends(get_current_admin),
):
    feeder = await db.feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]

    def find_col(end_name: str, kind: str):
        target_tokens = end_name.strip().lower().replace("_", " ").replace("-", " ").replace(".", " ").split()
        for idx, h in enumerate(headers):
            h_low = str(h).strip().lower()
            normalized = h_low.replace("_", " ").replace("-", " ").replace(".", " ")
            tokens = normalized.split()
            if all(t in tokens for t in target_tokens) and kind in tokens and "final" in tokens:
                return idx + 1
        return None

    def find_generic_col(end_label: str, kind: str):
        end_label = end_label.strip().lower()
        for idx, h in enumerate(headers):
            h_low = str(h).strip().lower()
            normalized = h_low.replace("_", " ").replace("-", " ").replace(".", " ")
            tokens = normalized.split()
            if end_label in tokens and kind in tokens and "final" in tokens:
                return idx + 1
        return None

    end1_imp_col = find_col(feeder["end1_name"], "import") or find_generic_col("end1", "import")
    end1_exp_col = find_col(feeder["end1_name"], "export") or find_generic_col("end1", "export")
    end2_imp_col = find_col(feeder["end2_name"], "import") or find_generic_col("end2", "import")
    end2_exp_col = find_col(feeder["end2_name"], "export") or find_generic_col("end2", "export")

    preview: list[dict[str, Any]] = []

    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=1).value
        if not date_cell:
            continue
        if isinstance(date_cell, datetime):
            date_str = date_cell.date().isoformat()
        else:
            val_str = str(date_cell).strip()
            formats = [
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%d-%b-%y",
                "%d-%b-%Y",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%d.%m.%Y",
            ]
            parsed = False
            for fmt in formats:
                try:
                    date_str = datetime.strptime(val_str, fmt).date().isoformat()
                    parsed = True
                    break
                except Exception:
                    continue
            if not parsed:
                continue

        def get_float(v):
            try:
                return float(v) if v is not None and str(v).strip() != "" else None
            except Exception:
                return None

        e1i = get_float(ws.cell(row=row, column=end1_imp_col).value) if end1_imp_col else None
        e1e = get_float(ws.cell(row=row, column=end1_exp_col).value) if end1_exp_col else None
        e2i = get_float(ws.cell(row=row, column=end2_imp_col).value) if end2_imp_col else None
        e2e = get_float(ws.cell(row=row, column=end2_exp_col).value) if end2_exp_col else None
        if e1i is None and e1e is None and e2i is None and e2e is None:
            continue
        preview.append(
            {
                "date": date_str,
                "end1_import_final": e1i,
                "end1_export_final": e1e,
                "end2_import_final": e2i,
                "end2_export_final": e2e,
            }
        )

    payload: dict[str, Any] = {
        "feeder_id": feeder_id,
        "entries": preview,
        "overwrite": overwrite,
        "year": year,
        "month": month,
    }
    return await admin_bulk_import_line_losses(payload=payload, current_admin=current_admin)


@api_router.post("/admin/bulk-import/line-losses/excel-preview/{feeder_id}")
async def admin_bulk_import_line_losses_excel_preview(
    feeder_id: str,
    year: int | None = None,
    month: int | None = None,
    file: UploadFile = File(...),
    current_admin: User = Depends(get_current_admin),
):
    feeder = await db.feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]

    def find_col(end_name: str, kind: str):
        target_tokens = end_name.strip().lower().replace("_", " ").replace("-", " ").replace(".", " ").split()
        for idx, h in enumerate(headers):
            h_low = str(h).strip().lower()
            normalized = h_low.replace("_", " ").replace("-", " ").replace(".", " ")
            tokens = normalized.split()
            if all(t in tokens for t in target_tokens) and kind in tokens and "final" in tokens:
                return idx + 1
        return None

    def find_generic_col(end_label: str, kind: str):
        end_label = end_label.strip().lower()
        for idx, h in enumerate(headers):
            h_low = str(h).strip().lower()
            normalized = h_low.replace("_", " ").replace("-", " ").replace(".", " ")
            tokens = normalized.split()
            if end_label in tokens and kind in tokens and "final" in tokens:
                return idx + 1
        return None

    end1_imp_col = find_col(feeder["end1_name"], "import") or find_generic_col("end1", "import")
    end1_exp_col = find_col(feeder["end1_name"], "export") or find_generic_col("end1", "export")
    end2_imp_col = find_col(feeder["end2_name"], "import") or find_generic_col("end2", "import")
    end2_exp_col = find_col(feeder["end2_name"], "export") or find_generic_col("end2", "export")

    preview: list[dict[str, Any]] = []

    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=1).value
        if not date_cell:
            continue
        if isinstance(date_cell, datetime):
            date_str = date_cell.date().isoformat()
        else:
            val_str = str(date_cell).strip()
            formats = [
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%d-%b-%y",
                "%d-%b-%Y",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%d.%m.%Y",
            ]
            parsed = False
            for fmt in formats:
                try:
                    date_str = datetime.strptime(val_str, fmt).date().isoformat()
                    parsed = True
                    break
                except Exception:
                    continue
            if not parsed:
                continue

        def get_float(v):
            try:
                return float(v) if v is not None and str(v).strip() != "" else None
            except Exception:
                return None

        e1i = get_float(ws.cell(row=row, column=end1_imp_col).value) if end1_imp_col else None
        e1e = get_float(ws.cell(row=row, column=end1_exp_col).value) if end1_exp_col else None
        e2i = get_float(ws.cell(row=row, column=end2_imp_col).value) if end2_imp_col else None
        e2e = get_float(ws.cell(row=row, column=end2_exp_col).value) if end2_exp_col else None
        if e1i is None and e1e is None and e2i is None and e2e is None:
            continue

        existing = await db.entries.find_one({"feeder_id": feeder_id, "date": date_str})

        item = {
            "date": date_str,
            "end1_import_final": e1i,
            "end1_export_final": e1e,
            "end2_import_final": e2i,
            "end2_export_final": e2e,
            "exists": bool(existing),
        }

        preview.append(item)

    if year is not None or month is not None:
        year_val = year
        month_val = month
        if year_val is not None and month_val is not None:
            filtered: list[dict[str, Any]] = []
            for item in preview:
                d = item.get("date")
                if not d:
                    continue
                try:
                    dt = datetime.strptime(d, "%Y-%m-%d")
                except Exception:
                    continue
                if dt.year == year_val and dt.month == month_val:
                    filtered.append(item)
            preview = filtered

    return preview

@api_router.post("/admin/bulk-import/energy/excel/{sheet_id}")
async def admin_bulk_import_energy_excel(
    sheet_id: str,
    year: int | None = None,
    month: int | None = None,
    overwrite: bool = False,
    file: UploadFile = File(...),
    current_admin: User = Depends(get_current_admin),
):
    sheet = await db.energy_sheets.find_one({"id": sheet_id}, {"_id": 0})
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    meters = await db.energy_meters.find({"sheet_id": sheet_id}, {"_id": 0}).to_list(100)
    if not meters:
        raise HTTPException(status_code=400, detail="No meters configured for this sheet")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]

    def find_final_col(meter_name: str):
        target = meter_name.strip().lower()
        for idx, h in enumerate(headers):
            h_low = h.strip().lower()
            if h_low == f"{target} final":
                return idx + 1
            normalized = h_low.replace("_", " ").replace("-", " ").replace(".", " ")
            tokens = normalized.split()
            if target in tokens and "final" in tokens:
                return idx + 1
        return None

    meter_final_cols: dict[str, int] = {}
    for m in meters:
        col = find_final_col(m["name"])
        if col:
            meter_final_cols[m["id"]] = col

    preview: list[dict[str, Any]] = []

    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=1).value
        if not date_cell:
            continue
        if isinstance(date_cell, datetime):
            date_str = date_cell.date().isoformat()
        else:
            val_str = str(date_cell).strip()
            formats = [
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%d-%b-%y",
                "%d-%b-%Y",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%d.%m.%Y",
            ]
            parsed = False
            for fmt in formats:
                try:
                    date_str = datetime.strptime(val_str, fmt).date().isoformat()
                    parsed = True
                    break
                except Exception:
                    continue
            if not parsed:
                continue

        readings: list[dict[str, Any]] = []
        for meter_id, col in meter_final_cols.items():
            val = ws.cell(row=row, column=col).value
            try:
                final = float(val) if val is not None else None
            except Exception:
                final = None
            if final is not None:
                readings.append({"meter_id": meter_id, "final": final})

        if not readings:
            continue

        preview.append({"date": date_str, "readings": readings})

    payload: dict[str, Any] = {
        "sheet_id": sheet_id,
        "entries": preview,
        "overwrite": overwrite,
        "year": year,
        "month": month,
    }
    return await admin_bulk_import_energy(payload=payload, current_admin=current_admin)


@api_router.post("/admin/bulk-import/energy/excel-preview/{sheet_id}")
async def admin_bulk_import_energy_excel_preview(
    sheet_id: str,
    year: int | None = None,
    month: int | None = None,
    file: UploadFile = File(...),
    current_admin: User = Depends(get_current_admin),
):
    sheet = await db.energy_sheets.find_one({"id": sheet_id}, {"_id": 0})
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    meters = await db.energy_meters.find({"sheet_id": sheet_id}, {"_id": 0}).to_list(100)
    if not meters:
        raise HTTPException(status_code=400, detail="No meters configured for this sheet")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]

    def find_final_col(meter_name: str):
        target = meter_name.strip().lower()
        for idx, h in enumerate(headers):
            h_low = h.strip().lower()
            if h_low == f"{target} final":
                return idx + 1
            normalized = h_low.replace("_", " ").replace("-", " ").replace(".", " ")
            tokens = normalized.split()
            if target in tokens and "final" in tokens:
                return idx + 1
        return None

    meter_final_cols: dict[str, int] = {}
    for m in meters:
        col = find_final_col(m["name"])
        if col:
            meter_final_cols[m["id"]] = col

    preview: list[dict[str, Any]] = []

    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=1).value
        if not date_cell:
            continue
        if isinstance(date_cell, datetime):
            date_str = date_cell.date().isoformat()
        else:
            val_str = str(date_cell).strip()
            formats = [
                "%Y-%m-%d",
                "%d-%m-%Y",
                "%d-%b-%y",
                "%d-%b-%Y",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%d.%m.%Y",
            ]
            parsed = False
            for fmt in formats:
                try:
                    date_str = datetime.strptime(val_str, fmt).date().isoformat()
                    parsed = True
                    break
                except Exception:
                    continue
            if not parsed:
                continue

        readings: list[dict[str, Any]] = []
        for meter_id, col in meter_final_cols.items():
            val = ws.cell(row=row, column=col).value
            try:
                final = float(val) if val is not None else None
            except Exception:
                final = None
            if final is not None:
                readings.append({"meter_id": meter_id, "final": final})

        if not readings:
            continue

        existing = await db.energy_entries.find_one({"sheet_id": sheet_id, "date": date_str})

        item = {
            "date": date_str,
            "readings": readings,
            "exists": bool(existing),
        }

        preview.append(item)

    if year is not None or month is not None:
        year_val = year
        month_val = month
        if year_val is not None and month_val is not None:
            filtered: list[dict[str, Any]] = []
            for item in preview:
                d = item.get("date")
                if not d:
                    continue
                try:
                    dt = datetime.strptime(d, "%Y-%m-%d")
                except Exception:
                    continue
                if dt.year == year_val and dt.month == month_val:
                    filtered.append(item)
            preview = filtered

    return preview

@api_router.post("/admin/bulk-import/max-min/excel/{feeder_id}")
async def admin_bulk_import_max_min_excel(
    feeder_id: str,
    year: int | None = None,
    month: int | None = None,
    overwrite: bool = False,
    file: UploadFile = File(...),
    current_admin: User = Depends(get_current_admin),
):
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]

    def _tokens(h: str) -> list[str]:
        return (
            h.lower()
            .replace("_", " ")
            .replace("-", " ")
            .replace(".", " ")
            .replace("(", " ")
            .replace(")", " ")
            .split()
        )

    def _find_col(required: set[str]) -> int | None:
        for idx, h in enumerate(headers):
            t = _tokens(h)
            if all(r in t for r in required):
                return idx + 1
        return None

    entries: list[dict[str, Any]] = []
    ict_ids: list[str] = []
    if feeder["type"] == "bus_station":
        ict_feeders = await db.max_min_feeders.find({"type": "ict_feeder"}, {"id": 1}).to_list(None)
        ict_ids = [f["id"] for f in ict_feeders]

    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=1).value
        if not date_cell:
            continue
        if isinstance(date_cell, datetime):
            date_str = date_cell.date().isoformat()
        else:
            val_str = str(date_cell).strip()
            fmts = ["%Y-%m-%d", "%d-%m-%Y", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"]
            parsed = False
            for fmt in fmts:
                try:
                    date_str = datetime.strptime(val_str, fmt).date().isoformat()
                    parsed = True
                    break
                except Exception:
                    continue
            if not parsed:
                continue
        data: dict[str, Any] = {}
        if feeder["type"] == "bus_station":
            v400_max_col = _find_col({"max", "bus", "voltage", "400kv"}) or _find_col({"max", "400kv"})
            v220_max_col = _find_col({"max", "bus", "voltage", "220kv"}) or _find_col({"max", "220kv"})
            v400_min_col = _find_col({"min", "bus", "voltage", "400kv"}) or _find_col({"min", "400kv"})
            v220_min_col = _find_col({"min", "bus", "voltage", "220kv"}) or _find_col({"min", "220kv"})
            max_time_col = _find_col({"max", "time"})
            min_time_col = _find_col({"min", "time"})
            load_max_mw_col = _find_col({"station", "load", "max", "mw"}) or _find_col({"max", "mw"})
            load_mvar_col = _find_col({"station", "load", "mvar"}) or _find_col({"mvar"})
            load_time_col = _find_col({"station", "load", "time"})

            def getf(c: int | None) -> float | None:
                try:
                    v = ws.cell(row=row, column=c).value if c else None
                    return float(v) if v is not None and str(v).strip() != "" else None
                except Exception:
                    return None

            def gets(c: int | None) -> str | None:
                v = ws.cell(row=row, column=c).value if c else None
                return str(v).strip() if v is not None else None

            max_time = gets(max_time_col)
            min_time = gets(min_time_col)

            station_time = gets(load_time_col)
            station_max_mw = getf(load_max_mw_col)
            station_mvar = getf(load_mvar_col)

            if ict_ids:
                ict_entries = await db.max_min_entries.find(
                    {"feeder_id": {"$in": ict_ids}, "date": date_str},
                    {"data.max": 1},
                ).to_list(None)

                times: list[str] = []
                total_mw = 0.0
                total_mvar = 0.0
                has_ict_data = False

                for e in ict_entries:
                    d = e.get("data", {}).get("max", {})
                    t = d.get("time")
                    mw = d.get("mw")
                    mvar = d.get("mvar")
                    if t:
                        times.append(t)
                    if mw is not None:
                        try:
                            total_mw += float(mw)
                        except Exception:
                            pass
                        has_ict_data = True
                    if mvar is not None:
                        try:
                            total_mvar += float(mvar)
                        except Exception:
                            pass

                if has_ict_data:
                    station_max_mw = total_mw
                    station_mvar = total_mvar

                times = [t for t in times if t]
                if times:
                    from collections import Counter

                    station_time = Counter(times).most_common(1)[0][0]

            data = {
                "max_bus_voltage_400kv": {"value": getf(v400_max_col), "time": max_time},
                "max_bus_voltage_220kv": {"value": getf(v220_max_col), "time": max_time},
                "min_bus_voltage_400kv": {"value": getf(v400_min_col), "time": min_time},
                "min_bus_voltage_220kv": {"value": getf(v220_min_col), "time": min_time},
                "station_load": {"max_mw": station_max_mw, "mvar": station_mvar, "time": station_time},
            }
            if all(
                x is None
                for x in [
                    data["max_bus_voltage_400kv"]["value"],
                    data["max_bus_voltage_220kv"]["value"],
                    data["min_bus_voltage_400kv"]["value"],
                    data["min_bus_voltage_220kv"]["value"],
                    data["station_load"]["max_mw"],
                    data["station_load"]["mvar"],
                ]
            ):
                continue
        elif feeder["type"] == "ict_feeder":
            max_amps_col = _find_col({"max", "amps"})
            max_mw_col = _find_col({"max", "mw"})
            max_mvar_col = _find_col({"max", "mvar"})
            max_time_col = _find_col({"max", "time"})
            min_amps_col = _find_col({"min", "amps"})
            min_mw_col = _find_col({"min", "mw"})
            min_mvar_col = _find_col({"min", "mvar"})
            min_time_col = _find_col({"min", "time"})
            avg_amps_col = _find_col({"avg", "amps"})
            avg_mw_col = _find_col({"avg", "mw"})

            def getf(c: int | None) -> float | None:
                try:
                    v = ws.cell(row=row, column=c).value if c else None
                    return float(v) if v is not None and str(v).strip() != "" else None
                except Exception:
                    return None

            def gets(c: int | None) -> str | None:
                v = ws.cell(row=row, column=c).value if c else None
                return str(v).strip() if v is not None else None

            max_amps_val = getf(max_amps_col)
            min_amps_val = getf(min_amps_col)
            avg_amps_val = getf(avg_amps_col)
            if avg_amps_val is None and max_amps_val is not None and min_amps_val is not None:
                avg_amps_val = (max_amps_val + min_amps_val) / 2

            max_mw_val = getf(max_mw_col)
            min_mw_val = getf(min_mw_col)
            avg_mw_val = getf(avg_mw_col)
            if avg_mw_val is None and max_mw_val is not None and min_mw_val is not None:
                avg_mw_val = (max_mw_val + min_mw_val) / 2

            data = {
                "max": {
                    "amps": max_amps_val,
                    "mw": max_mw_val,
                    "mvar": getf(max_mvar_col),
                    "time": gets(max_time_col),
                },
                "min": {
                    "amps": min_amps_val,
                    "mw": min_mw_val,
                    "mvar": getf(min_mvar_col),
                    "time": gets(min_time_col),
                },
                "avg": {"amps": avg_amps_val, "mw": avg_mw_val},
            }
            if all(
                x is None
                for x in [
                    data["max"]["amps"],
                    data["max"]["mw"],
                    data["min"]["amps"],
                    data["min"]["mw"],
                    data["avg"]["amps"],
                    data["avg"]["mw"],
                ]
            ):
                continue
        else:
            max_amps_col = _find_col({"max", "amps"})
            max_mw_col = _find_col({"max", "mw"})
            max_time_col = _find_col({"max", "time"})
            min_amps_col = _find_col({"min", "amps"})
            min_mw_col = _find_col({"min", "mw"})
            min_time_col = _find_col({"min", "time"})
            avg_amps_col = _find_col({"avg", "amps"})
            avg_mw_col = _find_col({"avg", "mw"})

            def getf(c: int | None) -> float | None:
                try:
                    v = ws.cell(row=row, column=c).value if c else None
                    return float(v) if v is not None and str(v).strip() != "" else None
                except Exception:
                    return None

            def gets(c: int | None) -> str | None:
                v = ws.cell(row=row, column=c).value if c else None
                return str(v).strip() if v is not None else None

            max_amps_val = getf(max_amps_col)
            min_amps_val = getf(min_amps_col)
            avg_amps_val = getf(avg_amps_col)
            if avg_amps_val is None and max_amps_val is not None and min_amps_val is not None:
                avg_amps_val = (max_amps_val + min_amps_val) / 2

            max_mw_val = getf(max_mw_col)
            min_mw_val = getf(min_mw_col)
            avg_mw_val = getf(avg_mw_col)
            if avg_mw_val is None and max_mw_val is not None and min_mw_val is not None:
                avg_mw_val = (max_mw_val + min_mw_val) / 2

            data = {
                "max": {"amps": max_amps_val, "mw": max_mw_val, "time": gets(max_time_col)},
                "min": {"amps": min_amps_val, "mw": min_mw_val, "time": gets(min_time_col)},
                "avg": {"amps": avg_amps_val, "mw": avg_mw_val},
            }
            if all(
                x is None
                for x in [
                    data["max"]["amps"],
                    data["max"]["mw"],
                    data["min"]["amps"],
                    data["min"]["mw"],
                    data["avg"]["amps"],
                    data["avg"]["mw"],
                ]
            ):
                continue
        entries.append({"date": date_str, "data": data})

    payload: dict[str, Any] = {
        "feeder_id": feeder_id,
        "entries": entries,
        "overwrite": overwrite,
        "year": year,
        "month": month,
    }
    return await admin_bulk_import_max_min(payload=payload, current_admin=current_admin)


@api_router.post("/admin/bulk-import/max-min/excel-preview/{feeder_id}")
async def admin_bulk_import_max_min_excel_preview(
    feeder_id: str,
    year: int | None = None,
    month: int | None = None,
    file: UploadFile = File(...),
    current_admin: User = Depends(get_current_admin),
):
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]

    def _tokens(h: str) -> list[str]:
        return (
            h.lower()
            .replace("_", " ")
            .replace("-", " ")
            .replace(".", " ")
            .replace("(", " ")
            .replace(")", " ")
            .split()
        )

    def _find_col(required: set[str]) -> int | None:
        for idx, h in enumerate(headers):
            t = _tokens(h)
            if all(r in t for r in required):
                return idx + 1
        return None

    entries: list[dict[str, Any]] = []
    ict_ids: list[str] = []
    if feeder["type"] == "bus_station":
        ict_feeders = await db.max_min_feeders.find({"type": "ict_feeder"}, {"id": 1}).to_list(None)
        ict_ids = [f["id"] for f in ict_feeders]

    for row in range(2, ws.max_row + 1):
        date_cell = ws.cell(row=row, column=1).value
        if not date_cell:
            continue
        if isinstance(date_cell, datetime):
            date_str = date_cell.date().isoformat()
        else:
            val_str = str(date_cell).strip()
            fmts = ["%Y-%m-%d", "%d-%m-%Y", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"]
            parsed = False
            for fmt in fmts:
                try:
                    date_str = datetime.strptime(val_str, fmt).date().isoformat()
                    parsed = True
                    break
                except Exception:
                    continue
            if not parsed:
                continue
        data: dict[str, Any] = {}
        if feeder["type"] == "bus_station":
            v400_max_col = _find_col({"max", "bus", "voltage", "400kv"}) or _find_col({"max", "400kv"})
            v220_max_col = _find_col({"max", "bus", "voltage", "220kv"}) or _find_col({"max", "220kv"})
            v400_min_col = _find_col({"min", "bus", "voltage", "400kv"}) or _find_col({"min", "400kv"})
            v220_min_col = _find_col({"min", "bus", "voltage", "220kv"}) or _find_col({"min", "220kv"})
            max_time_col = _find_col({"max", "time"})
            min_time_col = _find_col({"min", "time"})
            load_max_mw_col = _find_col({"station", "load", "max", "mw"}) or _find_col({"max", "mw"})
            load_mvar_col = _find_col({"station", "load", "mvar"}) or _find_col({"mvar"})
            load_time_col = _find_col({"station", "load", "time"})

            def getf(c: int | None) -> float | None:
                try:
                    v = ws.cell(row=row, column=c).value if c else None
                    return float(v) if v is not None and str(v).strip() != "" else None
                except Exception:
                    return None

            def gets(c: int | None) -> str | None:
                v = ws.cell(row=row, column=c).value if c else None
                return str(v).strip() if v is not None else None

            max_time = gets(max_time_col)
            min_time = gets(min_time_col)

            station_time = gets(load_time_col)
            station_max_mw = getf(load_max_mw_col)
            station_mvar = getf(load_mvar_col)

            if ict_ids:
                ict_entries = await db.max_min_entries.find(
                    {"feeder_id": {"$in": ict_ids}, "date": date_str},
                    {"data.max": 1},
                ).to_list(None)

                times: list[str] = []
                total_mw = 0.0
                total_mvar = 0.0
                has_ict_data = False

                for e in ict_entries:
                    d = e.get("data", {}).get("max", {})
                    t = d.get("time")
                    mw = d.get("mw")
                    mvar = d.get("mvar")
                    if t:
                        times.append(t)
                    if mw is not None:
                        try:
                            total_mw += float(mw)
                        except Exception:
                            pass
                        has_ict_data = True
                    if mvar is not None:
                        try:
                            total_mvar += float(mvar)
                        except Exception:
                            pass

                if has_ict_data:
                    station_max_mw = total_mw
                    station_mvar = total_mvar

                times = [t for t in times if t]
                if times:
                    from collections import Counter

                    station_time = Counter(times).most_common(1)[0][0]

            data = {
                "max_bus_voltage_400kv": {"value": getf(v400_max_col), "time": max_time},
                "max_bus_voltage_220kv": {"value": getf(v220_max_col), "time": max_time},
                "min_bus_voltage_400kv": {"value": getf(v400_min_col), "time": min_time},
                "min_bus_voltage_220kv": {"value": getf(v220_min_col), "time": min_time},
                "station_load": {"max_mw": station_max_mw, "mvar": station_mvar, "time": station_time},
            }
            if all(
                x is None
                for x in [
                    data["max_bus_voltage_400kv"]["value"],
                    data["max_bus_voltage_220kv"]["value"],
                    data["min_bus_voltage_400kv"]["value"],
                    data["min_bus_voltage_220kv"]["value"],
                    data["station_load"]["max_mw"],
                    data["station_load"]["mvar"],
                ]
            ):
                continue
        elif feeder["type"] == "ict_feeder":
            max_amps_col = _find_col({"max", "amps"})
            max_mw_col = _find_col({"max", "mw"})
            max_mvar_col = _find_col({"max", "mvar"})
            max_time_col = _find_col({"max", "time"})
            min_amps_col = _find_col({"min", "amps"})
            min_mw_col = _find_col({"min", "mw"})
            min_mvar_col = _find_col({"min", "mvar"})
            min_time_col = _find_col({"min", "time"})
            avg_amps_col = _find_col({"avg", "amps"})
            avg_mw_col = _find_col({"avg", "mw"})

            def getf(c: int | None) -> float | None:
                try:
                    v = ws.cell(row=row, column=c).value if c else None
                    return float(v) if v is not None and str(v).strip() != "" else None
                except Exception:
                    return None

            def gets(c: int | None) -> str | None:
                v = ws.cell(row=row, column=c).value if c else None
                return str(v).strip() if v is not None else None

            max_amps_val = getf(max_amps_col)
            min_amps_val = getf(min_amps_col)
            avg_amps_val = getf(avg_amps_col)
            if avg_amps_val is None and max_amps_val is not None and min_amps_val is not None:
                avg_amps_val = (max_amps_val + min_amps_val) / 2

            max_mw_val = getf(max_mw_col)
            min_mw_val = getf(min_mw_col)
            avg_mw_val = getf(avg_mw_col)
            if avg_mw_val is None and max_mw_val is not None and min_mw_val is not None:
                avg_mw_val = (max_mw_val + min_mw_val) / 2

            data = {
                "max": {
                    "amps": max_amps_val,
                    "mw": max_mw_val,
                    "mvar": getf(max_mvar_col),
                    "time": gets(max_time_col),
                },
                "min": {
                    "amps": min_amps_val,
                    "mw": min_mw_val,
                    "mvar": getf(min_mvar_col),
                    "time": gets(min_time_col),
                },
                "avg": {"amps": avg_amps_val, "mw": avg_mw_val},
            }
            if all(
                x is None
                for x in [
                    data["max"]["amps"],
                    data["max"]["mw"],
                    data["min"]["amps"],
                    data["min"]["mw"],
                    data["avg"]["amps"],
                    data["avg"]["mw"],
                ]
            ):
                continue
        else:
            max_amps_col = _find_col({"max", "amps"})
            max_mw_col = _find_col({"max", "mw"})
            max_time_col = _find_col({"max", "time"})
            min_amps_col = _find_col({"min", "amps"})
            min_mw_col = _find_col({"min", "mw"})
            min_time_col = _find_col({"min", "time"})
            avg_amps_col = _find_col({"avg", "amps"})
            avg_mw_col = _find_col({"avg", "mw"})

            def getf(c: int | None) -> float | None:
                try:
                    v = ws.cell(row=row, column=c).value if c else None
                    return float(v) if v is not None and str(v).strip() != "" else None
                except Exception:
                    return None

            def gets(c: int | None) -> str | None:
                v = ws.cell(row=row, column=c).value if c else None
                return str(v).strip() if v is not None else None

            max_amps_val = getf(max_amps_col)
            min_amps_val = getf(min_amps_col)
            avg_amps_val = getf(avg_amps_col)
            if avg_amps_val is None and max_amps_val is not None and min_amps_val is not None:
                avg_amps_val = (max_amps_val + min_amps_val) / 2

            max_mw_val = getf(max_mw_col)
            min_mw_val = getf(min_mw_col)
            avg_mw_val = getf(avg_mw_col)
            if avg_mw_val is None and max_mw_val is not None and min_mw_val is not None:
                avg_mw_val = (max_mw_val + min_mw_val) / 2

            data = {
                "max": {"amps": max_amps_val, "mw": max_mw_val, "time": gets(max_time_col)},
                "min": {"amps": min_amps_val, "mw": min_mw_val, "time": gets(min_time_col)},
                "avg": {"amps": avg_amps_val, "mw": avg_mw_val},
            }
            if all(
                x is None
                for x in [
                    data["max"]["amps"],
                    data["max"]["mw"],
                    data["min"]["amps"],
                    data["min"]["mw"],
                    data["avg"]["amps"],
                    data["avg"]["mw"],
                ]
            ):
                continue

        entries.append({"date": date_str, "data": data})

    preview: list[dict[str, Any]] = []
    for e in entries:
        date_str = e.get("date")
        if not date_str:
            continue
        existing = await db.max_min_entries.find_one(
            {"feeder_id": feeder_id, "date": date_str},
            {"_id": 0},
        )
        item = {"date": date_str, "data": e.get("data") or {}, "exists": bool(existing)}
        preview.append(item)

    if year is not None or month is not None:
        year_val = year
        month_val = month
        if year_val is not None and month_val is not None:
            filtered_preview: list[dict[str, Any]] = []
            for item in preview:
                d = item.get("date")
                if not d:
                    continue
                try:
                    dt = datetime.strptime(d, "%Y-%m-%d")
                except Exception:
                    continue
                if dt.year == year_val and dt.month == month_val:
                    filtered_preview.append(item)
            preview = filtered_preview

    return preview


@api_router.post("/admin/bulk-import/interruptions/excel/{feeder_id}")
async def admin_bulk_import_interruptions_excel(
    feeder_id: str,
    year: int | None = None,
    month: int | None = None,
    overwrite: bool = False,
    file: UploadFile = File(...),
    current_admin: User = Depends(get_current_admin),
):
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    if feeder.get("type") not in [
        "feeder_400kv",
        "feeder_220kv",
        "ict_feeder",
        "reactor_feeder",
        "bay_feeder",
    ]:
        raise HTTPException(
            status_code=400,
            detail="Interruptions supported only for 400KV, 220KV, ICT, Reactor and Bay feeders",
        )
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]

    def tokens(h: str) -> list[str]:
        return (
            h.lower()
            .replace("_", " ")
            .replace("-", " ")
            .replace(".", " ")
            .replace("/", " ")
            .split()
        )

    def find_col(*required: str) -> int | None:
        required_set = {r for r in required if r}
        for idx, h in enumerate(headers):
            t = tokens(h)
            if all(r in t for r in required_set):
                return idx + 1
        return None

    date_col = find_col("date")
    from_col = find_col("time", "from") or find_col("from")
    to_col = find_col("time", "to") or find_col("to")
    cause_col = find_col("cause", "interruption") or find_col("cause")
    relay_col = find_col("relay") or find_col("indications") or find_col("lc", "work")
    breakdown_col = find_col("breakdown")
    fault_ident_col = find_col("fault", "identified") or find_col("identified")
    fault_location_col = find_col("fault", "location") or find_col("location")
    remarks_col = find_col("remarks") or find_col("remark")
    action_col = find_col("action", "taken") or find_col("action")

    def parse_date_value(v: Any) -> str | None:
        if not v:
            return None
        if isinstance(v, datetime):
            return v.date().isoformat()
        s = str(v).strip()
        fmts = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"]
        for fmt in fmts:
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except Exception:
                continue
        return None

    def parse_time_value(v: Any) -> str | None:
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        return format_time(s)

    entries: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        if not date_col or not from_col or not to_col:
            break
        date_val = ws.cell(row=row, column=date_col).value
        if not date_val:
            continue
        date_str = parse_date_value(date_val)
        if not date_str:
            continue
        start_raw = ws.cell(row=row, column=from_col).value
        end_raw = ws.cell(row=row, column=to_col).value
        if not start_raw or not end_raw:
            continue
        start_time = parse_time_value(start_raw)
        if not start_time:
            continue
        end_s = str(end_raw).strip()
        if not end_s:
            continue
        end_date_str: str | None
        end_time_part: str
        if " " in end_s:
            date_part, time_part = end_s.rsplit(" ", 1)
            end_date_str = parse_date_value(date_part) or date_str
            end_time_part = time_part
        else:
            end_date_str = date_str
            end_time_part = end_s
        end_time = parse_time_value(end_time_part)
        if not end_time:
            continue
        duration_minutes: float | None = None
        try:
            start_dt = datetime.strptime(f"{date_str} {start_time}", "%Y-%m-%d %H:%M")
            end_dt = datetime.strptime(f"{end_date_str} {end_time}", "%Y-%m-%d %H:%M")
            if end_dt > start_dt:
                duration_minutes = (end_dt - start_dt).total_seconds() / 60.0
        except Exception:
            duration_minutes = None

        cause_val = ws.cell(row=row, column=cause_col).value if cause_col else None
        relay_val = ws.cell(row=row, column=relay_col).value if relay_col else None
        breakdown_val = ws.cell(row=row, column=breakdown_col).value if breakdown_col else None
        fault_ident_val = ws.cell(row=row, column=fault_ident_col).value if fault_ident_col else None
        fault_location_val = ws.cell(row=row, column=fault_location_col).value if fault_location_col else None
        remarks_val = ws.cell(row=row, column=remarks_col).value if remarks_col else None
        action_val = ws.cell(row=row, column=action_col).value if action_col else None

        cause_text = str(cause_val).strip() if cause_val is not None else ""
        data = {
            "start_time": start_time,
            "end_time": end_time,
            "end_date": end_date_str or date_str,
            "duration_minutes": duration_minutes,
            "description": cause_text or (str(remarks_val).strip() if remarks_val is not None else ""),
            "cause_of_interruption": cause_text or None,
            "relay_indications_lc_work": relay_val,
            "breakdown_declared": breakdown_val,
            "fault_identified_during_patrolling": fault_ident_val,
            "fault_location": fault_location_val,
            "remarks": remarks_val,
            "action_taken": action_val,
        }
        entries.append(
            {
                "feeder_id": feeder_id,
                "date": date_str,
                "start_time": start_time,
                "end_time": end_time,
                "end_date": end_date_str or date_str,
                "duration_minutes": duration_minutes,
                "description": data["description"],
                "cause_of_interruption": data["cause_of_interruption"],
                "relay_indications_lc_work": data["relay_indications_lc_work"],
                "breakdown_declared": data["breakdown_declared"],
                "fault_identified_during_patrolling": data["fault_identified_during_patrolling"],
                "fault_location": data["fault_location"],
                "remarks": data["remarks"],
                "action_taken": data["action_taken"],
            }
        )

    payload: dict[str, Any] = {
        "entries": entries,
        "overwrite": overwrite,
        "year": year,
        "month": month,
    }
    return await admin_bulk_import_interruptions(payload=payload, current_admin=current_admin)


@api_router.post("/admin/bulk-import/interruptions/excel-preview/{feeder_id}")
async def admin_bulk_import_interruptions_excel_preview(
    feeder_id: str,
    year: int | None = None,
    month: int | None = None,
    file: UploadFile = File(...),
    current_admin: User = Depends(get_current_admin),
):
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    if feeder.get("type") not in [
        "feeder_400kv",
        "feeder_220kv",
        "ict_feeder",
        "reactor_feeder",
        "bay_feeder",
    ]:
        raise HTTPException(
            status_code=400,
            detail="Interruptions supported only for 400KV, 220KV, ICT, Reactor and Bay feeders",
        )
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(ws.cell(row=1, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]

    def tokens(h: str) -> list[str]:
        return (
            h.lower()
            .replace("_", " ")
            .replace("-", " ")
            .replace(".", " ")
            .replace("/", " ")
            .split()
        )

    def find_col(*required: str) -> int | None:
        required_set = {r for r in required if r}
        for idx, h in enumerate(headers):
            t = tokens(h)
            if all(r in t for r in required_set):
                return idx + 1
        return None

    date_col = find_col("date")
    from_col = find_col("time", "from") or find_col("from")
    to_col = find_col("time", "to") or find_col("to")
    cause_col = find_col("cause", "interruption") or find_col("cause")
    relay_col = find_col("relay") or find_col("indications") or find_col("lc", "work")
    breakdown_col = find_col("breakdown")
    fault_ident_col = find_col("fault", "identified") or find_col("identified")
    fault_location_col = find_col("fault", "location") or find_col("location")
    remarks_col = find_col("remarks") or find_col("remark")
    action_col = find_col("action", "taken") or find_col("action")

    def parse_date_value(v: Any) -> str | None:
        if not v:
            return None
        if isinstance(v, datetime):
            return v.date().isoformat()
        s = str(v).strip()
        fmts = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"]
        for fmt in fmts:
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except Exception:
                continue
        return None

    def parse_time_value(v: Any) -> str | None:
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        return format_time(s)

    entries: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        if not date_col or not from_col or not to_col:
            break
        date_val = ws.cell(row=row, column=date_col).value
        if not date_val:
            continue
        date_str = parse_date_value(date_val)
        if not date_str:
            continue
        start_raw = ws.cell(row=row, column=from_col).value
        end_raw = ws.cell(row=row, column=to_col).value
        if not start_raw or not end_raw:
            continue
        start_time = parse_time_value(start_raw)
        if not start_time:
            continue
        end_s = str(end_raw).strip()
        if not end_s:
            continue
        if " " in end_s:
            date_part, time_part = end_s.rsplit(" ", 1)
            end_date_str = parse_date_value(date_part) or date_str
            end_time_part = time_part
        else:
            end_date_str = date_str
            end_time_part = end_s
        end_time = parse_time_value(end_time_part)
        if not end_time:
            continue
        duration_minutes: float | None = None
        try:
            start_dt = datetime.strptime(f"{date_str} {start_time}", "%Y-%m-%d %H:%M")
            end_dt = datetime.strptime(f"{end_date_str} {end_time}", "%Y-%m-%d %H:%M")
            if end_dt > start_dt:
                duration_minutes = (end_dt - start_dt).total_seconds() / 60.0
        except Exception:
            duration_minutes = None

        cause_val = ws.cell(row=row, column=cause_col).value if cause_col else None
        relay_val = ws.cell(row=row, column=relay_col).value if relay_col else None
        breakdown_val = ws.cell(row=row, column=breakdown_col).value if breakdown_col else None
        fault_ident_val = ws.cell(row=row, column=fault_ident_col).value if fault_ident_col else None
        fault_location_val = ws.cell(row=row, column=fault_location_col).value if fault_location_col else None
        remarks_val = ws.cell(row=row, column=remarks_col).value if remarks_col else None
        action_val = ws.cell(row=row, column=action_col).value if action_col else None

        cause_text = str(cause_val).strip() if cause_val is not None else ""
        entries.append(
            {
                "feeder_id": feeder_id,
                "date": date_str,
                "start_time": start_time,
                "end_time": end_time,
                "end_date": end_date_str or date_str,
                "duration_minutes": duration_minutes,
                "description": cause_text or (str(remarks_val).strip() if remarks_val is not None else ""),
                "cause_of_interruption": cause_text or None,
                "relay_indications_lc_work": relay_val,
                "breakdown_declared": breakdown_val,
                "fault_identified_during_patrolling": fault_ident_val,
                "fault_location": fault_location_val,
                "remarks": remarks_val,
                "action_taken": action_val,
            }
        )

    preview: list[dict[str, Any]] = []
    for e in entries:
        date_str = e.get("date")
        start_time = e.get("start_time")
        if not date_str or not start_time:
            continue
        existing = await db.interruption_entries.find_one(
            {"feeder_id": feeder_id, "date": date_str, "data.start_time": start_time},
            {"_id": 0},
        )
        item = dict(e)
        item["exists"] = bool(existing)
        preview.append(item)

    if year is not None or month is not None:
        year_val = year
        month_val = month
        if year_val is not None and month_val is not None:
            filtered: list[dict[str, Any]] = []
            for item in preview:
                d = item.get("date")
                if not d:
                    continue
                try:
                    dt = datetime.strptime(d, "%Y-%m-%d")
                except Exception:
                    continue
                if dt.year == year_val and dt.month == month_val:
                    filtered.append(item)
            preview = filtered

    return preview

# Max-Min Data Module Endpoints

@api_router.post("/max-min/init")
async def init_max_min_feeders(current_user: User = Depends(get_current_user)):
    feeders_data = [
        {"name": "Bus Voltages & Station Load", "type": "bus_station"},
        {"name": "400KV MAHESHWARAM-1", "type": "feeder_400kv"},
        {"name": "400KV MAHESHWARAM-2", "type": "feeder_400kv"},
        {"name": "400KV NARSAPUR-1", "type": "feeder_400kv"},
        {"name": "400KV NARSAPUR-2", "type": "feeder_400kv"},
        {"name": "400KV KETHIREDDYPALLY-1", "type": "feeder_400kv"},
        {"name": "400KV KETHIREDDYPALLY-2", "type": "feeder_400kv"},
        {"name": "400KV NIZAMABAD-1", "type": "feeder_400kv"},
        {"name": "400KV NIZAMABAD-2", "type": "feeder_400kv"},
        {"name": "220KV PARIGI-1", "type": "feeder_220kv"},
        {"name": "220KV PARIGI-2", "type": "feeder_220kv"},
        {"name": "220KV THANDUR", "type": "feeder_220kv"},
        {"name": "220KV GACHIBOWLI-1", "type": "feeder_220kv"},
        {"name": "220KV GACHIBOWLI-2", "type": "feeder_220kv"},
        {"name": "220KV KETHIREDDYPALLY", "type": "feeder_220kv"},
        {"name": "220KV YEDDUMAILARAM-1", "type": "feeder_220kv"},
        {"name": "220KV YEDDUMAILARAM-2", "type": "feeder_220kv"},
        {"name": "220KV SADASIVAPET-1", "type": "feeder_220kv"},
        {"name": "220KV SADASIVAPET-2", "type": "feeder_220kv"},
        {"name": "ICT-1 (315MVA)", "type": "ict_feeder"},
        {"name": "ICT-2 (315MVA)", "type": "ict_feeder"},
        {"name": "ICT-3 (315MVA)", "type": "ict_feeder"},
        {"name": "ICT-4 (500MVA)", "type": "ict_feeder"},
        {"name": "125MVAR Bus Reactor", "type": "reactor_feeder"},
        {"name": "4-14 Bay", "type": "bay_feeder"},
        {"name": "4-15 Bay", "type": "bay_feeder"},
        {"name": "4-16 Bay", "type": "bay_feeder"},
        {"name": "4-22 Bay", "type": "bay_feeder"},
    ]
    inserted = 0
    for feeder in feeders_data:
        existing = await db.max_min_feeders.find_one({"name": feeder["name"]})
        if existing:
            await db.max_min_feeders.update_one(
                {"_id": existing["_id"]},
                {"$set": {"type": feeder["type"]}},
            )
            continue
        feeder_obj = MaxMinFeeder(**feeder)
        doc = feeder_obj.model_dump()
        doc["created_at"] = doc["created_at"].isoformat()
        await db.max_min_feeders.insert_one(doc)
        inserted += 1
    total = await db.max_min_feeders.count_documents({})
    return {
        "message": "Max-Min feeders initialized or updated successfully",
        "inserted": inserted,
        "count": total,
    }

@api_router.get("/max-min/feeders", response_model=List[MaxMinFeeder])
async def get_max_min_feeders(current_user: User = Depends(get_current_user)):
    feeders = await db.max_min_feeders.find({}, {"_id": 0}).to_list(100)
    for feeder in feeders:
        if isinstance(feeder.get('created_at'), str):
            feeder['created_at'] = datetime.fromisoformat(feeder['created_at'])
    return feeders

@api_router.get("/max-min/entries/{feeder_id}", response_model=List[MaxMinEntry])
async def get_max_min_entries(
    feeder_id: str,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    query = {"feeder_id": feeder_id}
    if year and month:
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query['date'] = {"$gte": start_date, "$lt": end_date}
    
    entries = await db.max_min_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    # If it's the Bus + Station page, we might need to calculate Station Load on the fly if it's missing or outdated?
    # But user said "Auto-calculate from entered daily data".
    # Let's trust the stored data for now, but we might need a way to refresh it.
    
    for entry in entries:
        if isinstance(entry.get('created_at'), str):
            entry['created_at'] = datetime.fromisoformat(entry['created_at'])
        if isinstance(entry.get('updated_at'), str):
            entry['updated_at'] = datetime.fromisoformat(entry['updated_at'])
    return entries

@api_router.post("/max-min/entries", response_model=MaxMinEntry)
async def create_max_min_entry(entry_data: MaxMinEntryCreate, current_user: User = Depends(get_current_user)):
    # Check if entry exists
    existing_entry = await db.max_min_entries.find_one(
        {"feeder_id": entry_data.feeder_id, "date": entry_data.date},
        {"_id": 0}
    )
    
    if existing_entry:
        # Update existing
        update_data = {
            "data": entry_data.data,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.max_min_entries.update_one(
            {"id": existing_entry['id']},
            {"$set": update_data}
        )
        existing_entry['data'] = entry_data.data
        existing_entry['updated_at'] = update_data['updated_at']
        if isinstance(existing_entry.get('created_at'), str):
            existing_entry['created_at'] = datetime.fromisoformat(existing_entry['created_at'])
        if isinstance(existing_entry.get('updated_at'), str):
            existing_entry['updated_at'] = datetime.fromisoformat(existing_entry['updated_at'])
        return MaxMinEntry(**existing_entry)
    else:
        # Create new
        entry_obj = MaxMinEntry(
            feeder_id=entry_data.feeder_id,
            date=entry_data.date,
            data=entry_data.data
        )
        doc = entry_obj.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.max_min_entries.insert_one(doc)
        return entry_obj

@api_router.put("/max-min/entries/{entry_id}", response_model=MaxMinEntry)
async def update_max_min_entry(entry_id: str, entry_data: MaxMinEntryCreate, current_user: User = Depends(get_current_user)):
    existing_entry = await db.max_min_entries.find_one({"id": entry_id}, {"_id": 0})
    if not existing_entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    update_data = {
        "data": entry_data.data,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Also verify feeder_id and date if needed, but usually we trust the ID
    # However, user might change date in the modal, which is tricky.
    # Frontend sends date in body.
    if existing_entry['date'] != entry_data.date:
        # Check for conflict
        conflict = await db.max_min_entries.find_one(
            {"feeder_id": existing_entry['feeder_id'], "date": entry_data.date},
            {"_id": 0}
        )
        if conflict:
             raise HTTPException(status_code=400, detail="Entry for this date already exists")
        update_data['date'] = entry_data.date

    await db.max_min_entries.update_one(
        {"id": entry_id},
        {"$set": update_data}
    )
    
    updated_entry = await db.max_min_entries.find_one({"id": entry_id}, {"_id": 0})
    if isinstance(updated_entry.get('created_at'), str):
        updated_entry['created_at'] = datetime.fromisoformat(updated_entry['created_at'])
    if isinstance(updated_entry.get('updated_at'), str):
        updated_entry['updated_at'] = datetime.fromisoformat(updated_entry['updated_at'])
    return MaxMinEntry(**updated_entry)

@api_router.delete("/max-min/entries/{entry_id}")
async def delete_max_min_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    result = await db.max_min_entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted successfully"}

# ---------------------------------------------------------
# Max-Min Summary Logic Helpers
# ---------------------------------------------------------

DOUBLE_CIRCUIT_PAIRS = [
    {"names": ["400KV MAHESHWARAM-2", "400KV MAHESHWARAM-1"]},
    {"names": ["400KV NARSAPUR-1", "400KV NARSAPUR-2"]},
    {"names": ["400KV KETHIREDDYPALLY-1", "400KV KETHIREDDYPALLY-2"]},
    {"names": ["400KV NIZAMABAD-1", "400KV NIZAMABAD-2"]},
    {"names": ["220KV PARIGI-1", "220KV PARIGI-2"]},
    {"names": ["220KV GACHIBOWLI-1", "220KV GACHIBOWLI-2"]},
    {"names": ["220KV YEDDUMAILARAM-1", "220KV YEDDUMAILARAM-2"]},
    {"names": ["220KV SADASIVAPET-1", "220KV SADASIVAPET-2"]}
]
ICT_GROUP_NAMES = ["ICT-1 (315MVA)", "ICT-2 (315MVA)", "ICT-3 (315MVA)", "ICT-4 (500MVA)"]

def normalize_time(t):
    if not t: return ""
    s = str(t).strip()
    
    # Handle "YYYY-MM-DD HH:MM:SS" or similar
    if ' ' in s:
        s = s.split(' ')[-1]
        
    # Handle HH:MM:SS vs HH:MM
    if ':' in s:
        parts = s.split(':')
        if len(parts) >= 2:
            try:
                return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
            except ValueError:
                pass
    return s

def get_float_safe(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def get_feeder_group_info(feeder_name):
    # Check Double Circuit
    for pair in DOUBLE_CIRCUIT_PAIRS:
        if feeder_name in pair['names']:
            return True, [n for n in pair['names'] if n != feeder_name]
    
    # Check ICT
    if feeder_name in ICT_GROUP_NAMES:
        return True, [n for n in ICT_GROUP_NAMES if n != feeder_name]
        
    return False, []

def determine_leader(p_group_map):
    global_max_mw = -float('inf')
    candidate_leaders = [] # List of {'fid': fid, 'mw': mw, 'entry': entry}

    for fid, entries in p_group_map.items():
        local_max = -float('inf')
        local_entry = None
        for e in entries:
            mw = get_float_safe(e.get('data', {}).get('max', {}).get('mw'))
            if mw is not None and mw > local_max:
                local_max = mw
                local_entry = e
        
        if local_max > global_max_mw:
            global_max_mw = local_max
            candidate_leaders = [{'fid': fid, 'mw': local_max, 'entry': local_entry}]
        elif local_max == global_max_mw and local_max > -float('inf'):
            candidate_leaders.append({'fid': fid, 'mw': local_max, 'entry': local_entry})
    
    leader_id = None
    
    if not candidate_leaders:
         # Fallback if no data
         return list(p_group_map.keys())[0] if p_group_map else None

    if len(candidate_leaders) == 1:
        leader_id = candidate_leaders[0]['fid']
    elif len(candidate_leaders) > 1:
        # Sort candidates by ID for deterministic processing order (crucial for ties)
        candidate_leaders.sort(key=lambda x: str(x['fid']))

        # Tie Breaker: Calculate Sum of MW at the candidate's timestamp
        best_sum = -float('inf')
        best_leader = None
        
        for cand in candidate_leaders:
            c_fid = cand['fid']
            c_entry = cand['entry']
            c_date = c_entry['date']
            c_time = normalize_time(c_entry.get('data', {}).get('max', {}).get('time'))
            
            current_sum = 0
            valid_timestamp = True
            
            # Sum all partners at this timestamp
            for pid, pentries in p_group_map.items():
                p_entry = next((e for e in pentries if e['date'] == c_date), None)
                if not p_entry:
                    valid_timestamp = False
                    break
                
                p_time = normalize_time(p_entry.get('data', {}).get('max', {}).get('time'))
                if p_time != c_time:
                    valid_timestamp = False
                    break
                    
                val = get_float_safe(p_entry.get('data', {}).get('max', {}).get('mw'))
                if val is not None: current_sum += val
            
            if valid_timestamp and current_sum > best_sum:
                best_sum = current_sum
                best_leader = c_fid
        
        if best_leader:
            leader_id = best_leader
        else:
            # Fallback: Sort by ID for determinism
            candidate_leaders.sort(key=lambda x: x['fid'])
            leader_id = candidate_leaders[0]['fid']
            
    return leader_id

def calculate_standard_stats(period_entries, feeder_type):
    stats = {}
    
    if feeder_type == 'bus_station':
        # Initialize stats with defaults
        stats = {
            'max_400kv': -float('inf'), 'max_400kv_date': '-', 'max_400kv_time': '-',
            'min_400kv': float('inf'), 'min_400kv_date': '-', 'min_400kv_time': '-',
            'max_220kv': -float('inf'), 'max_220kv_date': '-', 'max_220kv_time': '-',
            'min_220kv': float('inf'), 'min_220kv_date': '-', 'min_220kv_time': '-',
            'max_load': -float('inf'), 'max_load_date': '-', 'max_load_time': '-', 'max_load_mvar': '-'
        }

        # 1. Find Max/Min Values
        for e in period_entries:
            d = e.get('data', {})
            
            # 400KV
            v = get_float_safe(d.get('max_bus_voltage_400kv', {}).get('value'))
            if v is not None and v > stats['max_400kv']: stats['max_400kv'] = v
            v = get_float_safe(d.get('min_bus_voltage_400kv', {}).get('value'))
            if v is not None and v < stats['min_400kv']: stats['min_400kv'] = v
            
            # 220KV
            v = get_float_safe(d.get('max_bus_voltage_220kv', {}).get('value'))
            if v is not None and v > stats['max_220kv']: stats['max_220kv'] = v
            v = get_float_safe(d.get('min_bus_voltage_220kv', {}).get('value'))
            if v is not None and v < stats['min_220kv']: stats['min_220kv'] = v

            # Load (Independent)
            v = get_float_safe(d.get('station_load', {}).get('max_mw'))
            if v is not None and v > stats['max_load']:
                stats['max_load'] = v
                stats['max_load_date'] = e['date']
                stats['max_load_time'] = d.get('station_load', {}).get('time', '-')
                stats['max_load_mvar'] = d.get('station_load', {}).get('mvar', '-')

        # 2. Collect Candidates & 3. Find Best Matches (Simplified for brevity as per original logic)
        # For simplicity in this refactor, we will just use the FIRST occurrence for Max/Min if simple comparison
        # But original logic had "Common Time Priority" for voltages.
        # Let's keep it simple or copy full logic? 
        # Copying full logic is safer.
        
        cands_max_400 = []
        cands_min_400 = []
        cands_max_220 = []
        cands_min_220 = []

        for e in period_entries:
            d = e.get('data', {})
            date = e['date']
            
            # Max 400
            v = get_float_safe(d.get('max_bus_voltage_400kv', {}).get('value'))
            if v == stats['max_400kv'] and v is not None:
                cands_max_400.append({'date': date, 'time': str(d.get('max_bus_voltage_400kv', {}).get('time', '')).strip()})
        
            # Min 400
            v = get_float_safe(d.get('min_bus_voltage_400kv', {}).get('value'))
            if v == stats['min_400kv'] and v is not None:
                cands_min_400.append({'date': date, 'time': str(d.get('min_bus_voltage_400kv', {}).get('time', '')).strip()})
        
            # Max 220
            v = get_float_safe(d.get('max_bus_voltage_220kv', {}).get('value'))
            if v == stats['max_220kv'] and v is not None:
                cands_max_220.append({'date': date, 'time': str(d.get('max_bus_voltage_220kv', {}).get('time', '')).strip()})
        
            # Min 220
            v = get_float_safe(d.get('min_bus_voltage_220kv', {}).get('value'))
            if v == stats['min_220kv'] and v is not None:
                cands_min_220.append({'date': date, 'time': str(d.get('min_bus_voltage_220kv', {}).get('time', '')).strip()})

        def find_best(list_a, list_b):
            for a in list_a:
                for b in list_b:
                    if a['date'] == b['date'] and a['time'] == b['time']:
                        return a, b
            res_a = list_a[0] if list_a else {'date': '-', 'time': '-'}
            res_b = list_b[0] if list_b else {'date': '-', 'time': '-'}
            return res_a, res_b

        final_max_400, final_max_220 = find_best(cands_max_400, cands_max_220)
        final_min_400, final_min_220 = find_best(cands_min_400, cands_min_220)

        stats['max_400kv_date'] = final_max_400['date']
        stats['max_400kv_time'] = final_max_400['time']
        stats['max_220kv_date'] = final_max_220['date']
        stats['max_220kv_time'] = final_max_220['time']

        stats['min_400kv_date'] = final_min_400['date']
        stats['min_400kv_time'] = final_min_400['time']
        stats['min_220kv_date'] = final_min_220['date']
        stats['min_220kv_time'] = final_min_220['time']
        
        # Cleanup
        if stats['max_400kv'] == -float('inf'): stats['max_400kv'] = '-'
        if stats['min_400kv'] == float('inf'): stats['min_400kv'] = '-'
        if stats['max_220kv'] == -float('inf'): stats['max_220kv'] = '-'
        if stats['min_220kv'] == float('inf'): stats['min_220kv'] = '-'
        if stats['max_load'] == -float('inf'): stats['max_load'] = '-'
        
    else:
        # Feeder / ICT Standard Logic
        stats = {
            'max_amps': -float('inf'), 'max_amps_date': '-', 'max_amps_time': '-',
            'min_amps': float('inf'), 'min_amps_date': '-', 'min_amps_time': '-',
            'max_mw': -float('inf'), 'max_mw_date': '-', 'max_mw_time': '-',
            'min_mw': float('inf'), 'min_mw_date': '-', 'min_mw_time': '-',
            'avg_amps': 0, 'avg_mw': 0
        }
        
        total_amps = 0; count_amps = 0
        total_mw = 0; count_mw = 0
        
        max_mw_entry = None
        min_mw_entry = None
        
        for e in period_entries:
            d = e.get('data', {})
            
            # Avg Amps
            max_a = get_float_safe(d.get('max', {}).get('amps'))
            min_a = get_float_safe(d.get('min', {}).get('amps'))
            avg_a = get_float_safe(d.get('avg', {}).get('amps'))
            if avg_a is None and max_a is not None and min_a is not None: avg_a = (max_a + min_a) / 2
            if avg_a is not None: total_amps += avg_a; count_amps += 1
                
            # Avg MW
            max_m = get_float_safe(d.get('max', {}).get('mw'))
            min_m = get_float_safe(d.get('min', {}).get('mw'))
            avg_m = get_float_safe(d.get('avg', {}).get('mw'))
            if avg_m is None and max_m is not None and min_m is not None: avg_m = (max_m + min_m) / 2
            if avg_m is not None: total_mw += avg_m; count_mw += 1
            
            # Find Max/Min MW Candidates
            if max_m is not None and max_m > stats['max_mw']:
                stats['max_mw'] = max_m
                max_mw_entry = e
            if min_m is not None and min_m < stats['min_mw']:
                stats['min_mw'] = min_m
                min_mw_entry = e
                
        # Fill Stats from Max/Min MW Entries
        if max_mw_entry:
            d = max_mw_entry.get('data', {})
            stats['max_mw_date'] = max_mw_entry['date']
            stats['max_mw_time'] = d.get('max', {}).get('time', '-')
            stats['max_amps'] = get_float_safe(d.get('max', {}).get('amps'))
            stats['max_amps_date'] = max_mw_entry['date']
            stats['max_amps_time'] = d.get('max', {}).get('time', '-')

        if min_mw_entry:
            d = min_mw_entry.get('data', {})
            stats['min_mw_date'] = min_mw_entry['date']
            stats['min_mw_time'] = d.get('min', {}).get('time', '-')
            stats['min_amps'] = get_float_safe(d.get('min', {}).get('amps'))
            stats['min_amps_date'] = min_mw_entry['date']
            stats['min_amps_time'] = d.get('min', {}).get('time', '-')
            
        # Cleanup
        if stats['max_amps'] is None or stats['max_amps'] == -float('inf'): stats['max_amps'] = '-'
        if stats['min_amps'] is None or stats['min_amps'] == float('inf'): stats['min_amps'] = '-'
        stats['avg_amps'] = round(total_amps / count_amps, 2) if count_amps > 0 else '-'
        if stats['max_mw'] == -float('inf'): stats['max_mw'] = '-'
        if stats['min_mw'] == float('inf'): stats['min_mw'] = '-'
        stats['avg_mw'] = round(total_mw / count_mw, 2) if count_mw > 0 else '-'
        
    return stats

def calculate_coincident_stats(leader_entries, current_feeder_id, group_entries_map, feeder_type):
    # This function uses leader_entries to find the "Best Coincident Timestamp"
    # And then returns the stats for the CURRENT feeder at that timestamp.
    
    # Base calculation for Averages (independent of coincidence)
    # We must calculate averages from the CURRENT feeder's entries, not Leader's.
    current_feeder_entries = group_entries_map.get(current_feeder_id, [])
    base_stats = calculate_standard_stats(current_feeder_entries, feeder_type)
    
    # Reset Max/Min fields to be populated by Coincident Logic
    base_stats.update({
        'max_mw': -float('inf'), 'max_mw_date': '-', 'max_mw_time': '-',
        'max_amps': -float('inf'), 'max_amps_date': '-', 'max_amps_time': '-',
        'min_mw': float('inf'), 'min_mw_date': '-', 'min_mw_time': '-',
        'min_amps': float('inf'), 'min_amps_date': '-', 'min_amps_time': '-'
    })

    # 1. Gather Candidates from LEADER
    max_candidates = []
    min_candidates = []
    
    for e in leader_entries:
        d = e.get('data', {})
        mw_max = get_float_safe(d.get('max', {}).get('mw'))
        time_max = normalize_time(d.get('max', {}).get('time'))
        if mw_max is not None:
            max_candidates.append({'entry': e, 'val': mw_max, 'date': e['date'], 'time': time_max})
            
        mw_min = get_float_safe(d.get('min', {}).get('mw'))
        time_min = normalize_time(d.get('min', {}).get('time'))
        if mw_min is not None:
            min_candidates.append({'entry': e, 'val': mw_min, 'date': e['date'], 'time': time_min})
            
    # Sort
    max_candidates.sort(key=lambda x: x['val'], reverse=True)
    min_candidates.sort(key=lambda x: x['val']) # Ascending
    
    # Find Coincident Max Timestamp
    found_max_timestamp = None # {date, time}
    for cand in max_candidates:
        c_date = cand['date']
        c_time = cand['time']
        
        # Check all partners
        all_match = True
        for fid, entries in group_entries_map.items():
            # Find entry for c_date
            partner_entry = next((p for p in entries if p['date'] == c_date), None)
            if not partner_entry:
                all_match = False
                break
            
            # Check time matches
            p_time = normalize_time(partner_entry.get('data', {}).get('max', {}).get('time'))
            if p_time != c_time:
                all_match = False
                break
        
        if all_match:
            found_max_timestamp = {'date': c_date, 'time': c_time}
            break
            
    # Find Coincident Min Timestamp
    found_min_timestamp = None
    for cand in min_candidates:
        c_date = cand['date']
        c_time = cand['time']
        
        all_match = True
        for fid, entries in group_entries_map.items():
            partner_entry = next((p for p in entries if p['date'] == c_date), None)
            if not partner_entry:
                all_match = False
                break
            p_time = normalize_time(partner_entry.get('data', {}).get('min', {}).get('time'))
            if p_time != c_time:
                all_match = False
                break
        
        if all_match:
            found_min_timestamp = {'date': c_date, 'time': c_time}
            break
            
    # Update Stats with CURRENT FEEDER values at found timestamps
    if found_max_timestamp:
        # Find entry for current feeder at this date
        entry = next((e for e in current_feeder_entries if e['date'] == found_max_timestamp['date']), None)
        if entry:
            d = entry.get('data', {})
            base_stats['max_mw'] = get_float_safe(d.get('max', {}).get('mw'))
            base_stats['max_mw_date'] = found_max_timestamp['date']
            base_stats['max_mw_time'] = found_max_timestamp['time'] # Use common time
            
            base_stats['max_amps'] = get_float_safe(d.get('max', {}).get('amps'))
            base_stats['max_amps_date'] = found_max_timestamp['date']
            base_stats['max_amps_time'] = found_max_timestamp['time']
            
            if base_stats['max_mw'] is None: base_stats['max_mw'] = '-'
            if base_stats['max_amps'] is None: base_stats['max_amps'] = '-'
    else:
        base_stats['max_mw'] = '-'
        base_stats['max_mw_date'] = '-'
        base_stats['max_mw_time'] = '-'
        base_stats['max_amps'] = '-'
        base_stats['max_amps_date'] = '-'
        base_stats['max_amps_time'] = '-'

    if found_min_timestamp:
        entry = next((e for e in current_feeder_entries if e['date'] == found_min_timestamp['date']), None)
        if entry:
            d = entry.get('data', {})
            base_stats['min_mw'] = get_float_safe(d.get('min', {}).get('mw'))
            base_stats['min_mw_date'] = found_min_timestamp['date']
            base_stats['min_mw_time'] = found_min_timestamp['time']
            
            base_stats['min_amps'] = get_float_safe(d.get('min', {}).get('amps'))
            base_stats['min_amps_date'] = found_min_timestamp['date']
            base_stats['min_amps_time'] = found_min_timestamp['time']

            if base_stats['min_mw'] is None: base_stats['min_mw'] = '-'
            if base_stats['min_amps'] is None: base_stats['min_amps'] = '-'
    else:
        base_stats['min_mw'] = '-'
        base_stats['min_mw_date'] = '-'
        base_stats['min_mw_time'] = '-'
        base_stats['min_amps'] = '-'
        base_stats['min_amps_date'] = '-'
        base_stats['min_amps_time'] = '-'
        
    return base_stats

@api_router.get("/max-min/summary/{feeder_id}/{year}/{month}")
async def get_monthly_summary(
    feeder_id: str,
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    # 1. Fetch Feeder
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        return JSONResponse(status_code=404, content={"message": "Feeder not found"})

    # 2. Check if Special Logic Applies
    # is_special = False
    # partner_names = []
    
    # # Check Double Circuit
    # for pair in DOUBLE_CIRCUIT_PAIRS:
    #     if feeder['name'] in pair['names']:
    #         is_special = True
    #         partner_names = [n for n in pair['names'] if n != feeder['name']]
    #         break
            
    # # Check ICT
    # if not is_special and feeder['name'] in ICT_GROUP_NAMES:
    #     is_special = True
    #     partner_names = [n for n in ICT_GROUP_NAMES if n != feeder['name']]

    # if not is_special:
    #     return None

    # 3. Fetch Partner Feeders & Entries
    # partner_feeders = await db.max_min_feeders.find({"name": {"$in": partner_names}}).to_list(100)
    # partner_ids = [p['id'] for p in partner_feeders]
    
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"

    # Fetch Target Entries
    target_entries = await db.max_min_entries.find(
        {"feeder_id": feeder_id, "date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0}
    ).to_list(1000)
    
    # Fetch Partner Entries
    # group_entries_map = {}
    # if partner_ids:
    #     p_entries = await db.max_min_entries.find(
    #         {"feeder_id": {"$in": partner_ids}, "date": {"$gte": start_date, "$lt": end_date}},
    #         {"_id": 0}
    #     ).to_list(5000)
        
    #     for e in p_entries:
    #         fid = e['feeder_id']
    #         if fid not in group_entries_map: group_entries_map[fid] = []
    #         group_entries_map[fid].append(e)

    # 4. Calculate Stats for Periods
    periods = [
        {"name": "1st to 15th", "start": 1, "end": 15},
        {"name": "16th to End", "start": 16, "end": 31},
        {"name": "Full Month", "start": 1, "end": 31}
    ]
    
    results = []
    
    for p in periods:
        p_target = [
            e for e in target_entries 
            if p["start"] <= int(e['date'].split('-')[2]) <= p["end"]
        ]
        
        # p_group_map = {}
        # # Add current feeder to map
        # p_group_map[feeder_id] = p_target

        # for fid, entries in group_entries_map.items():
        #     p_group_map[fid] = [
        #         e for e in entries
        #         if p["start"] <= int(e['date'].split('-')[2]) <= p["end"]
        #     ]
        
        # # Determine Leader (Feeder with highest Max MW in this period)
        # leader_id = determine_leader(p_group_map)
        # if not leader_id: leader_id = feeder_id

        # leader_entries = p_group_map[leader_id]
            
        # stats = calculate_coincident_stats(leader_entries, feeder_id, p_group_map, feeder['type'])
        stats = calculate_standard_stats(p_target, feeder['type'])
        stats['name'] = p['name']
        results.append(stats)
        
    return results

@api_router.get("/max-min/export/{feeder_id}/{year}/{month}")
async def export_max_min_data(
    feeder_id: str,
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        import io
        import traceback
        from fastapi.responses import JSONResponse

        feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
        if not feeder:
            raise HTTPException(status_code=404, detail="Feeder not found")
        
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
            
        entries = await db.max_min_entries.find(
            {"feeder_id": feeder_id, "date": {"$gte": start_date, "$lt": end_date}},
            {"_id": 0}
        ).sort("date", 1).to_list(1000)
        
        wb = Workbook()
        ws = wb.active
        ws.title = feeder['name'][:31]
        
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Define headers based on feeder type
        if feeder['type'] == 'bus_station':
            headers = [
                "Date",
                "Max Bus Voltage 400KV", "Max Bus Voltage 220KV", "Max Value Time",
                "Min Bus Voltage 400KV", "Min Bus Voltage 220KV", "Min Value Time",
                "Station Load Max MW", "Station Load MVAR", "Station Load Time"
            ]
        elif feeder['type'] == 'ict_feeder':
            headers = [
                "Date",
                "Max Amps", "Max MW", "Max MVAR", "Max Time",
                "Min Amps", "Min MW", "Min MVAR", "Min Time",
                "Avg Amps", "Avg MW"
            ]
        else: # non-ICT feeders
            headers = [
                "Date",
                "Max Amps", "Max MW", "Max Time",
                "Min Amps", "Min MW", "Min Time",
                "Avg Amps", "Avg MW"
            ]
            
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        for entry in entries:
            d = entry.get('data', {})
            row = []
            row.append(format_date(entry['date']))
            
            if feeder['type'] == 'bus_station':
                # Determine common time for Max Voltages
                max_time = d.get('max_bus_voltage', {}).get('time') or d.get('max_bus_voltage_400kv', {}).get('time') or d.get('max_bus_voltage_220kv', {}).get('time')
                
                # Determine common time for Min Voltages
                min_time = d.get('min_bus_voltage', {}).get('time') or d.get('min_bus_voltage_400kv', {}).get('time') or d.get('min_bus_voltage_220kv', {}).get('time')
                
                row.extend([
                    d.get('max_bus_voltage_400kv', {}).get('value', ''), 
                    d.get('max_bus_voltage_220kv', {}).get('value', ''), 
                    format_time(max_time),
                    
                    d.get('min_bus_voltage_400kv', {}).get('value', ''), 
                    d.get('min_bus_voltage_220kv', {}).get('value', ''), 
                    format_time(min_time),
                    
                    d.get('station_load', {}).get('max_mw', ''), 
                    d.get('station_load', {}).get('mvar', ''), 
                    format_time(d.get('station_load', {}).get('time', ''))
                ])
            elif feeder['type'] == 'ict_feeder':
                row.extend([
                    d.get('max', {}).get('amps', ''), d.get('max', {}).get('mw', ''), d.get('max', {}).get('mvar', ''), format_time(d.get('max', {}).get('time', '')),
                    d.get('min', {}).get('amps', ''), d.get('min', {}).get('mw', ''), d.get('min', {}).get('mvar', ''), format_time(d.get('min', {}).get('time', '')),
                    d.get('avg', {}).get('amps', ''), d.get('avg', {}).get('mw', '')
                ])
            else:
                row.extend([
                    d.get('max', {}).get('amps', ''), d.get('max', {}).get('mw', ''), format_time(d.get('max', {}).get('time', '')),
                    d.get('min', {}).get('amps', ''), d.get('min', {}).get('mw', ''), format_time(d.get('min', {}).get('time', '')),
                    d.get('avg', {}).get('amps', ''), d.get('avg', {}).get('mw', '')
                ])
            ws.append(row)
            
        
        # ---------------------------------------------------------
        # Monthly Summary Section
        # ---------------------------------------------------------
        
        # Define periods
        import calendar
        
        last_day = calendar.monthrange(year, month)[1]
        
        p1_start = f"{year}-{month:02d}-01"
        p1_end = f"{year}-{month:02d}-15"
        
        p2_start = f"{year}-{month:02d}-16"
        p2_end = f"{year}-{month:02d}-{last_day}"
        
        full_start = f"{year}-{month:02d}-01"
        full_end = f"{year}-{month:02d}-{last_day}"
        
        periods = [
            {"name": "1st to 15th", "start": p1_start, "end": p1_end},
            {"name": "16th to End", "start": p2_start, "end": p2_end},
            {"name": "Full Month", "start": full_start, "end": full_end}
        ]
        
        # Helper to parse float safely
        def get_float(val):
            try:
                return float(val)
            except (ValueError, TypeError):
                return None
                
        def format_time_local(t):
            if not t:
                return ''
            return format_time(t)

        # Helper to calculate stats for a period
        def calculate_period_stats(period_entries, feeder_type):
            stats = {}
            
            if feeder_type == 'bus_station':
                # Initialize stats with defaults
                stats = {
                    'max_400kv': -float('inf'), 'max_400kv_date': '-', 'max_400kv_time': '-',
                    'min_400kv': float('inf'), 'min_400kv_date': '-', 'min_400kv_time': '-',
                    'max_220kv': -float('inf'), 'max_220kv_date': '-', 'max_220kv_time': '-',
                    'min_220kv': float('inf'), 'min_220kv_date': '-', 'min_220kv_time': '-',
                    'max_load': -float('inf'), 'max_load_date': '-', 'max_load_time': '-'
                }

                # 1. Find Max/Min Values
                for e in period_entries:
                    d = e.get('data', {})
                    
                    # 400KV
                    v = get_float(d.get('max_bus_voltage_400kv', {}).get('value'))
                    if v is not None and v > stats['max_400kv']: stats['max_400kv'] = v
                    v = get_float(d.get('min_bus_voltage_400kv', {}).get('value'))
                    if v is not None and v < stats['min_400kv']: stats['min_400kv'] = v
                    
                    # 220KV
                    v = get_float(d.get('max_bus_voltage_220kv', {}).get('value'))
                    if v is not None and v > stats['max_220kv']: stats['max_220kv'] = v
                    v = get_float(d.get('min_bus_voltage_220kv', {}).get('value'))
                    if v is not None and v < stats['min_220kv']: stats['min_220kv'] = v

                    # Load (Independent)
                    v = get_float(d.get('station_load', {}).get('max_mw'))
                    if v is not None and v > stats['max_load']:
                        stats['max_load'] = v
                        stats['max_load_date'] = e['date']
                        stats['max_load_time'] = d.get('station_load', {}).get('time', '-')

                # 2. Collect Candidates
                cands_max_400 = []
                cands_min_400 = []
                cands_max_220 = []
                cands_min_220 = []

                for e in period_entries:
                    d = e.get('data', {})
                    date = e['date']
                    
                    # Max 400
                    v = get_float(d.get('max_bus_voltage_400kv', {}).get('value'))
                    if v == stats['max_400kv'] and v is not None:
                        cands_max_400.append({'date': date, 'time': str(d.get('max_bus_voltage_400kv', {}).get('time', '')).strip()})
                    
                    # Min 400
                    v = get_float(d.get('min_bus_voltage_400kv', {}).get('value'))
                    if v == stats['min_400kv'] and v is not None:
                        cands_min_400.append({'date': date, 'time': str(d.get('min_bus_voltage_400kv', {}).get('time', '')).strip()})
                    
                    # Max 220
                    v = get_float(d.get('max_bus_voltage_220kv', {}).get('value'))
                    if v == stats['max_220kv'] and v is not None:
                        cands_max_220.append({'date': date, 'time': str(d.get('max_bus_voltage_220kv', {}).get('time', '')).strip()})
                    
                    # Min 220
                    v = get_float(d.get('min_bus_voltage_220kv', {}).get('value'))
                    if v == stats['min_220kv'] and v is not None:
                        cands_min_220.append({'date': date, 'time': str(d.get('min_bus_voltage_220kv', {}).get('time', '')).strip()})

                # 3. Find Best Matches (Common Time Priority)
                def find_best(list_a, list_b):
                    # Try to find intersection
                    for a in list_a:
                        for b in list_b:
                            if a['date'] == b['date'] and a['time'] == b['time']:
                                return a, b
                    # No intersection, return first available
                    res_a = list_a[0] if list_a else {'date': '-', 'time': '-'}
                    res_b = list_b[0] if list_b else {'date': '-', 'time': '-'}
                    return res_a, res_b

                final_max_400, final_max_220 = find_best(cands_max_400, cands_max_220)
                final_min_400, final_min_220 = find_best(cands_min_400, cands_min_220)

                stats['max_400kv_date'] = final_max_400['date']
                stats['max_400kv_time'] = final_max_400['time']
                stats['max_220kv_date'] = final_max_220['date']
                stats['max_220kv_time'] = final_max_220['time']

                stats['min_400kv_date'] = final_min_400['date']
                stats['min_400kv_time'] = final_min_400['time']
                stats['min_220kv_date'] = final_min_220['date']
                stats['min_220kv_time'] = final_min_220['time']
                
                # Cleanup infinities
                if stats['max_400kv'] == -float('inf'): stats['max_400kv'] = '-'
                if stats['min_400kv'] == float('inf'): stats['min_400kv'] = '-'
                if stats['max_220kv'] == -float('inf'): stats['max_220kv'] = '-'
                if stats['min_220kv'] == float('inf'): stats['min_220kv'] = '-'
                if stats['max_load'] == -float('inf'): stats['max_load'] = '-'
                
            else:
                # Feeder / ICT
                # Initialize stats
                stats['max_amps'] = -float('inf'); stats['max_amps_date'] = '-'; stats['max_amps_time'] = '-'
                stats['min_amps'] = float('inf'); stats['min_amps_date'] = '-'; stats['min_amps_time'] = '-'
                stats['max_mw'] = -float('inf'); stats['max_mw_date'] = '-'; stats['max_mw_time'] = '-'
                stats['min_mw'] = float('inf'); stats['min_mw_date'] = '-'; stats['min_mw_time'] = '-'
                
                total_amps = 0; count_amps = 0
                total_mw = 0; count_mw = 0
                
                max_mw_entry = None
                min_mw_entry = None
                
                for e in period_entries:
                    d = e.get('data', {})
                    date = e['date']
                    
                    # Collect Averages (Amps)
                    max_amps = get_float(d.get('max', {}).get('amps'))
                    min_amps = get_float(d.get('min', {}).get('amps'))
                    avg_amps = get_float(d.get('avg', {}).get('amps'))
                    
                    # Auto-calc avg if missing
                    if avg_amps is None and max_amps is not None and min_amps is not None:
                        avg_amps = (max_amps + min_amps) / 2
                    
                    if avg_amps is not None:
                        total_amps += avg_amps
                        count_amps += 1
                        
                    # Collect Averages (MW) and Find Max/Min MW
                    max_mw = get_float(d.get('max', {}).get('mw'))
                    min_mw = get_float(d.get('min', {}).get('mw'))
                    avg_mw = get_float(d.get('avg', {}).get('mw'))
                    
                    # Auto-calc avg if missing
                    if avg_mw is None and max_mw is not None and min_mw is not None:
                        avg_mw = (max_mw + min_mw) / 2
                        
                    if avg_mw is not None:
                        total_mw += avg_mw
                        count_mw += 1
                    
                    # Find Max MW Entry
                    if max_mw is not None and max_mw > stats['max_mw']:
                        stats['max_mw'] = max_mw
                        max_mw_entry = e
                    
                    # Find Min MW Entry
                    if min_mw is not None and min_mw < stats['min_mw']:
                        stats['min_mw'] = min_mw
                        min_mw_entry = e
                
                # Apply Max MW Logic: Get corresponding Amps/Time from Max MW entry
                if max_mw_entry:
                    d = max_mw_entry.get('data', {})
                    stats['max_mw_date'] = max_mw_entry['date']
                    stats['max_mw_time'] = d.get('max', {}).get('time', '-')
                    
                    stats['max_amps'] = get_float(d.get('max', {}).get('amps'))
                    stats['max_amps_date'] = max_mw_entry['date']
                    stats['max_amps_time'] = d.get('max', {}).get('time', '-')
                
                # Apply Min MW Logic: Get corresponding Amps/Time from Min MW entry
                if min_mw_entry:
                    d = min_mw_entry.get('data', {})
                    stats['min_mw_date'] = min_mw_entry['date']
                    stats['min_mw_time'] = d.get('min', {}).get('time', '-')
                    
                    stats['min_amps'] = get_float(d.get('min', {}).get('amps'))
                    stats['min_amps_date'] = min_mw_entry['date']
                    stats['min_amps_time'] = d.get('min', {}).get('time', '-')
                
                # Cleanup and averages
                if stats['max_amps'] is None or stats['max_amps'] == -float('inf'): stats['max_amps'] = '-'
                if stats['min_amps'] is None or stats['min_amps'] == float('inf'): stats['min_amps'] = '-'
                stats['avg_amps'] = round(total_amps / count_amps, 2) if count_amps > 0 else '-'
                
                if stats['max_mw'] == -float('inf'): stats['max_mw'] = '-'
                if stats['min_mw'] == float('inf'): stats['min_mw'] = '-'
                stats['avg_mw'] = round(total_mw / count_mw, 2) if count_mw > 0 else '-'
                
            return stats

        # Add spacing
        ws.append([])
        ws.append([])
        
        # Summary Header
        summary_header_row = ws.max_row + 1
        ws.cell(row=summary_header_row, column=1, value="Monthly Summary Report")
        ws.cell(row=summary_header_row, column=1).font = Font(bold=True, size=14)
        
        # Summary Table Headers
        header_row = ws.max_row + 1
        
        if feeder['type'] == 'bus_station':
            summary_headers = [
                "Period", "Parameter", 
                "Maximum Value", "Date", "Time",
                "Minimum Value", "Date", "Time"
            ]
        else:
            summary_headers = [
                "Period", "Parameter", 
                "Maximum Value", "Date", "Time",
                "Minimum Value", "Date", "Time",
                "Average Value"
            ]
        
        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # Add border
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

        # Populate Summary Data
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center")
        
        for p in periods:
            # Filter entries for period
            p_entries = [e for e in entries if p['start'] <= e['date'] <= p['end']]
            stats = calculate_period_stats(p_entries, feeder['type'])
            
            start_row = ws.max_row + 1
            
            if feeder['type'] == 'bus_station':
                # Rows: 400KV, 220KV, Station Load
                rows_data = [
                    ("400KV Bus Voltage", stats['max_400kv'], format_date(stats['max_400kv_date']), format_time_local(stats['max_400kv_time']), stats['min_400kv'], format_date(stats['min_400kv_date']), format_time_local(stats['min_400kv_time'])),
                    ("220KV Bus Voltage", stats['max_220kv'], format_date(stats['max_220kv_date']), format_time_local(stats['max_220kv_time']), stats['min_220kv'], format_date(stats['min_220kv_date']), format_time_local(stats['min_220kv_time'])),
                    ("Station Load (MW)", stats['max_load'], format_date(stats['max_load_date']), format_time_local(stats['max_load_time']), "-", "-", "-")
                ]
                
                # Merge Period Cell
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+2, end_column=1)
                p_cell = ws.cell(row=start_row, column=1, value=p['name'])
                p_cell.alignment = center_align
                p_cell.border = thin_border
                # Apply border to merged cells (needs to be applied to all cells in range or use style)
                # Simplest is to apply to top-left and ensure others are handled, but openpyxl requires iterating range for borders
                for r in range(start_row, start_row+3):
                    ws.cell(row=r, column=1).border = thin_border
                
                for i, (param, max_v, max_d, max_t, min_v, min_d, min_t) in enumerate(rows_data):
                    r = start_row + i
                    ws.cell(row=r, column=2, value=param).border = thin_border
                    ws.cell(row=r, column=3, value=max_v).border = thin_border
                    ws.cell(row=r, column=3).alignment = center_align
                    ws.cell(row=r, column=4, value=max_d).border = thin_border
                    ws.cell(row=r, column=4).alignment = center_align
                    ws.cell(row=r, column=5, value=max_t).border = thin_border
                    ws.cell(row=r, column=5).alignment = center_align
                    ws.cell(row=r, column=6, value=min_v).border = thin_border
                    ws.cell(row=r, column=6).alignment = center_align
                    ws.cell(row=r, column=7, value=min_d).border = thin_border
                    ws.cell(row=r, column=7).alignment = center_align
                    ws.cell(row=r, column=8, value=min_t).border = thin_border
                    ws.cell(row=r, column=8).alignment = center_align

            else:
                # Rows: Amps, MW
                rows_data = [
                    ("Amps", stats['max_amps'], format_date(stats['max_amps_date']), format_time_local(stats['max_amps_time']), stats['min_amps'], format_date(stats['min_amps_date']), format_time_local(stats['min_amps_time']), stats['avg_amps']),
                    ("MW", stats['max_mw'], format_date(stats['max_mw_date']), format_time_local(stats['max_mw_time']), stats['min_mw'], format_date(stats['min_mw_date']), format_time_local(stats['min_mw_time']), stats['avg_mw'])
                ]
                
                # Merge Period Cell
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+1, end_column=1)
                p_cell = ws.cell(row=start_row, column=1, value=p['name'])
                p_cell.alignment = center_align
                p_cell.border = thin_border
                for r in range(start_row, start_row+2):
                    ws.cell(row=r, column=1).border = thin_border
                    
                for i, (param, max_v, max_d, max_t, min_v, min_d, min_t, avg_v) in enumerate(rows_data):
                    r = start_row + i
                    ws.cell(row=r, column=2, value=param).border = thin_border
                    ws.cell(row=r, column=3, value=max_v).border = thin_border
                    ws.cell(row=r, column=3).alignment = center_align
                    ws.cell(row=r, column=4, value=max_d).border = thin_border
                    ws.cell(row=r, column=4).alignment = center_align
                    ws.cell(row=r, column=5, value=max_t).border = thin_border
                    ws.cell(row=r, column=5).alignment = center_align
                    ws.cell(row=r, column=6, value=min_v).border = thin_border
                    ws.cell(row=r, column=6).alignment = center_align
                    ws.cell(row=r, column=7, value=min_d).border = thin_border
                    ws.cell(row=r, column=7).alignment = center_align
                    ws.cell(row=r, column=8, value=min_t).border = thin_border
                    ws.cell(row=r, column=8).alignment = center_align
                    ws.cell(row=r, column=9, value=avg_v).border = thin_border
                    ws.cell(row=r, column=9).alignment = center_align
        
        # Auto-width columns with wrapping support
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            # Calculate max length based on DATA only (skip header at row 1)
            for cell in col[1:]:
                try:
                    val = str(cell.value or "")
                    if len(val) > max_length:
                        max_length = len(val)
                except:
                    pass
            
            # Determine width: minimal 8, maximal 20, or fit data
            # We ignore header length because we enabled wrap_text for headers
            adjusted_width = max(8, min(max_length + 2, 20))
            ws.column_dimensions[column].width = adjusted_width
                    
        # Adjust widths for summary section if needed
        # We already adjusted for data, but summary might be wider?
        # Actually, headers are usually wide enough.
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{feeder['name']}_{year}_{month:02d}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JSONResponse(
            status_code=500,
            content={"detail": str(e)},
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Credentials": "true",
            }
        )

@api_router.post("/init-feeders")
async def initialize_feeders(current_user: User = Depends(get_current_user)):
    existing_count = await db.feeders.count_documents({})
    if existing_count > 0:
        return {"message": "Feeders already initialized", "count": existing_count}
    
    feeders_data = [
        {"name": "400 KV Shankarpally-Narsapur-1", "end1": "Shankarapally", "end2": "Narsapur", "mf": [0.2, 0.2, 1, 1]},
        {"name": "400 KV Shankarpally-MHRM-2", "end1": "Shankarapally", "end2": "Maheshwaram-2", "mf": [0.2, 0.2, 1, 1]},
        {"name": "400 KV Shankarpally-MHRM-1", "end1": "Shankarapally", "end2": "Maheshwaram-1", "mf": [0.2, 0.2, 1, 1]},
        {"name": "220 KV Sadasivapet-1", "end1": "Shankarapally", "end2": "Sadasivapet", "mf": [1, 1, 1.6, 1.6]},
        {"name": "220 KV Sadasivapet-2", "end1": "Shankarapally", "end2": "Sadasivapet", "mf": [1, 1, 1.6, 1.6]},
        {"name": "220 KV Parigi-1", "end1": "Shankarapally", "end2": "Parigi", "mf": [1, 1, 1, 1]},
        {"name": "220 KV Parigi-2", "end1": "Shankarapally", "end2": "Parigi", "mf": [1, 1, 1, 1]},
        {"name": "220 KV Tandur", "end1": "Shankarapally", "end2": "Tandur", "mf": [1, 1, 1, 1]},
        {"name": "220 KV Gachibowli-1", "end1": "Shankarapally", "end2": "Gachibowli", "mf": [2, 2, 2, 2]},
        {"name": "220 KV Gachibowli-2", "end1": "Shankarapally", "end2": "Gachibowli", "mf": [2, 2, 2, 2]},
        {"name": "220 KV KethiReddyPally", "end1": "Shankarapally", "end2": "KethiReddyPally", "mf": [1, 1, 1, 1]},
        {"name": "400 KV KethiReddyPally-1", "end1": "Shankarapally", "end2": "KethiReddyPally", "mf": [1, 1, 1, 1]},
        {"name": "220 KV Yeddumailaram-1", "end1": "Shankarapally", "end2": "Yeddumailaram", "mf": [1, 1, 1, 1]},
        {"name": "220 KV Yeddumailaram-2", "end1": "Shankarapally", "end2": "Yeddumailaram", "mf": [1, 1, 1, 1]},
        {"name": "400 KV Shankarpally-Narsapur-2", "end1": "Shankarapally", "end2": "Narsapur", "mf": [0.2, 0.2, 1, 1]},
        {"name": "400 KV Nizamabad-1&2", "end1": "Shankarapally", "end2": "Nizamabad", "mf": [1000, 1000, 1000, 1000]},
        {"name": "400 KV KethiReddyPally-2", "end1": "Shankarapally", "end2": "KethiReddyPally", "mf": [1, 1, 1, 1]}
    ]
    
    feeders = []
    for fd in feeders_data:
        feeder_obj = Feeder(
            name=fd['name'],
            end1_name=fd['end1'],
            end2_name=fd['end2'],
            end1_import_mf=fd['mf'][0],
            end1_export_mf=fd['mf'][1],
            end2_import_mf=fd['mf'][2],
            end2_export_mf=fd['mf'][3]
        )
        doc = feeder_obj.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        feeders.append(doc)
    
    await db.feeders.insert_many(feeders)
    return {"message": "Feeders initialized successfully", "count": len(feeders)}

@api_router.post("/energy/init")
async def initialize_energy_module(current_user: User = Depends(get_current_user)):
    # Check if already initialized
    if await db.energy_sheets.count_documents({}) > 0:
        return {"message": "Energy module already initialized"}

    # Create Sheets
    sheets_data = ["ICT-1", "ICT-2", "ICT-3", "ICT-4", "33KV"]
    sheet_ids = {}
    
    for name in sheets_data:
        sheet = EnergySheet(name=name)
        doc = sheet.model_dump()
        await db.energy_sheets.insert_one(doc)
        sheet_ids[name] = sheet.id

    # Create Meters
    meters = []
    
    # ICT Meters
    for i in range(1, 5):
        sheet_name = f"ICT-{i}"
        s_id = sheet_ids[sheet_name]
        meters.append(EnergyMeter(sheet_id=s_id, name="HV", mf=0.1 if i==1 else 1.0, unit="MWH")) # Example MF, user said fixed but didn't specify all. using 1.0 default or from screenshot
        meters.append(EnergyMeter(sheet_id=s_id, name="IV", mf=1.5 if i==1 else 1.0, unit="MWH"))
        # Note: Screenshot for ICT-1 shows HV MF=0.1, IV MF=1.5. I'll use those for ICT-1. Others 1.0 for now.

    # 33KV Meters
    s33_id = sheet_ids["33KV"]
    meters.append(EnergyMeter(sheet_id=s33_id, name="Donthapally", mf=1000, unit="KWH")) # Screenshot says MF(0.2) but column says 1000? Wait.
    # Screenshot 33KV: "MF(0.2)" in header? No, "MF(0.2)" might be the CT/PT ratio or something.
    # But column C says "1000". And formula D=(B-A)*C. So MF is 1000.
    # Another meter: "33KV Kandi", MF column says "4".
    meters.append(EnergyMeter(sheet_id=s33_id, name="Kandi", mf=4, unit="KWH"))

    for m in meters:
        doc = m.model_dump()
        await db.energy_meters.insert_one(doc)

    return {"message": "Energy module initialized", "sheets": len(sheets_data), "meters": len(meters)}

@api_router.get("/energy/sheets")
async def get_energy_sheets(current_user: User = Depends(get_current_user)):
    sheets = await db.energy_sheets.find({}, {"_id": 0}).to_list(100)
    result = []
    for sheet in sheets:
        meters = await db.energy_meters.find({"sheet_id": sheet['id']}, {"_id": 0}).sort("order", 1).to_list(100)
        sheet['meters'] = meters
        result.append(sheet)
    return result

@api_router.get("/energy/entries/{sheet_id}")
async def get_energy_entries(
    sheet_id: str,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    query = {"sheet_id": sheet_id}
    if year and month:
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
        query['date'] = {"$gte": start_date, "$lt": end_date}
    
    entries = await db.energy_entries.find(query, {"_id": 0}).sort("date", 1).to_list(1000)
    for entry in entries:
        if isinstance(entry.get('created_at'), str):
            entry['created_at'] = datetime.fromisoformat(entry['created_at'])
        if isinstance(entry.get('updated_at'), str):
            entry['updated_at'] = datetime.fromisoformat(entry['updated_at'])
    return entries

@api_router.post("/energy/entries")
async def save_energy_entry(
    entry_input: EnergyEntryInput,
    current_user: User = Depends(get_current_user)
):
    # Get previous day's entry for initials
    date_obj = datetime.strptime(entry_input.date, "%Y-%m-%d")
    prev_date = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
    prev_entry = await db.energy_entries.find_one(
        {"sheet_id": entry_input.sheet_id, "date": prev_date},
        {"_id": 0}
    )
    
    prev_finals = {}
    if prev_entry:
        for r in prev_entry['readings']:
            prev_finals[r['meter_id']] = r['final']
            
    # Calculate readings
    readings = []
    total_consumption = 0
    
    for r_in in entry_input.readings:
        meter = await db.energy_meters.find_one({"id": r_in.meter_id})
        if not meter:
            continue
            
        initial = prev_finals.get(r_in.meter_id, 0.0)
        consumption = (r_in.final - initial) * meter['mf']
        total_consumption += consumption
        
        readings.append(EnergyReading(
            meter_id=r_in.meter_id,
            initial=initial,
            final=r_in.final,
            consumption=consumption
        ))
        
    # Upsert entry
    existing = await db.energy_entries.find_one(
        {"sheet_id": entry_input.sheet_id, "date": entry_input.date}
    )
    
    entry_data = EnergyEntry(
        sheet_id=entry_input.sheet_id,
        date=entry_input.date,
        readings=readings,
        total_consumption=total_consumption
    )
    
    doc = entry_data.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    if existing:
        # Preserve original created_at
        doc['created_at'] = existing['created_at']
        await db.energy_entries.replace_one({"_id": existing['_id']}, doc)
    else:
        await db.energy_entries.insert_one(doc)
        
    return entry_data

@api_router.get("/line-losses/export-all/{year}/{month}")
async def export_all_line_losses(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    feeders = await db.feeders.find({}, {"_id": 0}).to_list(100)
    
    # Sort feeders based on predefined order
    FEEDER_ORDER = [
        "400KV MAHESHWARAM-2",
        "400KV MAHESHWARAM-1",
        "400KV NARSAPUR-1",
        "400KV NARSAPUR-2",
        "400KV KETHIREDDYPALLY-1",
        "400KV KETHIREDDYPALLY-2",
        "400KV NIZAMABAD-1",
        "400KV NIZAMABAD-2",
        "220KV PARIGI-1",
        "220KV PARIGI-2",
        "220KV THANDUR",
        "220KV GACHIBOWLI-1",
        "220KV GACHIBOWLI-2",
        "220KV KETHIREDDYPALLY",
        "220KV YEDDUMAILARAM-1",
        "220KV YEDDUMAILARAM-2",
        "220KV SADASIVAPET-1",
        "220KV SADASIVAPET-2"
    ]
    
    feeders.sort(key=lambda x: FEEDER_ORDER.index(x['name']) if x['name'] in FEEDER_ORDER else 999)
    
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
        
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    if not feeders:
        ws = wb.create_sheet("No Data")
        ws.cell(row=1, column=1, value="No feeders found")
        
    for feeder in feeders:
        ws = wb.create_sheet(title=feeder['name'][:30]) # Sheet name limit 31 chars
        
        headers = [
            "Date",
            f"{feeder['end1_name']} Import Initial",
            f"{feeder['end1_name']} Import Final",
            f"{feeder['end1_name']} Import MF",
            f"{feeder['end1_name']} Import Consumption",
            f"{feeder['end1_name']} Export Initial",
            f"{feeder['end1_name']} Export Final",
            f"{feeder['end1_name']} Export MF",
            f"{feeder['end1_name']} Export Consumption",
            f"{feeder['end2_name']} Import Initial",
            f"{feeder['end2_name']} Import Final",
            f"{feeder['end2_name']} Import MF",
            f"{feeder['end2_name']} Import Consumption",
            f"{feeder['end2_name']} Export Initial",
            f"{feeder['end2_name']} Export Final",
            f"{feeder['end2_name']} Export MF",
            f"{feeder['end2_name']} Export Consumption",
            "% Loss"
        ]
        
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        entries = await db.entries.find(
            {"feeder_id": feeder['id'], "date": {"$gte": start_date, "$lt": end_date}},
            {"_id": 0}
        ).sort("date", 1).to_list(1000)
        
        for entry in entries:
            ws.append([
                format_date(entry['date']),
                entry['end1_import_initial'],
                entry['end1_import_final'],
                feeder['end1_import_mf'],
                entry['end1_import_consumption'],
                entry['end1_export_initial'],
                entry['end1_export_final'],
                feeder['end1_export_mf'],
                entry['end1_export_consumption'],
                entry['end2_import_initial'],
                entry['end2_import_final'],
                feeder['end2_import_mf'],
                entry['end2_import_consumption'],
                entry['end2_export_initial'],
                entry['end2_export_final'],
                feeder['end2_export_mf'],
                entry['end2_export_consumption'],
                entry['loss_percent']
            ])
            
        # Auto-fit columns with wrap text logic
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            # Skip header row for width calculation to avoid overly wide columns
            for cell in col[1:]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Set width based on data content with a minimum and maximum limit
            # This allows headers to wrap if they are long
            adjusted_width = max(max_length + 2, 12)
            ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Line_Losses_{month}-{year}.xlsx"}
    )


@api_router.get("/max-min/export-all/{year}/{month}")
async def export_all_max_min_data(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    # Fetch all feeders
    feeders = await db.max_min_feeders.find({}, {"_id": 0}).to_list(100)
    
    # Sort feeders based on predefined order
    FEEDER_ORDER = [
        "Bus Voltages & Station Load",
        "400KV MAHESHWARAM-2",
        "400KV MAHESHWARAM-1",
        "400KV NARSAPUR-1",
        "400KV NARSAPUR-2",
        "400KV KETHIREDDYPALLY-1",
        "400KV KETHIREDDYPALLY-2",
        "400KV NIZAMABAD-1",
        "400KV NIZAMABAD-2",
        "ICT-1 (315MVA)",
        "ICT-2 (315MVA)",
        "ICT-3 (315MVA)",
        "ICT-4 (500MVA)",
        "220KV PARIGI-1",
        "220KV PARIGI-2",
        "220KV THANDUR",
        "220KV GACHIBOWLI-1",
        "220KV GACHIBOWLI-2",
        "220KV KETHIREDDYPALLY",
        "220KV YEDDUMAILARAM-1",
        "220KV YEDDUMAILARAM-2",
        "220KV SADASIVAPET-1",
        "220KV SADASIVAPET-2"
    ]
    
    feeders.sort(key=lambda x: FEEDER_ORDER.index(x['name']) if x['name'] in FEEDER_ORDER else 999)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"All Feeders {month}-{year}"
    
    # Styles
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    current_col = 1
    
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
    
    import calendar
    last_day = calendar.monthrange(year, month)[1]

    for feeder in feeders:
        entries = await db.max_min_entries.find(
            {"feeder_id": feeder['id'], "date": {"$gte": start_date, "$lt": end_date}},
            {"_id": 0}
        ).sort("date", 1).to_list(1000)
        
        # Headers
        if feeder['type'] == 'bus_station':
            headers = [
                "Date",
                "Max Bus Voltage 400KV Value", "Time",
                "Max Bus Voltage 220KV Value", "Time",
                "Min Bus Voltage 400KV Value", "Time",
                "Min Bus Voltage 220KV Value", "Time",
                "Station Load Max MW", "Station Load MVAR", "Station Load Time"
            ]
        elif feeder['type'] == 'ict_feeder':
            headers = ["Date", "Max Amps", "Max MW", "Max MVAR", "Max Time", "Min Amps", "Min MW", "Min MVAR", "Min Time", "Avg Amps", "Avg MW"]
        else:
            headers = ["Date", "Max Amps", "Max MW", "Max Time", "Min Amps", "Min MW", "Min Time", "Avg Amps", "Avg MW"]
            
        # Write Feeder Name
        ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + len(headers) - 1)
        cell = ws.cell(row=1, column=current_col, value=feeder['name'])
        cell.alignment = center_align
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        for i in range(len(headers)):
            ws.cell(row=1, column=current_col+i).border = thin_border
        
        # Write Headers
        for i, h in enumerate(headers):
            cell = ws.cell(row=2, column=current_col + i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # Write Data
        row_idx = 3
        for entry in entries:
            d = entry.get('data', {})
            row_data = [format_date(entry['date'])]
            if feeder['type'] == 'bus_station':
                row_data.extend([
                    d.get('max_bus_voltage_400kv', {}).get('value', ''), format_time(d.get('max_bus_voltage_400kv', {}).get('time', '')),
                    d.get('max_bus_voltage_220kv', {}).get('value', ''), format_time(d.get('max_bus_voltage_220kv', {}).get('time', '')),
                    d.get('min_bus_voltage_400kv', {}).get('value', ''), format_time(d.get('min_bus_voltage_400kv', {}).get('time', '')),
                    d.get('min_bus_voltage_220kv', {}).get('value', ''), format_time(d.get('min_bus_voltage_220kv', {}).get('time', '')),
                    d.get('station_load', {}).get('max_mw', ''), d.get('station_load', {}).get('mvar', ''), format_time(d.get('station_load', {}).get('time', ''))
                ])
            elif feeder['type'] == 'ict_feeder':
                row_data.extend([
                    d.get('max', {}).get('amps', ''), d.get('max', {}).get('mw', ''), d.get('max', {}).get('mvar', ''), format_time(d.get('max', {}).get('time', '')),
                    d.get('min', {}).get('amps', ''), d.get('min', {}).get('mw', ''), d.get('min', {}).get('mvar', ''), format_time(d.get('min', {}).get('time', '')),
                    d.get('avg', {}).get('amps', ''), d.get('avg', {}).get('mw', '')
                ])
            else:
                row_data.extend([
                    d.get('max', {}).get('amps', ''), d.get('max', {}).get('mw', ''), format_time(d.get('max', {}).get('time', '')),
                    d.get('min', {}).get('amps', ''), d.get('min', {}).get('mw', ''), format_time(d.get('min', {}).get('time', '')),
                    d.get('avg', {}).get('amps', ''), d.get('avg', {}).get('mw', '')
                ])
            
            for i, val in enumerate(row_data):
                cell = ws.cell(row=row_idx, column=current_col + i, value=val)
                cell.alignment = center_align
                cell.border = thin_border
            row_idx += 1
            
        # Write Stats (Summary)
        row_idx += 1
        periods = [
            {"name": "1st to 15th", "start": f"{year}-{month:02d}-01", "end": f"{year}-{month:02d}-15"},
            {"name": "16th to End", "start": f"{year}-{month:02d}-16", "end": f"{year}-{month:02d}-{last_day}"},
            {"name": "Full Month", "start": f"{year}-{month:02d}-01", "end": f"{year}-{month:02d}-{last_day}"}
        ]
        
        for p in periods:
            p_entries = [e for e in entries if p['start'] <= e['date'] <= p['end']]
            stats = calculate_period_stats(p_entries, feeder['type'])
            
            # Header for Period
            ws.merge_cells(start_row=row_idx, start_column=current_col, end_row=row_idx, end_column=current_col + len(headers) - 1)
            cell = ws.cell(row=row_idx, column=current_col, value=f"Summary: {p['name']}")
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            for i in range(len(headers)): ws.cell(row=row_idx, column=current_col+i).border = thin_border
            row_idx += 1
            
            if feeder['type'] == 'bus_station':
                # Write stats rows
                labels = ["Max 400KV", "Min 400KV", "Max 220KV", "Min 220KV", "Max Load"]
                vals = [
                    (stats['max_400kv'], format_date(stats['max_400kv_date']), format_time(stats['max_400kv_time'])),
                    (stats['min_400kv'], format_date(stats['min_400kv_date']), format_time(stats['min_400kv_time'])),
                    (stats['max_220kv'], format_date(stats['max_220kv_date']), format_time(stats['max_220kv_time'])),
                    (stats['min_220kv'], format_date(stats['min_220kv_date']), format_time(stats['min_220kv_time'])),
                    (stats['max_load'], format_date(stats['max_load_date']), format_time(stats['max_load_time']))
                ]
                for label, val_tuple in zip(labels, vals):
                    ws.cell(row=row_idx, column=current_col, value=label).border = thin_border
                    ws.cell(row=row_idx, column=current_col+1, value=val_tuple[0]).border = thin_border
                    ws.cell(row=row_idx, column=current_col+2, value=val_tuple[1]).border = thin_border
                    ws.cell(row=row_idx, column=current_col+3, value=val_tuple[2]).border = thin_border
                    row_idx += 1
            else:
                labels = ["Max Amps", "Min Amps", "Max MW", "Min MW"]
                vals = [
                    (stats['max_amps'], format_date(stats['max_amps_date']), format_time(stats['max_amps_time'])),
                    (stats['min_amps'], format_date(stats['min_amps_date']), format_time(stats['min_amps_time'])),
                    (stats['max_mw'], format_date(stats['max_mw_date']), format_time(stats['max_mw_time'])),
                    (stats['min_mw'], format_date(stats['min_mw_date']), format_time(stats['min_mw_time']))
                ]
                for label, val_tuple in zip(labels, vals):
                    ws.cell(row=row_idx, column=current_col, value=label).border = thin_border
                    ws.cell(row=row_idx, column=current_col+1, value=val_tuple[0]).border = thin_border
                    ws.cell(row=row_idx, column=current_col+2, value=val_tuple[1]).border = thin_border
                    ws.cell(row=row_idx, column=current_col+3, value=val_tuple[2]).border = thin_border
                    row_idx += 1
            row_idx += 1 # Gap between periods
            
        # Auto-fit columns
        for i in range(len(headers)):
            col_letter = ws.cell(row=2, column=current_col+i).column_letter
            max_len = 0
            for row in range(1, row_idx):
                val = ws.cell(row=row, column=current_col+i).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = max_len + 2
            
        current_col += len(headers) + 1 # Gap between feeders

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=MaxMin_All_{year}_{month:02d}.xlsx"}
    )

@api_router.get("/interruptions/export/{feeder_id}/{year}/{month}")
async def export_interruptions_feeder(
    feeder_id: str,
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    if feeder.get("type") not in ["feeder_400kv", "feeder_220kv", "ict_feeder", "reactor_feeder"]:
        raise HTTPException(status_code=400, detail="Interruptions supported only for 400KV, 220KV, ICT and Reactor feeders")
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
    entries = await db.interruption_entries.find(
        {"feeder_id": feeder_id, "date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0},
    ).sort("date", 1).to_list(1000)
    wb = Workbook()
    ws = wb.active
    ws.title = feeder["name"][:31]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    headers = [
        "Date",
        "Time From",
        "Time To",
        "Duration",
        "Cause of Interruption",
        "Relay Indications / LC Work carried out",
        "Breakdown Declared (Yes/No)",
        "Fault Identified During Patrolling",
        "Fault Location",
        "Remarks",
        "Action Taken",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for entry in entries:
        data = entry.get("data", {})
        end_date_str = data.get("end_date") or entry["date"]
        end_time_str = format_time(data.get("end_time") or "")
        if end_date_str == entry["date"]:
            to_value = end_time_str
        else:
            to_value = f"{format_date(end_date_str)} {end_time_str}".strip()
        ws.append(
            [
                format_date(entry["date"]),
                format_time(data.get("start_time") or ""),
                to_value,
                format_duration_hhmm(data.get("duration_minutes")),
                data.get("cause_of_interruption") or "",
                data.get("relay_indications_lc_work") or "",
                data.get("breakdown_declared") or "",
                data.get("fault_identified_during_patrolling") or "",
                data.get("fault_location") or "",
                data.get("remarks") or "",
                data.get("action_taken") or "",
            ]
        )
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Interruptions_{feeder['name']}_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

@api_router.get("/interruptions/export-all/{year}/{month}")
async def export_interruptions_all(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    feeders = await db.max_min_feeders.find(
        {"type": {"$in": ["feeder_400kv", "feeder_220kv", "ict_feeder", "reactor_feeder", "bay_feeder"]}},
        {"_id": 0},
    ).to_list(100)
    def feeder_sort_key(f):
        name = f.get("name", "")
        ftype = f.get("type")
        is_bay = ftype == "bay_feeder"
        if ftype == "feeder_400kv" and not is_bay:
            group = 0
        elif is_bay and name.startswith("4-"):
            group = 1
        elif ftype == "feeder_220kv" and not is_bay:
            group = 2
        elif is_bay and name.startswith("2-"):
            group = 3
        elif ftype == "ict_feeder":
            group = 4
        elif ftype == "reactor_feeder":
            group = 5
        else:
            group = 6
        return (group, name)
    feeders = sorted(feeders, key=feeder_sort_key)
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    if not feeders:
        ws = wb.create_sheet("No Data")
        ws.cell(row=1, column=1, value="No feeders found")
    created_any_sheet = False

    for feeder in feeders:
        entries = await db.interruption_entries.find(
            {"feeder_id": feeder["id"], "date": {"$gte": start_date, "$lt": end_date}},
            {"_id": 0},
        ).sort("date", 1).to_list(1000)
        if not entries:
            continue

        ws = wb.create_sheet(title=feeder["name"][:30])
        created_any_sheet = True

        headers = [
            "Date",
            "Time From",
            "Time To",
            "Duration",
            "Cause of Interruption",
            "Relay Indications / LC Work carried out",
            "Breakdown Declared (Yes/No)",
            "Fault Identified During Patrolling",
            "Fault Location",
            "Remarks",
            "Action Taken",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for entry in entries:
            data = entry.get("data", {})
            end_date_str = data.get("end_date") or entry["date"]
            end_time_str = format_time(data.get("end_time") or "")
            if end_date_str == entry["date"]:
                to_value = end_time_str
            else:
                to_value = f"{format_date(end_date_str)} {end_time_str}".strip()
            ws.append(
                [
                    format_date(entry["date"]),
                    format_time(data.get("start_time") or ""),
                    to_value,
                    format_duration_hhmm(data.get("duration_minutes")),
                    data.get("cause_of_interruption") or "",
                    data.get("relay_indications_lc_work") or "",
                    data.get("breakdown_declared") or "",
                    data.get("fault_identified_during_patrolling") or "",
                    data.get("fault_location") or "",
                    data.get("remarks") or "",
                    data.get("action_taken") or "",
                ]
            )

    if not created_any_sheet and feeders:
        ws = wb.create_sheet("No Data")
        ws.cell(row=1, column=1, value="No interruptions found for this period")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Interruptions_All_{year}_{month:02d}.xlsx"},
    )

async def _get_interruptions_report_data(year: int, month: int):
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"

    feeders = await db.max_min_feeders.find(
        {"type": {"$in": ["feeder_400kv", "feeder_220kv", "ict_feeder", "reactor_feeder", "bay_feeder"]}},
        {"_id": 0},
    ).to_list(200)

    feeders_by_id = {f["id"]: f for f in feeders}

    all_entries = await db.interruption_entries.find(
        {"date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0},
    ).to_list(20000)

    entries_by_feeder = {}
    for entry in all_entries:
        fid = entry.get("feeder_id")
        if not fid or fid not in feeders_by_id:
            continue
        if fid not in entries_by_feeder:
            entries_by_feeder[fid] = []
        entries_by_feeder[fid].append(entry)

    for fid, rows in entries_by_feeder.items():
        rows.sort(
            key=lambda e: (
                e.get("date") or "",
                (e.get("data") or {}).get("start_time") or "",
            )
        )

    def split_cause_and_relay(text: str):
        base = text or ""
        lower = base.lower()

        def strip_time(s: str) -> str:
            # Remove patterns like "at 10:15 hrs" or "at 9.30 hours"
            return re.sub(r"\bat\s+\d{1,2}[:\.]\d{2}\s*(?:hrs?|hours?)?\b", "", s, flags=re.IGNORECASE)

        def trim_edges(s: str) -> str:
            return re.sub(r"^[\s\.\-,:]+", "", re.sub(r"[\s\.\-,:]+$", "", s))

        if "a/r success" in lower and "with following indications" in lower:
            idx_follow = lower.index("with following indications")
            relay = base[idx_follow + len("with following indications") :]
            relay = trim_edges(relay)
            cause_segment = base[:idx_follow]
            idx_ar = lower.index("a/r success")
            if idx_ar != -1:
                cause_segment = base[idx_ar:idx_follow]
            cause = strip_time(cause_segment)
            cause = trim_edges(cause)
            return cause, relay

        if "lc issued" in lower:
            idx_lc = lower.index("lc issued")
            segment = base[idx_lc:]
            lower_seg = segment.lower()
            idx_for = lower_seg.find(" for ")
            if idx_for != -1:
                cause_segment = segment[:idx_for]
                relay_segment = segment[idx_for:]
            else:
                cause_segment = segment
                relay_segment = ""
            cause = strip_time(cause_segment)
            cause = trim_edges(cause)
            relay = trim_edges(relay_segment or "")
            return cause, relay

        idx_for_plain = lower.find(" for ")
        if idx_for_plain != -1:
            cause_segment = base[:idx_for_plain]
            cause = strip_time(cause_segment)
            cause = trim_edges(cause)
            relay = trim_edges(base[idx_for_plain:])
            return cause, relay

        return base, ""

    def build_groups(target_types, bay_prefix=None):
        relevant = []
        for f in feeders:
            ftype = f.get("type")
            name = f.get("name", "")
            if ftype in target_types:
                relevant.append(f)
            elif ftype == "bay_feeder" and bay_prefix and name.startswith(bay_prefix):
                relevant.append(f)
        relevant.sort(key=lambda x: x.get("name", ""))

        groups = []
        for feeder in relevant:
            fid = feeder.get("id")
            feeder_entries = entries_by_feeder.get(fid, [])
            if not feeder_entries:
                continue

            rows = []
            for idx, entry in enumerate(feeder_entries, 1):
                data = entry.get("data") or {}
                end_date_str = data.get("end_date") or entry.get("date")
                end_time_raw = data.get("end_time") or ""
                end_time_fmt = format_time(end_time_raw) if end_time_raw else ""
                if not end_time_fmt:
                    to_value = ""
                elif end_date_str == entry.get("date"):
                    to_value = end_time_fmt
                else:
                    to_value = f"{format_date(end_date_str)} {end_time_fmt}"

                duration_minutes = data.get("duration_minutes")
                duration_str = format_duration_hhmm(duration_minutes)

                remarks_parts = []
                remarks_val = data.get("remarks")
                action_val = data.get("action_taken")
                if remarks_val:
                    remarks_parts.append(str(remarks_val))
                if action_val:
                    remarks_parts.append(str(action_val))
                remarks_combined = "\n".join(remarks_parts)

                description_val = data.get("description") or ""
                cause_val = data.get("cause_of_interruption") or ""
                relay_val = data.get("relay_indications_lc_work") or ""

                if (not cause_val or cause_val == description_val) and description_val:
                    parsed_cause, parsed_relay = split_cause_and_relay(description_val)
                    cause_val = parsed_cause or description_val
                    if not relay_val:
                        relay_val = parsed_relay or ""

                rows.append(
                    {
                        "sl_no": idx,
                        "date": format_date(entry.get("date")),
                        "time_from": format_time(data.get("start_time") or ""),
                        "time_to": to_value,
                        "duration": duration_str,
                        "cause": cause_val,
                        "relay": relay_val,
                        "breakdown": data.get("breakdown_declared") or "",
                        "fault_identified": data.get("fault_identified_during_patrolling") or "",
                        "fault_location": data.get("fault_location") or "",
                        "remarks": remarks_combined,
                    }
                )

            if rows:
                groups.append({"name": feeder.get("name", ""), "rows": rows})

        return groups

    month_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    if 1 <= month <= 12:
        month_label = month_labels[month - 1]
    else:
        month_label = str(month)

    sections = [
        {
            "id": "400kv",
            "title": "400KV Feeders",
            "header": f"Interruptions of 400KV SS Shankarpally 400KV Feeders for the Month of {month_label}-{year}",
            "groups": build_groups(["feeder_400kv"], bay_prefix="4-"),
        },
        {
            "id": "220kv",
            "title": "220KV Feeders",
            "header": f"Interruptions of 400KV SS Shankarpally 220KV Feeders for the Month of {month_label}-{year}",
            "groups": build_groups(["feeder_220kv"], bay_prefix="2-"),
        },
        {
            "id": "ict_reactor",
            "title": "ICTs & 125MVAR Reactor",
            "header": f"Interruptions of 400KV SS Shankarpally ICTs / Bays / Reactor for the Month of {month_label}-{year}",
            "groups": build_groups(["ict_feeder", "reactor_feeder"]),
        },
    ]

    return {
        "sections": sections,
        "month_label": month_label,
        "year": year,
    }

async def _generate_interruptions_report_wb(year: int, month: int):
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter

    data = await _get_interruptions_report_data(year, month)

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    headers = [
        "Sl. No",
        "Date",
        "Time From",
        "Time To",
        "Duration",
        "Cause of Interruption",
        "Relay Indications / LC Work carried out",
        "Break down declared or Not",
        "Fault identified in patrolling",
        "Fault Location",
        "Remarks and action taken",
    ]

    sections = data.get("sections") or []
    for section in sections:
        ws = wb.create_sheet(title=section.get("title", "")[:31] or "Sheet")

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        c = ws.cell(row=1, column=1, value=section.get("header", ""))
        c.font = Font(bold=True)
        c.alignment = center_align
        c.border = thin_border

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        row_idx = 3
        groups = section.get("groups") or []
        if not groups:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            cell = ws.cell(row=row_idx, column=1, value="No interruptions found for this period")
            cell.alignment = center_align
            cell.border = thin_border
            continue

        for group in groups:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
            cell = ws.cell(row=row_idx, column=1, value=group.get("name", ""))
            cell.font = Font(bold=True)
            cell.alignment = center_align
            cell.border = thin_border
            row_idx += 1

            for row in group.get("rows") or []:
                values = [
                    row.get("sl_no"),
                    row.get("date"),
                    row.get("time_from"),
                    row.get("time_to"),
                    row.get("duration"),
                    row.get("cause"),
                    row.get("relay"),
                    row.get("breakdown"),
                    row.get("fault_identified"),
                    row.get("fault_location"),
                    row.get("remarks"),
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx <= 5 or col_idx >= 8 and col_idx <= 10:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                row_idx += 1

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    return wb


async def _get_mis_interruptions_report_data(year: int, month: int):
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"

    feeders = await db.max_min_feeders.find(
        {"type": {"$in": ["feeder_400kv", "feeder_220kv", "ict_feeder", "reactor_feeder", "bay_feeder"]}},
        {"_id": 0},
    ).to_list(200)

    feeders_by_id = {f["id"]: f for f in feeders if f.get("id")}

    all_entries = await db.interruption_entries.find(
        {"date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0},
    ).to_list(20000)

    def _combined_text(entry: Dict[str, Any]) -> str:
        data = entry.get("data") or {}
        parts = [
            data.get("cause_of_interruption"),
            data.get("relay_indications_lc_work"),
            data.get("remarks"),
            data.get("description"),
            data.get("action_taken"),
        ]
        return " ".join(str(p) for p in parts if p) or ""

    def classify_category(entry: Dict[str, Any]) -> str:
        combined = _combined_text(entry)
        lower = combined.lower()
        breakdown_flag = False
        flag_val = str(data.get("breakdown_declared") or "").strip().lower()
        if flag_val.startswith("y"):
            breakdown_flag = True
        if "breakdown declared" in lower or "under breakdown condition" in lower:
            breakdown_flag = True
        if breakdown_flag:
            return "breakdown"
        lc_nbfc_keywords = [
            "lc issued",
            "lc applied",
            "under lc",
            "line clear",
            "line clear for",
            "nbfc issued",
            "nbfc applied",
            "under nbfc",
            "nbfc for",
            "nbfc returned",
            "hand tripped",
            "hand trip",
            "hand-tripped",
        ]
        if any(k in lower for k in lc_nbfc_keywords):
            return "lc_nbfc_ht"
        return "faulty_tripping"

    def classify_lc_nbfc_label(entry: Dict[str, Any]) -> str:
        combined = _combined_text(entry)
        lower = combined.lower()
        if "nbfc" in lower:
            return "NBFC"
        ht_keywords = [
            "hand tripped",
            "hand trip",
            "hand-tripped",
            " ht ",
            "(ht",
            "ht)",
        ]
        if any(k in lower for k in ht_keywords):
            return "HT"
        return "LC"

    summary_by_feeder: Dict[str, Dict[str, Any]] = {}
    for f in feeders:
        fid = f.get("id")
        if not fid:
            continue
        summary_by_feeder[fid] = {
            "lc_nbfc_ht_count": 0,
            "lc_nbfc_ht_duration": 0.0,
            "lc_count": 0,
            "nbfc_count": 0,
            "ht_count": 0,
            "breakdown_count": 0,
            "breakdown_duration": 0.0,
            "faulty_tripping_count": 0,
            "faulty_tripping_duration": 0.0,
        }

    for entry in all_entries:
        fid = entry.get("feeder_id")
        if not fid or fid not in feeders_by_id:
            continue
        if fid not in summary_by_feeder:
            continue
        data = entry.get("data") or {}
        duration_raw = data.get("duration_minutes") or 0
        try:
            minutes = float(duration_raw)
        except Exception:
            minutes = 0.0
        category = classify_category(entry)
        if category == "lc_nbfc_ht":
            count_key = "lc_nbfc_ht_count"
            dur_key = "lc_nbfc_ht_duration"
            label = classify_lc_nbfc_label(entry)
        elif category == "breakdown":
            count_key = "breakdown_count"
            dur_key = "breakdown_duration"
        else:
            count_key = "faulty_tripping_count"
            dur_key = "faulty_tripping_duration"
        summary = summary_by_feeder[fid]
        summary[count_key] += 1
        summary[dur_key] += minutes
        if category == "lc_nbfc_ht":
            if label == "LC":
                summary["lc_count"] += 1
            elif label == "NBFC":
                summary["nbfc_count"] += 1
            elif label == "HT":
                summary["ht_count"] += 1

    month_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    if 1 <= month <= 12:
        month_label = month_labels[month - 1]
    else:
        month_label = str(month)

    def section_key_for_feeder(feeder: Dict[str, Any]) -> Optional[str]:
        ftype = feeder.get("type")
        name = feeder.get("name", "")
        if ftype == "feeder_400kv":
            return "400kv"
        if ftype == "feeder_220kv":
            return "220kv"
        if ftype == "ict_feeder":
            return "ict_reactor"
        if ftype == "reactor_feeder":
            return "ict_reactor"
        if ftype == "bay_feeder":
            return "bay"
        return None

    def _build_mis_interruptions_flat_rows(
        feeders_list: List[Dict[str, Any]],
        summary_map: Dict[str, Dict[str, Any]],
        section_key_fn,
    ) -> List[Dict[str, Any]]:
        ordered_sections = ["400kv", "220kv", "ict_reactor", "bay"]
        feeders_by_section: Dict[str, List[Dict[str, Any]]] = {k: [] for k in ordered_sections}
        for f in feeders_list:
            fid = f.get("id")
            if not fid or fid not in summary_map:
                continue
            skey = section_key_fn(f)
            if not skey or skey not in feeders_by_section:
                continue
            feeders_by_section[skey].append(f)
        for flist in feeders_by_section.values():
            flist.sort(key=lambda x: x.get("name", ""))
        rows: List[Dict[str, Any]] = []
        sl = 1
        for skey in ordered_sections:
            for feeder in feeders_by_section.get(skey, []):
                fid = feeder.get("id")
                if not fid:
                    continue
                summary = summary_map.get(fid) or {}
                faulty_count = summary.get("faulty_tripping_count", 0)
                faulty_duration_minutes = summary.get("faulty_tripping_duration", 0.0)
                breakdown_count = summary.get("breakdown_count", 0)
                breakdown_duration_minutes = summary.get("breakdown_duration", 0.0)
                lc_nbfc_count = summary.get("lc_nbfc_ht_count", 0)
                lc_nbfc_duration_minutes = summary.get("lc_nbfc_ht_duration", 0.0)
                lc_count = summary.get("lc_count", 0)
                nbfc_count = summary.get("nbfc_count", 0)
                ht_count = summary.get("ht_count", 0)
                has_data = any(
                    [
                        faulty_count,
                        faulty_duration_minutes,
                        breakdown_count,
                        breakdown_duration_minutes,
                        lc_nbfc_count,
                        lc_nbfc_duration_minutes,
                    ]
                )
                if not has_data:
                    continue
                lc_parts: List[str] = []
                if lc_count:
                    lc_parts.append(f"{lc_count}(LC)")
                if nbfc_count:
                    lc_parts.append(f"{nbfc_count}(NBFC)")
                if ht_count:
                    lc_parts.append(f"{ht_count}(HT)")
                lc_nbfc_display = "\n".join(lc_parts) if lc_parts else ""
                rows.append(
                    {
                        "sl_no": sl,
                        "feeder_name": feeder.get("name", ""),
                        "faulty_ol_count": 0,
                        "faulty_el_count": faulty_count,
                        "faulty_other_count": 0,
                        "faulty_duration": format_duration_hhmm(faulty_duration_minutes),
                        "incoming_count": 0,
                        "incoming_duration": "",
                        "load_relief_count": 0,
                        "load_relief_duration": "",
                        "equipment_failures_count": 0,
                        "equipment_failures_duration": "",
                        "breakdown_count": breakdown_count,
                        "breakdown_duration": format_duration_hhmm(breakdown_duration_minutes),
                        "pre_arranged_count": 0,
                        "pre_arranged_duration": "",
                        "lc_nbfc_count": lc_nbfc_display,
                        "lc_nbfc_duration": format_duration_hhmm(lc_nbfc_duration_minutes),
                        "remarks": "",
                    }
                )
                sl += 1
        return rows

    return {
        "rows": _build_mis_interruptions_flat_rows(feeders, summary_by_feeder, section_key_for_feeder),
        "month_label": month_label,
        "year": year,
    }


async def _generate_mis_interruptions_report_wb(year: int, month: int):
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter

    data = await _get_mis_interruptions_report_data(year, month)

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    header_font = Font(bold=True, color="000000")

    month_label = data.get("month_label") or ""
    year_val = data.get("year") or year

    wb = Workbook()
    ws = wb.active
    ws.title = "Statement-16"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
    c1 = ws.cell(row=1, column=1, value="STATEMENT-16")
    c1.font = Font(bold=True)
    c1.alignment = center_align

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=19)
    header_text = f"PARTICULARS OF INTERRUPTIONS FOR THE MONTH OF {month_label} - {year_val} IN O & M-I DIVISION"
    c2 = ws.cell(row=2, column=1, value=header_text)
    c2.font = Font(bold=True)
    c2.alignment = center_align

    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=6)
    ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=8)
    ws.merge_cells(start_row=3, start_column=9, end_row=3, end_column=10)
    ws.merge_cells(start_row=3, start_column=11, end_row=3, end_column=12)
    ws.merge_cells(start_row=3, start_column=13, end_row=3, end_column=14)
    ws.merge_cells(start_row=3, start_column=15, end_row=3, end_column=16)
    ws.merge_cells(start_row=3, start_column=17, end_row=3, end_column=18)
    ws.merge_cells(start_row=3, start_column=19, end_row=4, end_column=19)

    headers = {
        (3, 1): "Sl. No",
        (3, 2): "Name of the feeder",
        (3, 3): "Faulty Trippings",
        (3, 7): "Incoming Supply",
        (3, 9): "Load Relief",
        (3, 11): "Equipment Failures",
        (3, 13): "Break Downs",
        (3, 15): "Pre Arranged",
        (3, 17): "LC/NBFC",
        (3, 19): "Remarks",
        (4, 3): "O/L",
        (4, 4): "E/L",
        (4, 5): "Others",
        (4, 6): "Duration",
        (4, 7): "No",
        (4, 8): "Duration",
        (4, 9): "No",
        (4, 10): "Duration",
        (4, 11): "No",
        (4, 12): "Duration",
        (4, 13): "No",
        (4, 14): "Duration",
        (4, 15): "No",
        (4, 16): "Duration",
        (4, 17): "No",
        (4, 18): "Duration",
    }

    for (row_idx, col_idx), value in headers.items():
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    row_idx = 5
    rows = data.get("rows") or []
    if not rows:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=19)
        cell = ws.cell(row=row_idx, column=1, value="No interruptions found for this period")
        cell.alignment = center_align
        cell.border = thin_border
    else:
        for row in rows:
            values = [
                row.get("sl_no"),
                row.get("feeder_name"),
                row.get("faulty_ol_count"),
                row.get("faulty_el_count"),
                row.get("faulty_other_count"),
                row.get("faulty_duration"),
                row.get("incoming_count"),
                row.get("incoming_duration"),
                row.get("load_relief_count"),
                row.get("load_relief_duration"),
                row.get("equipment_failures_count"),
                row.get("equipment_failures_duration"),
                row.get("breakdown_count"),
                row.get("breakdown_duration"),
                row.get("pre_arranged_count"),
                row.get("pre_arranged_duration"),
                row.get("lc_nbfc_count"),
                row.get("lc_nbfc_duration"),
                row.get("remarks"),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx in [1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            row_idx += 1

    for col_idx in range(1, 20):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)

    return wb


async def _generate_fortnight_report_wb(year: int, month: int):
    try:
        import io
        import calendar
        import traceback
        from openpyxl.utils import get_column_letter

        # Fetch all feeders
        feeders = await db.max_min_feeders.find({}, {"_id": 0}).to_list(100)
        
        # Sort feeders based on predefined order (INCLUDING ICTs)
        FEEDER_ORDER = [
            "Bus Voltages & Station Load",
            "400KV MAHESHWARAM-2",
            "400KV MAHESHWARAM-1",
            "400KV NARSAPUR-1",
            "400KV NARSAPUR-2",
            "400KV KETHIREDDYPALLY-1",
            "400KV KETHIREDDYPALLY-2",
            "400KV NIZAMABAD-1",
            "400KV NIZAMABAD-2",
            "220KV PARIGI-1",
            "220KV PARIGI-2",
            "220KV THANDUR",
            "220KV GACHIBOWLI-1",
            "220KV GACHIBOWLI-2",
            "220KV KETHIREDDYPALLY",
            "220KV YEDDUMAILARAM-1",
            "220KV YEDDUMAILARAM-2",
            "220KV SADASIVAPET-1",
            "220KV SADASIVAPET-2",
            "ICT-1 (315MVA)",
            "ICT-2 (315MVA)",
            "ICT-3 (315MVA)",
            "ICT-4 (500MVA)"
        ]
        
        ordered_feeders = []
        # Add Bus Station first
        bus_feeder = next((x for x in feeders if x.get('type') == "bus_station"), None)
        if bus_feeder:
            ordered_feeders.append(bus_feeder)
            
        # Add others in order
        for name in FEEDER_ORDER:
            if bus_feeder and name == bus_feeder.get('name'): continue
            f = next((x for x in feeders if x.get('name') == name), None)
            if f:
                ordered_feeders.append(f)
                
        
        # Helper to determine rating
        def get_rating(name):
            if "400KV" in name: return "400KV"
            if "220KV" in name: return "220KV"
            if "315MVA" in name: return "315MVA"
            if "500MVA" in name: return "500MVA"
            return "-"

        # Split feeders
        main_feeders = [f for f in ordered_feeders if f['type'] != 'ict_feeder' and f['type'] != 'bus_station']
        ict_feeders = [f for f in ordered_feeders if f['type'] == 'ict_feeder']
        bus_station_feeder = next((f for f in ordered_feeders if f['type'] == 'bus_station'), None)

        wb = Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)
            
        last_day = calendar.monthrange(year, month)[1]
        month_name = calendar.month_name[month]
        
        periods = [
            {"name": "1-15", "start": f"{year}-{month:02d}-01", "end": f"{year}-{month:02d}-15"},
            {"name": "16-End", "start": f"{year}-{month:02d}-16", "end": f"{year}-{month:02d}-{last_day}"},
            {"name": "Full Month", "start": f"{year}-{month:02d}-01", "end": f"{year}-{month:02d}-{last_day}"}
        ]
        
        header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        # header_font = Font(bold=True, color="000000") # Black text
        title_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Fetch all entries for the month
        all_entries = await db.max_min_entries.find(
            {"date": {"$gte": f"{year}-{month:02d}-01", "$lte": f"{year}-{month:02d}-{last_day}"}},
            {"_id": 0}
        ).to_list(10000)

        entries_by_feeder = {}
        for e in all_entries:
            fid = e.get('feeder_id')
            if fid:
                if fid not in entries_by_feeder:
                    entries_by_feeder[fid] = []
                entries_by_feeder[fid].append(e)

        for p in periods:
            ws = wb.create_sheet(title=p['name'])
            
            # Row 1: Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
            title_cell = ws.cell(row=1, column=1, value=f"MAX DEMAND OF EACH FEEDER and ICT DURING THE MONTH OF {month_name}-{year} ({p['name']})")
            title_cell.font = title_font
            title_cell.alignment = center_align
            
            # Row 2: Main Headers
            # Col 1: Sl.No
            ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
            c = ws.cell(row=2, column=1, value="Sl.No")
            c.font = bold_font; c.alignment = center_align; c.border = thin_border
            ws.cell(row=3, column=1).border = thin_border
            
            # Col 2: Name of the feeder
            ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
            c = ws.cell(row=2, column=2, value="Name of the feeder")
            c.font = bold_font; c.alignment = center_align; c.border = thin_border
            ws.cell(row=3, column=2).border = thin_border
            
            # Col 3: Rating
            ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)
            c = ws.cell(row=2, column=3, value="Rating")
            c.font = bold_font; c.alignment = center_align; c.border = thin_border
            ws.cell(row=3, column=3).border = thin_border
            
            # Col 4-7: Max Demand reached during
            ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=7)
            c = ws.cell(row=2, column=4, value="Max Demand reached during")
            c.font = bold_font; c.alignment = center_align; c.border = thin_border
            for i in range(5, 8): ws.cell(row=2, column=i).border = thin_border
            
            # Col 8-11: Min Demand reached
            ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=11)
            c = ws.cell(row=2, column=8, value="Min Demand reached")
            c.font = bold_font; c.alignment = center_align; c.border = thin_border
            for i in range(9, 12): ws.cell(row=2, column=i).border = thin_border
            
            # Col 12: Remarks
            ws.merge_cells(start_row=2, start_column=12, end_row=3, end_column=12)
            c = ws.cell(row=2, column=12, value="Remarks")
            c.font = bold_font; c.alignment = center_align; c.border = thin_border
            ws.cell(row=3, column=12).border = thin_border
            
            # Row 3: Sub Headers
            sub_headers = ["AMPS", "MW", "Date", "Time", "AMPS", "MW", "Date", "Time"]
            for i, h in enumerate(sub_headers):
                c = ws.cell(row=3, column=4+i, value=h)
                c.font = bold_font; c.alignment = center_align; c.border = thin_border
                
            row_idx = 4
            
            # 1. Main Feeders
            for i, feeder in enumerate(main_feeders, 1):
                f_entries = entries_by_feeder.get(feeder['id'], [])
                p_entries = [e for e in f_entries if p['start'] <= e['date'] <= p['end']]
                
                # Check for Grouping Logic
                is_special, partner_names = get_feeder_group_info(feeder['name'])
                if is_special:
                    p_group_map = {}
                    p_group_map[feeder['id']] = p_entries
                    
                    # Gather partner entries
                    for pname in partner_names:
                         p_feeder = next((x for x in feeders if x['name'] == pname), None)
                         if p_feeder:
                             pf_entries = entries_by_feeder.get(p_feeder['id'], [])
                             pp_entries = [e for e in pf_entries if p['start'] <= e['date'] <= p['end']]
                             p_group_map[p_feeder['id']] = pp_entries
                    
                    leader_id = determine_leader(p_group_map)
                    if not leader_id: leader_id = feeder['id']
                    
                    leader_entries = p_group_map.get(leader_id, [])
                    stats = calculate_coincident_stats(leader_entries, feeder['id'], p_group_map, feeder['type'])
                else:
                    stats = calculate_period_stats(p_entries, feeder['type'])
                
                ws.cell(row=row_idx, column=1, value=i).border = thin_border
                ws.cell(row=row_idx, column=1).alignment = center_align
                
                ws.cell(row=row_idx, column=2, value=feeder['name']).border = thin_border
                
                ws.cell(row=row_idx, column=3, value=get_rating(feeder['name'])).border = thin_border
                ws.cell(row=row_idx, column=3).alignment = center_align
                
                # Max
                ws.cell(row=row_idx, column=4, value=stats.get('max_amps', '-')).border = thin_border
                ws.cell(row=row_idx, column=5, value=stats.get('max_mw', '-')).border = thin_border
                ws.cell(row=row_idx, column=6, value=format_date(stats.get('max_mw_date'))).border = thin_border
                ws.cell(row=row_idx, column=7, value=format_time(stats.get('max_mw_time'))).border = thin_border
                
                # Min
                ws.cell(row=row_idx, column=8, value=stats.get('min_amps', '-')).border = thin_border
                ws.cell(row=row_idx, column=9, value=stats.get('min_mw', '-')).border = thin_border
                ws.cell(row=row_idx, column=10, value=format_date(stats.get('min_mw_date'))).border = thin_border
                ws.cell(row=row_idx, column=11, value=format_time(stats.get('min_mw_time'))).border = thin_border
                
                # Remarks
                ws.cell(row=row_idx, column=12, value="").border = thin_border
                
                # Center align data cells
                for c in range(4, 12): ws.cell(row=row_idx, column=c).alignment = center_align
                
                row_idx += 1
                
            # 2. ICT Separator
            row_idx += 1 # Gap
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
            c = ws.cell(row=row_idx, column=1, value="ICT'S")
            c.font = bold_font; c.alignment = center_align
            # No border for separator? Or strictly follow image? Usually just text.
            row_idx += 1
            
            # 3. ICT Feeders
            start_sl = len(main_feeders) + 1
            for i, feeder in enumerate(ict_feeders, start_sl):
                f_entries = entries_by_feeder.get(feeder['id'], [])
                p_entries = [e for e in f_entries if p['start'] <= e['date'] <= p['end']]
                
                # Check for Grouping Logic (ICTs are in the group list)
                is_special, partner_names = get_feeder_group_info(feeder['name'])
                if is_special:
                    p_group_map = {}
                    p_group_map[feeder['id']] = p_entries
                    
                    # Gather partner entries
                    for pname in partner_names:
                         p_feeder = next((x for x in feeders if x['name'] == pname), None)
                         if p_feeder:
                             pf_entries = entries_by_feeder.get(p_feeder['id'], [])
                             pp_entries = [e for e in pf_entries if p['start'] <= e['date'] <= p['end']]
                             p_group_map[p_feeder['id']] = pp_entries
                    
                    leader_id = determine_leader(p_group_map)
                    if not leader_id: leader_id = feeder['id']
                    
                    leader_entries = p_group_map.get(leader_id, [])
                    stats = calculate_coincident_stats(leader_entries, feeder['id'], p_group_map, feeder['type'])
                else:
                    stats = calculate_period_stats(p_entries, feeder['type'])
                
                ws.cell(row=row_idx, column=1, value=i).border = thin_border
                ws.cell(row=row_idx, column=1).alignment = center_align
                
                ws.cell(row=row_idx, column=2, value=feeder['name']).border = thin_border
                
                ws.cell(row=row_idx, column=3, value=get_rating(feeder['name'])).border = thin_border
                ws.cell(row=row_idx, column=3).alignment = center_align
                
                # Max
                ws.cell(row=row_idx, column=4, value=stats.get('max_amps', '-')).border = thin_border
                ws.cell(row=row_idx, column=5, value=stats.get('max_mw', '-')).border = thin_border
                ws.cell(row=row_idx, column=6, value=format_date(stats.get('max_mw_date'))).border = thin_border
                ws.cell(row=row_idx, column=7, value=format_time(stats.get('max_mw_time'))).border = thin_border
                
                # Min
                ws.cell(row=row_idx, column=8, value=stats.get('min_amps', '-')).border = thin_border
                ws.cell(row=row_idx, column=9, value=stats.get('min_mw', '-')).border = thin_border
                ws.cell(row=row_idx, column=10, value=format_date(stats.get('min_mw_date'))).border = thin_border
                ws.cell(row=row_idx, column=11, value=format_time(stats.get('min_mw_time'))).border = thin_border
                
                # Remarks
                ws.cell(row=row_idx, column=12, value="").border = thin_border
                
                # Center align data cells
                for c in range(4, 12): ws.cell(row=row_idx, column=c).alignment = center_align
                
                row_idx += 1
                
            # 4. Station Load (Bottom)
            row_idx += 2
            if bus_station_feeder:
                f_entries = entries_by_feeder.get(bus_station_feeder['id'], [])
                p_entries = [e for e in f_entries if p['start'] <= e['date'] <= p['end']]
                stats = calculate_period_stats(p_entries, bus_station_feeder['type'])
                
                # Header
                ws.cell(row=row_idx, column=2, value="Station Load in MW").border = thin_border
                ws.cell(row=row_idx, column=2).font = bold_font; ws.cell(row=row_idx, column=2).alignment = center_align
                
                ws.cell(row=row_idx, column=3, value="Time").border = thin_border
                ws.cell(row=row_idx, column=3).font = bold_font; ws.cell(row=row_idx, column=3).alignment = center_align
                
                ws.cell(row=row_idx, column=4, value="Date").border = thin_border
                ws.cell(row=row_idx, column=4).font = bold_font; ws.cell(row=row_idx, column=4).alignment = center_align
                
                row_idx += 1
                
                # Value
                ws.cell(row=row_idx, column=1, value="Month").border = thin_border # Based on image 2 "Month" cell
                ws.cell(row=row_idx, column=2, value=stats.get('max_load', '-')).border = thin_border
                ws.cell(row=row_idx, column=2).alignment = center_align
                
                ws.cell(row=row_idx, column=3, value=format_time(stats.get('max_load_time'))).border = thin_border
                ws.cell(row=row_idx, column=3).alignment = center_align
                
                ws.cell(row=row_idx, column=4, value=format_date(stats.get('max_load_date'))).border = thin_border
                ws.cell(row=row_idx, column=4).alignment = center_align
                    
            # Set fixed width for Sl.No
            ws.column_dimensions['A'].width = 6
            
            # Auto-fit other columns
            for i in range(2, ws.max_column + 1):
                col_letter = get_column_letter(i)
                max_len = 0
                for cell in ws[col_letter]:
                    # Skip the first row (Title) to avoid skewed width
                    if cell.row == 1: continue
                    try:
                        if cell.value and len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except: pass
                ws.column_dimensions[col_letter].width = max_len + 2
                
        return wb
    except Exception as e:
        import traceback
        import os
        error_msg = traceback.format_exc()
        print(error_msg)
        raise e

@api_router.get("/reports/fortnight/{year}/{month}")
async def generate_fortnight_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        import calendar
        wb = await _generate_fortnight_report_wb(year, month)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        month_name = calendar.month_name[month]
        filename = f"Fortnight_Report_{month_name}_{year}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"detail": str(e)},
            headers={
                "Access-Control-Allow-Origin": "http://localhost:3000",
                "Access-Control-Allow-Credentials": "true"
            }
        )

@api_router.get("/reports/fortnight/preview/{year}/{month}")
async def preview_fortnight_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        import calendar
        
        # Fetch all feeders
        feeders = await db.max_min_feeders.find({}, {"_id": 0}).to_list(100)
        
        # Sort feeders based on predefined order (INCLUDING ICTs)
        FEEDER_ORDER = [
            "Bus Voltages & Station Load",
            "400KV MAHESHWARAM-2",
            "400KV MAHESHWARAM-1",
            "400KV NARSAPUR-1",
            "400KV NARSAPUR-2",
            "400KV KETHIREDDYPALLY-1",
            "400KV KETHIREDDYPALLY-2",
            "400KV NIZAMABAD-1",
            "400KV NIZAMABAD-2",
            "220KV PARIGI-1",
            "220KV PARIGI-2",
            "220KV THANDUR",
            "220KV GACHIBOWLI-1",
            "220KV GACHIBOWLI-2",
            "220KV KETHIREDDYPALLY",
            "220KV YEDDUMAILARAM-1",
            "220KV YEDDUMAILARAM-2",
            "220KV SADASIVAPET-1",
            "220KV SADASIVAPET-2",
            "ICT-1 (315MVA)",
            "ICT-2 (315MVA)",
            "ICT-3 (315MVA)",
            "ICT-4 (500MVA)"
        ]
        
        ordered_feeders = []
        # Add Bus Station first
        bus_feeder = next((x for x in feeders if x.get('type') == "bus_station"), None)
        if bus_feeder:
            ordered_feeders.append(bus_feeder)
            
        # Add others in order
        for name in FEEDER_ORDER:
            if bus_feeder and name == bus_feeder.get('name'): continue
            f = next((x for x in feeders if x.get('name') == name), None)
            if f:
                ordered_feeders.append(f)
                
        # Helper to determine rating
        def get_rating(name):
            if "400KV" in name: return "400KV"
            if "220KV" in name: return "220KV"
            if "315MVA" in name: return "315MVA"
            if "500MVA" in name: return "500MVA"
            return "-"

        # Split feeders
        main_feeders = [f for f in ordered_feeders if f['type'] != 'ict_feeder' and f['type'] != 'bus_station']
        ict_feeders = [f for f in ordered_feeders if f['type'] == 'ict_feeder']
        bus_station_feeder = next((f for f in ordered_feeders if f['type'] == 'bus_station'), None)

        last_day = calendar.monthrange(year, month)[1]
        
        periods = [
            {"name": "1-15", "start": f"{year}-{month:02d}-01", "end": f"{year}-{month:02d}-15"},
            {"name": "16-End", "start": f"{year}-{month:02d}-16", "end": f"{year}-{month:02d}-{last_day}"},
            {"name": "Full Month", "start": f"{year}-{month:02d}-01", "end": f"{year}-{month:02d}-{last_day}"},
        ]
        
        # Fetch all entries for the month
        all_entries = await db.max_min_entries.find(
            {"date": {"$gte": f"{year}-{month:02d}-01", "$lte": f"{year}-{month:02d}-{last_day}"}},
            {"_id": 0}
        ).to_list(10000)

        entries_by_feeder = {}
        for e in all_entries:
            fid = e.get('feeder_id')
            if fid:
                if fid not in entries_by_feeder:
                    entries_by_feeder[fid] = []
                entries_by_feeder[fid].append(e)

        preview_data = {"periods": []}

        for p in periods:
            period_data = {
                "name": p['name'],
                "main_feeders": [],
                "ict_feeders": [],
                "station_load": None
            }
            
            # 1. Main Feeders
            for i, feeder in enumerate(main_feeders, 1):
                f_entries = entries_by_feeder.get(feeder['id'], [])
                p_entries = [e for e in f_entries if p['start'] <= e['date'] <= p['end']]
                
                # Check for Grouping Logic
                is_special, partner_names = get_feeder_group_info(feeder['name'])
                if is_special:
                    p_group_map = {}
                    p_group_map[feeder['id']] = p_entries
                    
                    # Gather partner entries
                    for pname in partner_names:
                         p_feeder = next((x for x in feeders if x['name'] == pname), None)
                         if p_feeder:
                             pf_entries = entries_by_feeder.get(p_feeder['id'], [])
                             pp_entries = [e for e in pf_entries if p['start'] <= e['date'] <= p['end']]
                             p_group_map[p_feeder['id']] = pp_entries
                    
                    leader_id = determine_leader(p_group_map)
                    if not leader_id: leader_id = feeder['id']
                    
                    leader_entries = p_group_map.get(leader_id, [])
                    stats = calculate_coincident_stats(leader_entries, feeder['id'], p_group_map, feeder['type'])
                else:
                    stats = calculate_period_stats(p_entries, feeder['type'])
                
                period_data["main_feeders"].append({
                    "sl_no": i,
                    "name": feeder['name'],
                    "rating": get_rating(feeder['name']),
                    "max_amps": stats.get('max_amps', '-'),
                    "max_mw": stats.get('max_mw', '-'),
                    "max_mw_date": format_date(stats.get('max_mw_date')),
                    "max_mw_time": format_time(stats.get('max_mw_time')),
                    "min_amps": stats.get('min_amps', '-'),
                    "min_mw": stats.get('min_mw', '-'),
                    "min_mw_date": format_date(stats.get('min_mw_date')),
                    "min_mw_time": format_time(stats.get('min_mw_time')),
                })
                
            # 2. ICT Feeders
            start_sl = len(main_feeders) + 1
            for i, feeder in enumerate(ict_feeders, start_sl):
                f_entries = entries_by_feeder.get(feeder['id'], [])
                p_entries = [e for e in f_entries if p['start'] <= e['date'] <= p['end']]
                
                # Check for Grouping Logic
                is_special, partner_names = get_feeder_group_info(feeder['name'])
                if is_special:
                    p_group_map = {}
                    p_group_map[feeder['id']] = p_entries
                    
                    # Gather partner entries
                    for pname in partner_names:
                         p_feeder = next((x for x in feeders if x['name'] == pname), None)
                         if p_feeder:
                             pf_entries = entries_by_feeder.get(p_feeder['id'], [])
                             pp_entries = [e for e in pf_entries if p['start'] <= e['date'] <= p['end']]
                             p_group_map[p_feeder['id']] = pp_entries
                    
                    leader_id = determine_leader(p_group_map)
                    if not leader_id: leader_id = feeder['id']
                    
                    leader_entries = p_group_map.get(leader_id, [])
                    stats = calculate_coincident_stats(leader_entries, feeder['id'], p_group_map, feeder['type'])
                else:
                    stats = calculate_period_stats(p_entries, feeder['type'])
                
                period_data["ict_feeders"].append({
                    "sl_no": i,
                    "name": feeder['name'],
                    "rating": get_rating(feeder['name']),
                    "max_amps": stats.get('max_amps', '-'),
                    "max_mw": stats.get('max_mw', '-'),
                    "max_mw_date": format_date(stats.get('max_mw_date')),
                    "max_mw_time": format_time(stats.get('max_mw_time')),
                    "min_amps": stats.get('min_amps', '-'),
                    "min_mw": stats.get('min_mw', '-'),
                    "min_mw_date": format_date(stats.get('min_mw_date')),
                    "min_mw_time": format_time(stats.get('min_mw_time')),
                })

            # 3. Station Load (Calculated from ICTs Sum)
            ict_feeders_data = period_data.get("ict_feeders", [])
            
            total_station_load = 0.0
            station_load_time = '-'
            station_load_date = '-'
            valid_load_found = False
            
            for ict in ict_feeders_data:
                mw_val = ict.get('max_mw')
                
                # Filter out invalid values
                if mw_val == '-' or mw_val == -float('inf') or mw_val is None:
                    continue
                    
                try:
                    val = float(mw_val)
                    total_station_load += val
                    valid_load_found = True
                    
                    # Capture time from the first contributing ICT
                    if station_load_time == '-' or station_load_time == '':
                         station_load_time = ict.get('max_mw_time', '-')
                         station_load_date = ict.get('max_mw_date', '-')
                except (ValueError, TypeError):
                    continue
            
            period_data["station_load"] = {
                "max_load": round(total_station_load, 2) if valid_load_found else '-',
                "max_load_time": station_load_time,
                "max_load_date": station_load_date
            }
                
            preview_data["periods"].append(period_data)

        return preview_data
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": "Failed to generate Fortnight preview."})

@api_router.get("/reports/max-min/daily-preview/{date_str}")
async def preview_max_min_daily_report(
    date_str: str,
    current_user: User = Depends(get_current_user)
):
    try:
        # Validate date format
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return JSONResponse(status_code=400, content={"detail": "Invalid date format. Use YYYY-MM-DD"})

        # Fetch all feeders
        feeders = await db.max_min_feeders.find({}, {"_id": 0}).to_list(100)
        
        # Sort feeders based on predefined order (INCLUDING ICTs) - Same as Fortnight
        FEEDER_ORDER = [
            "Bus Voltages & Station Load",
            "400KV MAHESHWARAM-2",
            "400KV MAHESHWARAM-1",
            "400KV NARSAPUR-1",
            "400KV NARSAPUR-2",
            "400KV KETHIREDDYPALLY-1",
            "400KV KETHIREDDYPALLY-2",
            "400KV NIZAMABAD-1",
            "400KV NIZAMABAD-2",
            "220KV PARIGI-1",
            "220KV PARIGI-2",
            "220KV THANDUR",
            "220KV GACHIBOWLI-1",
            "220KV GACHIBOWLI-2",
            "220KV KETHIREDDYPALLY",
            "220KV YEDDUMAILARAM-1",
            "220KV YEDDUMAILARAM-2",
            "220KV SADASIVAPET-1",
            "220KV SADASIVAPET-2",
            "ICT-1 (315MVA)",
            "ICT-2 (315MVA)",
            "ICT-3 (315MVA)",
            "ICT-4 (500MVA)"
        ]
        
        ordered_feeders = []
        # Add Bus Station first
        bus_feeder = next((x for x in feeders if x.get('type') == "bus_station"), None)
        if bus_feeder:
            ordered_feeders.append(bus_feeder)
            
        # Add others in order
        for name in FEEDER_ORDER:
            if bus_feeder and name == bus_feeder.get('name'): continue
            f = next((x for x in feeders if x.get('name') == name), None)
            if f:
                ordered_feeders.append(f)
                
        # Helper to determine rating
        def get_rating(name):
            if "400KV" in name: return "400KV"
            if "220KV" in name: return "220KV"
            if "315MVA" in name: return "315MVA"
            if "500MVA" in name: return "500MVA"
            return "-"

        # Split feeders
        main_feeders = [f for f in ordered_feeders if f['type'] != 'ict_feeder' and f['type'] != 'bus_station']
        ict_feeders = [f for f in ordered_feeders if f['type'] == 'ict_feeder']
        bus_station_feeder = next((f for f in ordered_feeders if f['type'] == 'bus_station'), None)
        if not bus_station_feeder:
             bus_station_feeder = next((x for x in feeders if x.get('type') == "bus_station"), None)

        # Single period for the day
        periods = [
            {"name": "Daily Details", "start": date_str, "end": date_str}
        ]
        
        # Fetch entries for the specific date
        all_entries = await db.max_min_entries.find(
            {"date": date_str},
            {"_id": 0}
        ).to_list(10000)

        entries_by_feeder = {}
        for e in all_entries:
            fid = e.get('feeder_id')
            if fid:
                if fid not in entries_by_feeder:
                    entries_by_feeder[fid] = []
                entries_by_feeder[fid].append(e)

        preview_data = {"periods": []}

        for p in periods:
            period_data = {
                "name": p['name'],
                "main_feeders": [],
                "ict_feeders": [],
                "station_load": None
            }
            
            # 1. Main Feeders
            for i, feeder in enumerate(main_feeders, 1):
                f_entries = entries_by_feeder.get(feeder['id'], [])
                p_entries = f_entries # Already filtered by date
                
                # Restore original coincident grouping logic for double-circuit/ICT groups
                is_special, partner_names = get_feeder_group_info(feeder['name'])
                if is_special:
                    p_group_map = {}
                    p_group_map[feeder['id']] = p_entries
                    
                    # Gather partner entries
                    for pname in partner_names:
                        p_feeder = next((x for x in feeders if x['name'] == pname), None)
                        if p_feeder:
                            pf_entries = entries_by_feeder.get(p_feeder['id'], [])
                            p_group_map[p_feeder['id']] = pf_entries
                    
                    leader_id = determine_leader(p_group_map)
                    if not leader_id:
                        leader_id = feeder['id']
                    
                    leader_entries = p_group_map.get(leader_id, [])
                    stats = calculate_coincident_stats(leader_entries, feeder['id'], p_group_map, feeder['type'])
                else:
                    stats = calculate_period_stats(p_entries, feeder['type'])
                
                period_data["main_feeders"].append({
                    "sl_no": i,
                    "name": feeder['name'],
                    "rating": get_rating(feeder['name']),
                    "max_amps": stats.get('max_amps', '-'),
                    "max_mw": stats.get('max_mw', '-'),
                    "max_mw_date": format_date(stats.get('max_mw_date')),
                    "max_mw_time": format_time(stats.get('max_mw_time')),
                    "min_amps": stats.get('min_amps', '-'),
                    "min_mw": stats.get('min_mw', '-'),
                    "min_mw_date": format_date(stats.get('min_mw_date')),
                    "min_mw_time": format_time(stats.get('min_mw_time')),
                })
                
            # 2. ICT Feeders
            start_sl = len(main_feeders) + 1
            for i, feeder in enumerate(ict_feeders, start_sl):
                f_entries = entries_by_feeder.get(feeder['id'], [])
                p_entries = f_entries
                
                # Restore original coincident grouping logic for ICT groups
                is_special, partner_names = get_feeder_group_info(feeder['name'])
                if is_special:
                    p_group_map = {}
                    p_group_map[feeder['id']] = p_entries
                    
                    # Gather partner entries
                    for pname in partner_names:
                        p_feeder = next((x for x in feeders if x['name'] == pname), None)
                        if p_feeder:
                            pf_entries = entries_by_feeder.get(p_feeder['id'], [])
                            p_group_map[p_feeder['id']] = pf_entries
                    
                    leader_id = determine_leader(p_group_map)
                    if not leader_id:
                        leader_id = feeder['id']
                    
                    leader_entries = p_group_map.get(leader_id, [])
                    stats = calculate_coincident_stats(leader_entries, feeder['id'], p_group_map, feeder['type'])
                else:
                    stats = calculate_period_stats(p_entries, feeder['type'])
                
                period_data["ict_feeders"].append({
                    "sl_no": i,
                    "name": feeder['name'],
                    "rating": get_rating(feeder['name']),
                    "max_amps": stats.get('max_amps', '-'),
                    "max_mw": stats.get('max_mw', '-'),
                    "max_mw_date": format_date(stats.get('max_mw_date')),
                    "max_mw_time": format_time(stats.get('max_mw_time')),
                    "min_amps": stats.get('min_amps', '-'),
                    "min_mw": stats.get('min_mw', '-'),
                    "min_mw_date": format_date(stats.get('min_mw_date')),
                    "min_mw_time": format_time(stats.get('min_mw_time')),
                })

            # 3. Station Load (Calculated from ICTs Sum)
            ict_feeders_data = period_data.get("ict_feeders", [])
            
            total_station_load = 0.0
            station_load_time = '-'
            station_load_date = '-'
            valid_load_found = False
            
            for ict in ict_feeders_data:
                mw_val = ict.get('max_mw')
                
                # Filter out invalid values
                if mw_val == '-' or mw_val == -float('inf') or mw_val is None:
                    continue
                    
                try:
                    val = float(mw_val)
                    total_station_load += val
                    valid_load_found = True
                    
                    # Capture time from the first contributing ICT
                    if station_load_time == '-' or station_load_time == '':
                         station_load_time = ict.get('max_mw_time', '-')
                         station_load_date = ict.get('max_mw_date', '-')
                except (ValueError, TypeError):
                    continue
            
            period_data["station_load"] = {
                "max_load": round(total_station_load, 2) if valid_load_found else '-',
                "max_load_time": station_load_time,
                "max_load_date": station_load_date
            }
                
            preview_data["periods"].append(period_data)

        return preview_data
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": "Failed to generate Daily preview."})

@api_router.get("/reports/energy/daily-preview/{date_str}")
async def preview_energy_daily_report(
    date_str: str,
    current_user: User = Depends(get_current_user)
):
    try:
        # Validate date format
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return JSONResponse(status_code=400, content={"detail": "Invalid date format. Use YYYY-MM-DD"})

        # Fetch all sheets
        sheets = await db.energy_sheets.find({}, {"_id": 0}).sort("name", 1).to_list(100)
        
        # Fetch all meters
        all_meters = await db.energy_meters.find({}, {"_id": 0}).to_list(1000)
        
        # Fetch entries for the date
        entries = await db.energy_entries.find(
            {"date": date_str},
            {"_id": 0}
        ).to_list(1000)
        
        # Map entries by sheet_id
        entries_by_sheet = {e['sheet_id']: e for e in entries}
        
        result_sheets = []
        
        for sheet in sheets:
            sheet_meters = [m for m in all_meters if m.get('sheet_id') == sheet['id']]
            # Sort meters by order if available, else by name
            sheet_meters.sort(key=lambda x: x.get('order', 999))
            
            entry = entries_by_sheet.get(sheet['id'])
            
            meter_data = []
            
            readings_map = {}
            if entry:
                readings_map = {r['meter_id']: r for r in entry.get('readings', [])}
                
            for idx, meter in enumerate(sheet_meters, 1):
                reading = readings_map.get(meter['id'])
                meter_data.append({
                    "sl_no": idx,
                    "name": meter['name'],
                    "initial": reading['initial'] if reading else None,
                    "final": reading['final'] if reading else None,
                    "mf": meter.get('mf', 1),
                    "consumption": reading['consumption'] if reading else None,
                    "unit": meter.get('unit', 'KWH')
                })
            
            result_sheets.append({
                "name": sheet['name'],
                "meters": meter_data
            })
            
        return {"sheets": result_sheets}
        
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": "Failed to generate Energy Daily preview."})


@api_router.get("/reports/daily-max-mva/preview/{year}/{month}")
async def get_daily_max_mva_preview(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        import math
        import calendar
        
        # 1. Find "Bus Voltages & Station Load" feeder
        feeder = await db.max_min_feeders.find_one({"name": "Bus Voltages & Station Load"})
        if not feeder:
            feeder = await db.max_min_feeders.find_one({"type": "bus_station"})
            
        if not feeder:
            return JSONResponse(status_code=404, content={"detail": "Bus Station feeder not found"})

        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"

        entries = await db.max_min_entries.find(
            {"feeder_id": feeder['id'], "date": {"$gte": start_date, "$lt": end_date}},
            {"_id": 0}
        ).sort("date", 1).to_list(1000)
        
        report_data = []
        
        # Helper to parse float
        def parse_float(val):
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        last_day = calendar.monthrange(year, month)[1]
        entries_map = {e['date']: e for e in entries}
        
        for day in range(1, last_day + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            entry = entries_map.get(date_str)
            
            row = {
                "date": format_date(date_str),
                "mw": "",
                "mvar": "",
                "time": "",
                "mva": ""
            }
            
            if entry:
                d = entry.get('data', {}).get('station_load', {})
                mw = d.get('max_mw')
                mvar = d.get('mvar')
                time = d.get('time')
                
                row["mw"] = mw if mw is not None else ""
                row["mvar"] = mvar if mvar is not None else ""
                row["time"] = format_time(time) if time else ""
                
                mw_val = parse_float(mw)
                mvar_val = parse_float(mvar)
                
                if mw_val is not None and mvar_val is not None:
                    mva = math.sqrt(mw_val**2 + mvar_val**2)
                    row["mva"] = f"{mva:.2f}"
            
            report_data.append(row)
            
        return report_data
        
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        with open("error_log.txt", "w") as f:
            f.write(error_msg)
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": str(e)})


async def _generate_daily_max_mva_wb(year: int, month: int):
    try:
        import math
        import calendar
        import io
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
        from openpyxl.utils import get_column_letter
        
        feeder = await db.max_min_feeders.find_one({"name": "Bus Voltages & Station Load"})
        if not feeder:
            feeder = await db.max_min_feeders.find_one({"type": "bus_station"})
            
        if not feeder:
            raise HTTPException(status_code=404, detail="Bus Station feeder not found")

        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"

        entries = await db.max_min_entries.find(
            {"feeder_id": feeder['id'], "date": {"$gte": start_date, "$lt": end_date}},
            {"_id": 0}
        ).sort("date", 1).to_list(1000)
        
        def parse_float(val):
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        wb = Workbook()
        ws = wb.active
        ws.title = f"Daily Max MVA {month}-{year}"
        
        # Styles
        header_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        headers = ["Date", "MW", "MVAR", "Time", "MVA"]
        
        for i, h in enumerate(headers):
            cell = ws.cell(row=1, column=i+1, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        entries_map = {e['date']: e for e in entries}
        last_day = calendar.monthrange(year, month)[1]
        
        row_idx = 2
        for day in range(1, last_day + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            formatted_date = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d-%b-%y")
            
            entry = entries_map.get(date_str)
            
            ws.cell(row=row_idx, column=1, value=formatted_date).border = thin_border
            ws.cell(row=row_idx, column=1).alignment = center_align
            
            for c in range(2, 6):
                ws.cell(row=row_idx, column=c).border = thin_border
                ws.cell(row=row_idx, column=c).alignment = center_align
            
            if entry:
                d = entry.get('data', {}).get('station_load', {})
                mw = d.get('max_mw')
                mvar = d.get('mvar')
                time = d.get('time')
                
                if mw is not None: ws.cell(row=row_idx, column=2, value=mw)
                if mvar is not None: ws.cell(row=row_idx, column=3, value=mvar)
                if time: ws.cell(row=row_idx, column=4, value=format_time(time))
                
                mw_val = parse_float(mw)
                mvar_val = parse_float(mvar)
                
                if mw_val is not None and mvar_val is not None:
                    mva = math.sqrt(mw_val**2 + mvar_val**2)
                    ws.cell(row=row_idx, column=5, value=f"{mva:.2f}")
            
            row_idx += 1

        for i in range(1, 6):
            col_letter = get_column_letter(i)
            max_len = 0
            for cell in ws[col_letter]:
                try:
                    if cell.value and len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except: pass
            ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

        return wb
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise e

@api_router.get("/reports/daily-max-mva/export/{year}/{month}")
async def export_daily_max_mva_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        import io
        wb = await _generate_daily_max_mva_wb(year, month)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Daily_Max_MVA_{month}-{year}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": str(e)})




async def _generate_energy_export_wb(year: int, month: int):
    sheets = await db.energy_sheets.find({}, {"_id": 0}).to_list(100)
    sheets.sort(key=lambda x: x['name'])
    
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
        
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for sheet in sheets:
        ws = wb.create_sheet(title=sheet['name'])
        
        meters = await db.energy_meters.find({"sheet_id": sheet['id']}, {"_id": 0}).to_list(100)
        
        entries = await db.energy_entries.find(
            {"sheet_id": sheet['id'], "date": {"$gte": start_date, "$lt": end_date}},
            {"_id": 0}
        ).sort("date", 1).to_list(1000)
        
        headers = ["Date"]
        for m in meters:
            headers.extend([
                f"{m['name']} Initial",
                f"{m['name']} Final",
                f"{m['name']} MF",
                f"{m['name']} Consumption"
            ])
        headers.append("Total Consumption")
        
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        for entry in entries:
            row = [format_date(entry['date'])]
            readings_map = {r['meter_id']: r for r in entry['readings']}
            
            for m in meters:
                r = readings_map.get(m['id'])
                if r:
                    row.extend([r['initial'], r['final'], m['mf'], r['consumption']])
                else:
                    row.extend(['-', '-', m['mf'], '-'])
            
            row.append(entry['total_consumption'])
            ws.append(row)
            
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
    return wb

@api_router.get("/energy/export-all/{year}/{month}")
async def export_all_energy_sheets(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    wb = await _generate_energy_export_wb(year, month)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Energy_Consumption_{month}-{year}.xlsx"}
    )

@api_router.get("/energy/export/{sheet_id}/{year}/{month}")
async def export_energy_sheet(
    sheet_id: str,
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    sheet = await db.energy_sheets.find_one({"id": sheet_id}, {"_id": 0})
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
        
    meters = await db.energy_meters.find({"sheet_id": sheet_id}, {"_id": 0}).to_list(100)
    
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
        
    entries = await db.energy_entries.find(
        {"sheet_id": sheet_id, "date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{sheet['name']} - {month}-{year}"
    
    # Headers
    headers = ["Date"]
    for m in meters:
        headers.extend([
            f"{m['name']} Initial",
            f"{m['name']} Final",
            f"{m['name']} MF",
            f"{m['name']} Consumption"
        ])
    headers.append("Total Consumption")
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    # Data
    for entry in entries:
        row = [format_date(entry['date'])]
        readings_map = {r['meter_id']: r for r in entry['readings']}
        
        for m in meters:
            r = readings_map.get(m['id'])
            if r:
                row.extend([r['initial'], r['final'], m['mf'], r['consumption']])
            else:
                row.extend([0, 0, m['mf'], 0])
                
        row.append(entry['total_consumption'])
        ws.append(row)
        
    # Auto-width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{sheet['name']}_{year}_{month:02d}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.put("/energy/entries/{entry_id}", response_model=EnergyEntry)
async def update_energy_entry(
    entry_id: str,
    entry_input: EnergyEntryInput,
    current_user: User = Depends(get_current_user)
):
    # Verify entry exists
    existing = await db.energy_entries.find_one({"id": entry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Entry not found")

    # Get previous day's entry for initials
    date_obj = datetime.strptime(entry_input.date, "%Y-%m-%d")
    prev_date = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
    prev_entry = await db.energy_entries.find_one(
        {"sheet_id": entry_input.sheet_id, "date": prev_date},
        {"_id": 0}
    )
    
    prev_finals = {}
    if prev_entry:
        for r in prev_entry['readings']:
            prev_finals[r['meter_id']] = r['final']
            
    # Calculate readings
    readings = []
    total_consumption = 0
    
    for r_in in entry_input.readings:
        meter = await db.energy_meters.find_one({"id": r_in.meter_id})
        if not meter:
            continue
            
        initial = prev_finals.get(r_in.meter_id, 0.0)
        consumption = (r_in.final - initial) * meter['mf']
        total_consumption += consumption
        
        readings.append(EnergyReading(
            meter_id=r_in.meter_id,
            initial=initial,
            final=r_in.final,
            consumption=consumption
        ))
        
    # Update entry
    entry_data = EnergyEntry(
        id=entry_id, # Preserve ID
        sheet_id=entry_input.sheet_id,
        date=entry_input.date,
        readings=readings,
        total_consumption=total_consumption,
        created_at=existing['created_at'], # Preserve created_at
        updated_at=datetime.now(timezone.utc)
    )
    
    doc = entry_data.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.energy_entries.replace_one({"id": entry_id}, doc)

    # Update next day's initial values if exists
    date_obj = datetime.strptime(entry_input.date, "%Y-%m-%d")
    next_date = (date_obj + timedelta(days=1)).strftime("%Y-%m-%d")
    next_entry = await db.energy_entries.find_one(
        {"sheet_id": entry_input.sheet_id, "date": next_date}
    )
    
    if next_entry:
        # Create a map of current finals
        current_finals = {r.meter_id: r.final for r in readings}
        
        next_readings = []
        next_total_consumption = 0
        
        for r in next_entry['readings']:
            if r['meter_id'] in current_finals:
                new_initial = current_finals[r['meter_id']]
                
                # Fetch meter to get MF
                meter = await db.energy_meters.find_one({"id": r['meter_id']})
                if meter:
                    new_consumption = (r['final'] - new_initial) * meter['mf']
                    r['initial'] = new_initial
                    r['consumption'] = new_consumption
            
            next_readings.append(r)
            next_total_consumption += r.get('consumption', 0)
            
        next_entry['readings'] = next_readings
        next_entry['total_consumption'] = next_total_consumption
        next_entry['updated_at'] = datetime.now(timezone.utc).isoformat()
        
        await db.energy_entries.replace_one({"_id": next_entry['_id']}, next_entry)
        
    return entry_data

@api_router.delete("/energy/entries/{entry_id}")
async def delete_energy_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    result = await db.energy_entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted successfully"}

async def get_boundary_meter_data(year: int, month: int):
    # 1. Find 33KV Sheet
    sheet = await db.energy_sheets.find_one({"name": "33KV"})
    if not sheet:
        raise HTTPException(status_code=404, detail="33KV Sheet not found")
    
    # 2. Find Meters (Donthapally & Kandi)
    meters_cursor = db.energy_meters.find({"sheet_id": sheet['id']})
    meters = await meters_cursor.to_list(100)
    
    target_meters = []
    # Map for easy lookup and to preserve order
    # User Order: 1. Donthapally, 2. Kandi
    donthapally = next((m for m in meters if "Donthapally" in m['name']), None)
    kandi = next((m for m in meters if "Kandi" in m['name']), None)
    
    if donthapally:
        target_meters.append({
            "meter": donthapally, 
            "display_name": "33KV Donthanpally\nservice no RRS:1445\nSerial no:14754097"
        })
    if kandi:
        target_meters.append({
            "meter": kandi, 
            "display_name": "33KV Kandi(BDL)\nservice no RRS:1445\nSerial no:04932269"
        })
        
    if not target_meters:
         raise HTTPException(status_code=404, detail="Target meters (Donthapally/Kandi) not found")

    # 3. Date Calculations
    # Current Month Start and End
    import calendar
    last_day = calendar.monthrange(year, month)[1]
    month_start_date = f"{year}-{month:02d}-01"
    month_end_date = f"{year}-{month:02d}-{last_day}"
    
    # Previous Month Last Day (for display)
    first_date_obj = datetime(year, month, 1)
    prev_month_last_date_obj = first_date_obj - timedelta(days=1)
    prev_month_str = prev_month_last_date_obj.strftime("%d-%b-%y")
    current_month_end_str = datetime(year, month, last_day).strftime("%d-%b-%y")
    
    # 4. Fetch Readings
    # We need:
    # Initial: Initial reading of the first day of the month (or Final of prev month last day)
    # Final: Final reading of the last day of the month
    
    # Fetch first day entry
    start_entry = await db.energy_entries.find_one({
        "sheet_id": sheet['id'],
        "date": month_start_date
    })
    
    # Fetch last day entry
    end_entry = await db.energy_entries.find_one({
        "sheet_id": sheet['id'],
        "date": month_end_date
    })
    
    report_data = []
    
    for item in target_meters:
        meter = item['meter']
        meter_id = meter['id']
        mf = meter['mf']
        
        initial_val = 0.0
        final_val = 0.0
        
        # Get Initial
        if start_entry:
            reading = next((r for r in start_entry['readings'] if r['meter_id'] == meter_id), None)
            if reading:
                initial_val = reading['initial']
                
        # Get Final
        if end_entry:
            reading = next((r for r in end_entry['readings'] if r['meter_id'] == meter_id), None)
            if reading:
                final_val = reading['final']
        
        diff = final_val - initial_val
        consumption = diff * mf
        
        report_data.append({
            "name": item['display_name'],
            "initial": initial_val,
            "final": final_val,
            "diff": diff,
            "mf": mf,
            "consumption": consumption
        })
        
    return {
        "report_data": report_data,
        "prev_month_str": prev_month_str,
        "current_month_end_str": current_month_end_str,
        "month_name": calendar.month_name[month]
    }

@api_router.get("/reports/status/{year}/{month}")
async def check_reports_status(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    import calendar
    
    try:
        # 1. Calculate Expected Counts
        last_day = calendar.monthrange(year, month)[1]
        days_in_month = last_day # Number of days
        
        start_date = f"{year}-{month:02d}-01"
        end_date = f"{year}-{month:02d}-{last_day}"
        
        # 2. Check Line Losses (entries)
        # Count active feeders
        num_feeders = await db.feeders.count_documents({})
        expected_entries = num_feeders * days_in_month
        
        actual_entries = await db.entries.count_documents({
            "date": {"$gte": start_date, "$lte": end_date}
        })
        
        line_losses_ready = actual_entries >= expected_entries
        
        # 3. Check Max Min Reports (max_min_entries)
        # For report readiness we only require that each day of the month
        # has at least one Max–Min entry (for any feeder), not that
        # every feeder has data for every single day.
        max_min_dates = await db.max_min_entries.distinct(
            "date",
            {"date": {"$gte": start_date, "$lte": end_date}},
        )
        max_min_ready = len(max_min_dates) >= days_in_month
        
        # 4. Check Boundary Meter (energy_entries - 33KV)
        sheet_33kv = await db.energy_sheets.find_one({"name": "33KV"})
        boundary_meter_ready = False
        if sheet_33kv:
            actual_energy_entries = await db.energy_entries.count_documents({
                "sheet_id": sheet_33kv['id'],
                "date": {"$gte": start_date, "$lte": end_date}
            })
            boundary_meter_ready = actual_energy_entries >= days_in_month
        
        interruptions_count = await db.interruption_entries.count_documents({
            "date": {"$gte": start_date, "$lte": end_date}
        })
        interruptions_ready = interruptions_count > 0
            
        # 5. Compile Status
        missing_reports = []
        
        if not line_losses_ready:
            missing_reports.append("Line Losses")
            missing_reports.append("New Line Losses Report")
        
        if not max_min_ready:
            # These all depend on max_min data
            missing_reports.append("Fortnight")
            missing_reports.append("Daily Max MVA (SAP)")
            missing_reports.append("KPI")
            missing_reports.append("PTR Max–Min (Format-1)")
            missing_reports.append("TL Max Loading (Format-4)")
            
        if not boundary_meter_ready:
            missing_reports.append("Boundary Meter Reading (33KV)")
        
        if not interruptions_ready:
            missing_reports.append("Interruptions")
            missing_reports.append("MIS Interruption Details")
            
        all_ready = len(missing_reports) == 0
        
        return {
            "all_ready": all_ready,
            "missing_reports": missing_reports
        }

    except Exception as e:
        print(f"Error checking report status: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/reports/boundary-meter-33kv/data/{year}/{month}")
async def get_boundary_meter_report_json(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    return await get_boundary_meter_data(year, month)

async def _generate_boundary_meter_wb(year: int, month: int):
    data = await get_boundary_meter_data(year, month)
    report_data = data['report_data']
    prev_month_str = data['prev_month_str']
    current_month_end_str = data['current_month_end_str']
    month_name = data['month_name']

    # 5. Generate Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Boundary Meter Report"
    
    # Styling
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(bold=True)
    
    # Title Row
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Boundary Meter Readings of 400KV Shankarpally for the Month of {month_name[:3]}'{str(year)[-2:]}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_align
    title_cell.border = thin_border
    
    # Sub-header
    ws.merge_cells('A2:H2')
    ws['A2'].value = "33KV Station Supply"
    ws['A2'].alignment = center_align
    ws['A2'].font = bold_font
    ws['A2'].border = thin_border
    
    # Header Rows (3 & 4)
    # Sl. No. (A3:A4)
    ws.merge_cells('A3:A4')
    ws['A3'] = "Sl.\nNo."
    
    # Name of Feeder (B3:B4)
    ws.merge_cells('B3:B4')
    ws['B3'] = "Name of Feeder"
    
    # Timing (C3:D3)
    ws.merge_cells('C3:D3')
    ws['C3'] = "Timing"
    ws['C4'] = "From"
    ws['D4'] = "To"
    
    # Initial Reading (E3:E4)
    ws.merge_cells('E3:E4')
    ws['E3'] = "Initial\nReading"
    
    # Final Reading (F3:F4)
    ws.merge_cells('F3:F4')
    ws['F3'] = "Final\nReading"
    
    # Diff (G3:G4)
    ws.merge_cells('G3:G4')
    ws['G3'] = "Diff."
    
    # M.F. (H3:H4)
    ws.merge_cells('H3:H4')
    ws['H3'] = "M.F."
    
    # Consumption (I3:I4) -> Wait, user image has "Consumption" then "Remarks".
    # User image columns: Sl No, Name, Timing(From, To), Initial, Final, Diff, MF, Consumption, Remarks
    # My Columns: A, B, C, D, E, F, G, H, I, J
    # Let's re-map:
    # A: Sl No
    # B: Name
    # C: Timing From
    # D: Timing To
    # E: Initial
    # F: Final
    # G: Diff
    # H: MF
    # I: Consumption
    # J: Remarks
    
    # Adjust Merges
    ws.unmerge_cells('A1:H1')
    ws.merge_cells('A1:J1')
    ws.unmerge_cells('A2:H2')
    ws.merge_cells('A2:J2')
    
    ws.merge_cells('I3:I4')
    ws['I3'] = "Consumption"
    
    ws.merge_cells('J3:J4')
    ws['J3'] = "Remarks"
    
    # Apply styles to header
    for row in ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = center_align
            cell.font = bold_font
            cell.border = thin_border
            
    # Data Rows
    start_row = 5
    for idx, data in enumerate(report_data, 1):
        ws[f'A{start_row}'] = idx
        ws[f'B{start_row}'] = data['name']
        ws[f'C{start_row}'] = f"{prev_month_str}\n12:00 Hrs"
        ws[f'D{start_row}'] = f"{current_month_end_str}\n12:00 Hrs"
        ws[f'E{start_row}'] = data['initial']
        ws[f'F{start_row}'] = data['final']
        ws[f'G{start_row}'] = data['diff']
        ws[f'H{start_row}'] = data['mf']
        ws[f'I{start_row}'] = data['consumption']
        ws[f'J{start_row}'] = "-"
        
        # Style row
        for col in range(1, 11):
            cell = ws.cell(row=start_row, column=col)
            cell.alignment = center_align
            cell.border = thin_border
            # Number formatting
            if col in [5, 6, 7, 9]: # Readings, Diff, Consumption
                cell.number_format = '0.00'
                
        # Row height for multiline text
        ws.row_dimensions[start_row].height = 60
        start_row += 1
        
    # Column Widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10

    return wb

@api_router.get("/reports/boundary-meter-33kv/{year}/{month}")
async def generate_boundary_meter_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        import calendar
        from io import BytesIO
        
        # We need month_name for the filename, but it's not directly returned by the wb function.
        # However, we can re-calculate it or fetch it.
        # But _generate_boundary_meter_wb calculates it internally.
        # Maybe we should return (wb, filename) from the internal function?
        # Or just recalculate it here.
        
        wb = await _generate_boundary_meter_wb(year, month)
        month_name = calendar.month_name[month]
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Boundary_Meter_Report_{month_name}_{year}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers=headers
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": str(e)})



# --- KPI Report Logic ---

KPI_FEEDER_DETAILS = {
    "400KV MAHESHWARAM-2": {"length_km": 84.13, "length_ckm": 84.13, "conductor": "ACSR Twin Moose", "capacity": 1786},
    "400KV MAHESHWARAM-1": {"length_km": 84.13, "length_ckm": 84.13, "conductor": "ACSR Twin Moose", "capacity": 1786},
    "400KV NARSAPUR-1": {"length_km": 60.92, "length_ckm": 60.92, "conductor": "ACSR Twin Moose", "capacity": 1786},
    "400KV NARSAPUR-2": {"length_km": 60.92, "length_ckm": 60.92, "conductor": "ACSR Twin Moose", "capacity": 1786},
    "400KV KETHIREDDYPALLY-1": {"length_km": 51.13, "length_ckm": 51.13, "conductor": "ACSR Quad Moose", "capacity": 3572},
    "400KV KETHIREDDYPALLY-2": {"length_km": 51.13, "length_ckm": 51.13, "conductor": "ACSR Quad Moose", "capacity": 3572},
    "400KV NIZAMABAD-1": {"length_km": 139, "length_ckm": 139, "conductor": "ACSR Twin Moose", "capacity": 1786},
    "400KV NIZAMABAD-2": {"length_km": 139, "length_ckm": 139, "conductor": "ACSR Twin Moose", "capacity": 1786},
    "220KV PARIGI-1": {"length_km": 45.05, "length_ckm": 45.05, "conductor": "ACSR Moose", "capacity": 795},
    "220KV PARIGI-2": {"length_km": 45.05, "length_ckm": 45.05, "conductor": "ACSR Moose", "capacity": 795},
    "220KV THANDUR": {"length_km": 85.057, "length_ckm": 85.057, "conductor": "ACSR Moose", "capacity": 795},
    "220KV GACHIBOWLI-1": {"length_km": 28.57, "length_ckm": 28.57, "conductor": "HTLS", "capacity": 1600},
    "220KV GACHIBOWLI-2": {"length_km": 23.3, "length_ckm": 23.3, "conductor": "HTLS", "capacity": 1600},
    "220KV KETHIREDDYPALLY": {"length_km": 29.44, "length_ckm": 29.44, "conductor": "ACSR Moose", "capacity": 795},
    "220KV YEDDUMAILARAM-1": {"length_km": 10.3, "length_ckm": 10.3, "conductor": "ACSR Moose", "capacity": 795},
    "220KV YEDDUMAILARAM-2": {"length_km": 10.2, "length_ckm": 10.2, "conductor": "ACSR Zebra", "capacity": 795},
    "220KV SADASIVAPET-1": {"length_km": 37.151, "length_ckm": 37.151, "conductor": "ACSR Zebra", "capacity": 795},
    "220KV SADASIVAPET-2": {"length_km": 37.151, "length_ckm": 37.151, "conductor": "ACSR Zebra", "capacity": 795},
}

FEEDER_ORDER_KPI = [
    "400KV MAHESHWARAM-2", "400KV MAHESHWARAM-1", 
    "400KV NARSAPUR-1", "400KV NARSAPUR-2",
    "400KV KETHIREDDYPALLY-1", "400KV KETHIREDDYPALLY-2",
    "400KV NIZAMABAD-1", "400KV NIZAMABAD-2",
    "220KV PARIGI-1", "220KV PARIGI-2",
    "220KV THANDUR",
    "220KV GACHIBOWLI-1", "220KV GACHIBOWLI-2",
    "220KV KETHIREDDYPALLY",
    "220KV YEDDUMAILARAM-1", "220KV YEDDUMAILARAM-2",
    "220KV SADASIVAPET-1", "220KV SADASIVAPET-2"
]

ICT_ORDER_KPI = ["ICT-1 (315MVA)", "ICT-2 (315MVA)", "ICT-3 (315MVA)", "ICT-4 (500MVA)"]

def calculate_kpi_stats(entries, feeder_type):
    if not entries:
        return {"avg_val": 0, "max_val": 0}
        
    total_avg = 0
    max_val = 0
    count = 0
    
    # For Lines: track Max MW to determine which Amps to pick
    max_mw_found = -1.0
    
    for e in entries:
        d = e.get('data', {})
        if not d: continue
        
        max_data = d.get('max') or {}
        avg_data = d.get('avg') or {}
        
        if feeder_type == 'ict_feeder':
            # ICT: Max Demand (MW) and Avg Load (MW)
            # Max MW comes from max.mw
            # Avg MW comes from avg.mw
            curr_max = float(max_data.get('mw', 0) or 0)
            curr_avg = float(avg_data.get('mw', 0) or 0)
            
            if curr_max > max_val:
                max_val = curr_max
            
            if curr_avg > 0:
                total_avg += curr_avg
                count += 1
        else:
            # Line: Max Line Loading (Amps) and Avg Loading (Amps)
            # Logic Update: Max Amps should be derived from the entry with Max MW
            curr_mw = float(max_data.get('mw', 0) or 0)
            curr_amps = float(max_data.get('amps', 0) or 0)
            
            if curr_mw > max_mw_found:
                max_mw_found = curr_mw
                max_val = curr_amps
            
            curr_avg = float(avg_data.get('amps', 0) or 0)
            if curr_avg > 0:
                total_avg += curr_avg
                count += 1
                
    avg_val = total_avg / count if count > 0 else 0
    return {"avg_val": avg_val, "max_val": max_val}

@api_router.get("/reports/kpi/preview/{year}/{month}")
async def get_kpi_preview(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        import calendar
        
        # Date Range
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"
            
        # 1. Lines Data
        lines_data = []
        feeders = await db.max_min_feeders.find({"type": {"$in": ["feeder_400kv", "feeder_220kv"]}}, {"_id": 0}).to_list(100)
        
        # Sort feeders
        feeders.sort(key=lambda x: FEEDER_ORDER_KPI.index(x['name']) if x['name'] in FEEDER_ORDER_KPI else 999)
        
        all_entries = await db.max_min_entries.find(
            {"date": {"$gte": start_date, "$lt": end_date}},
            {"_id": 0}
        ).to_list(5000)
        
        entries_by_feeder = {}
        for e in all_entries:
            if e['feeder_id'] not in entries_by_feeder:
                entries_by_feeder[e['feeder_id']] = []
            entries_by_feeder[e['feeder_id']].append(e)
            
        for idx, f in enumerate(feeders):
            if f['name'] not in FEEDER_ORDER_KPI: continue # Only show defined feeders
            
            details = KPI_FEEDER_DETAILS.get(f['name'], {})
            entries = entries_by_feeder.get(f['id'], [])
            stats = calculate_kpi_stats(entries, f['type'])
            
            avg_load = stats['avg_val']
            max_load = stats['max_val']
            pct_load = (avg_load / max_load * 100) if max_load > 0 else 0
            
            lines_data.append({
                "sl_no": idx + 1,
                "zone": "Metro Zone",
                "circle": "Metro- West",
                "feeder_name": f['name'],
                "length_km": details.get('length_km', '-'),
                "length_ckm": details.get('length_ckm', '-'),
                "conductor": details.get('conductor', '-'),
                "capacity": details.get('capacity', '-'),
                "avg_loading": f"{avg_load:.2f}" if avg_load else "-",
                "max_loading": f"{max_load:.2f}" if max_load else "-",
                "pct_loading": f"{pct_load:.2f}" if pct_load else "-"
            })
            
        # 2. ICT Data
        ict_data = []
        ict_feeders = await db.max_min_feeders.find({"type": "ict_feeder"}, {"_id": 0}).to_list(100)
        ict_feeders.sort(key=lambda x: ICT_ORDER_KPI.index(x['name']) if x['name'] in ICT_ORDER_KPI else 999)
        
        for idx, f in enumerate(ict_feeders):
            entries = entries_by_feeder.get(f['id'], [])
            stats = calculate_kpi_stats(entries, f['type'])
            
            avg_load = stats['avg_val']
            max_demand = stats['max_val']
            pct_load = (avg_load / max_demand * 100) if max_demand > 0 else 0
            
            # Format Name: ICT-1 (315MVA) -> 315 MVA ICT-1
            name_parts = f['name'].split(' ')
            formatted_name = f['name']
            if len(name_parts) >= 2:
                 try:
                     ict_part = name_parts[0] # ICT-1
                     mva_part = name_parts[1].replace('(', '').replace(')', '') # 315MVA
                     formatted_name = f"{mva_part[:3]} MVA {ict_part}" # 315 MVA ICT-1
                     if "500" in mva_part: formatted_name = "500MVA ICT-4"
                 except:
                     pass
            
            ict_data.append({
                "sl_no": idx + 1,
                "zone": "Merto Zone",
                "circle": "OMC- Metro- West Circle",
                "substation": "400KV Shankarpally",
                "capacity_name": formatted_name,
                "proposed": "-",
                "max_demand": f"{max_demand:.2f}" if max_demand else "-",
                "avg_load": f"{avg_load:.2f}" if avg_load else "-",
                "pct_loading": f"{pct_load:.2f}" if pct_load else "-"
            })
            
        return {"lines": lines_data, "icts": ict_data}
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"detail": str(e)})

async def _generate_kpi_report_wb(year: int, month: int):
    import calendar
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter
    
    # --- Common Styles ---
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(bold=True)
    
    wb = Workbook()
    
    # --- Data Prep ---
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"
    
    month_name = calendar.month_name[month]
    
    all_entries = await db.max_min_entries.find(
        {"date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0}
    ).to_list(5000)
    
    entries_by_feeder = {}
    for e in all_entries:
        if e['feeder_id'] not in entries_by_feeder:
            entries_by_feeder[e['feeder_id']] = []
        entries_by_feeder[e['feeder_id']].append(e)

    # ================= SHEET 1: Over Loading of Lines =================
    ws1 = wb.active
    ws1.title = "Over Loading of Lines"
    
    # Header Rows
    ws1.merge_cells('A1:L1')
    c = ws1.cell(row=1, column=1, value="400/220KV SHANKARPALLY SUBSTATION")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws1.merge_cells('A2:L2')
    c = ws1.cell(row=2, column=1, value=f"STATEMENT 20: FOR THE MONTH OF {month_name}-{year}")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws1.merge_cells('A3:L3')
    c = ws1.cell(row=3, column=1, value="Overloading of Lines")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    headers = [
        "Sl. No.", "Name of Zone", "Circle", "Name of feeder", 
        "Length in KM (Total Length)", "Length in CKM", "Type of Conductor",
        "Current Carrying Capacity", "Avg.Loading (Amps)", "Max.Line loading (Amps)",
        "% line loading", "Remarks"
    ]
    
    for i, h in enumerate(headers):
        cell = ws1.cell(row=4, column=i+1, value=h)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        
    feeders = await db.max_min_feeders.find({"type": {"$in": ["feeder_400kv", "feeder_220kv"]}}, {"_id": 0}).to_list(100)
    feeders.sort(key=lambda x: FEEDER_ORDER_KPI.index(x['name']) if x['name'] in FEEDER_ORDER_KPI else 999)
    
    row_idx = 5
    sl_no = 1
    for f in feeders:
        if f['name'] not in FEEDER_ORDER_KPI: continue
        
        details = KPI_FEEDER_DETAILS.get(f['name'], {})
        entries = entries_by_feeder.get(f['id'], [])
        stats = calculate_kpi_stats(entries, f['type'])
        
        avg_load = stats['avg_val']
        max_load = stats['max_val']
        pct_load = (avg_load / max_load * 100) if max_load > 0 else 0
        
        row_data = [
            sl_no,
            "Metro Zone",
            "Metro- West",
            f['name'],
            details.get('length_km', '-'),
            details.get('length_ckm', '-'),
            details.get('conductor', '-'),
            details.get('capacity', '-'),
            f"{avg_load:.2f}" if avg_load else "-",
            f"{max_load:.2f}" if max_load else "-",
            f"{pct_load:.2f}" if pct_load else "-",
            "-"
        ]
        
        for i, val in enumerate(row_data):
            cell = ws1.cell(row=row_idx, column=i+1, value=val)
            cell.alignment = center_align
            cell.border = thin_border
            
        row_idx += 1
        sl_no += 1
        
    for i in range(1, ws1.max_column + 1):
        col_letter = get_column_letter(i)
        max_len = 0
        for cell in ws1[col_letter]:
            if cell.row <= 3: continue
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except: pass
        ws1.column_dimensions[col_letter].width = max(max_len + 2, 10)


    # ================= SHEET 2: ICT'S =================
    ws2 = wb.create_sheet(title="ICT'S")
    
    ws2.merge_cells('A1:J1')
    c = ws2.cell(row=1, column=1, value="400/220KV SHANKARAPALLY")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws2.merge_cells('A2:J2')
    c = ws2.cell(row=2, column=1, value="ANNEXURE XVIII")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws2.merge_cells('A3:J3')
    c = ws2.cell(row=3, column=1, value="REDUCTION OF TRANSMISSION LINE FORMATS")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws2.merge_cells('A4:J4')
    c = ws2.cell(row=4, column=1, value=f"(A) Details of overloading of PTRs (70% and above) For {month_name}-{year}")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    headers = [
        "Sl.No", "Name of Zone", "TL&ss Circle", "Name of Substation", 
        "Existing ICT capacity (MVA)", "Proposed augmentation of ICT capacity",
        "Max. Demand reached during the month (MW)", "Average load in MW",
        "Average Percentage of ICT loading", "Remarks"
    ]
    
    for i, h in enumerate(headers):
        cell = ws2.cell(row=5, column=i+1, value=h)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        
    ict_feeders = await db.max_min_feeders.find({"type": "ict_feeder"}, {"_id": 0}).to_list(100)
    ict_feeders.sort(key=lambda x: ICT_ORDER_KPI.index(x['name']) if x['name'] in ICT_ORDER_KPI else 999)
    
    row_idx = 6
    sl_no = 1
    for f in ict_feeders:
        entries = entries_by_feeder.get(f['id'], [])
        stats = calculate_kpi_stats(entries, f['type'])
        
        avg_load = stats['avg_val']
        max_demand = stats['max_val']
        pct_load = (avg_load / max_demand * 100) if max_demand > 0 else 0
        
        name_parts = f['name'].split(' ')
        formatted_name = f['name']
        if len(name_parts) >= 2:
                try:
                    ict_part = name_parts[0] 
                    mva_part = name_parts[1].replace('(', '').replace(')', '')
                    formatted_name = f"{mva_part[:3]} MVA {ict_part}"
                    if "500" in mva_part: formatted_name = "500MVA ICT-4"
                except: pass
        
        row_data = [
            sl_no,
            "Merto Zone", 
            "OMC- Metro- West Circle",
            "400KV Shankarpally",
            formatted_name,
            "-",
            f"{max_demand:.2f}" if max_demand else "-",
            f"{avg_load:.2f}" if avg_load else "-",
            f"{pct_load:.2f}" if pct_load else "-",
            "-"
        ]
        
        for i, val in enumerate(row_data):
            cell = ws2.cell(row=row_idx, column=i+1, value=val)
            cell.alignment = center_align
            cell.border = thin_border
            
        row_idx += 1
        sl_no += 1

    for i in range(1, ws2.max_column + 1):
        col_letter = get_column_letter(i)
        max_len = 0
        for cell in ws2[col_letter]:
            if cell.row <= 4: continue
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except: pass
        ws2.column_dimensions[col_letter].width = max(max_len + 2, 12)
        
    return wb

@api_router.get("/reports/kpi/export/{year}/{month}")
async def export_kpi_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        import calendar
        wb = await _generate_kpi_report_wb(year, month)
        month_name = calendar.month_name[month]

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"KPI_Report_{month_name}_{year}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        with open("export_error_log.txt", "w") as f:
            f.write(error_msg)
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": str(e)})

@api_router.get("/reports/interruptions/preview/{year}/{month}")
async def preview_interruptions_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        data = await _get_interruptions_report_data(year, month)
        return data
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": str(e)})


# --- Line Losses Report ---

async def _get_line_losses_report_data(year: int, month: int) -> List[Dict[str, Any]]:
    import calendar

    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"

    all_feeders = await db.feeders.find({}, {"_id": 0}).to_list(100)

    FEEDER_MAPPING = {
        "400 KV Shankarpally-MHRM-2": "400KV MAHESHWARAM-2",
        "400 KV Shankarpally-MHRM-1": "400KV MAHESHWARAM-1",
        "400 KV Shankarpally-Narsapur-1": "400KV NARSAPUR-1",
        "400 KV Shankarpally-Narsapur-2": "400KV NARSAPUR-2",
        "400 KV KethiReddyPally-1": "400KV KETHIREDDYPALLY-1",
        "400 KV KethiReddyPally-2": "400KV KETHIREDDYPALLY-2",
        "400 KV Nizamabad-1&2": "400KV NIZAMABAD-1 & 2",
        "220 KV Parigi-1": "220KV PARIGI-1",
        "220 KV Parigi-2": "220KV PARIGI-2",
        "220 KV Tandur": "220KV THANDUR",
        "220 KV Gachibowli-1": "220KV GACHIBOWLI-1",
        "220 KV Gachibowli-2": "220KV GACHIBOWLI-2",
        "220 KV KethiReddyPally": "220KV KETHIREDDYPALLY",
        "220 KV Yeddumailaram-1": "220KV YEDDUMAILARAM-1",
        "220 KV Yeddumailaram-2": "220KV YEDDUMAILARAM-2",
        "220 KV Sadasivapet-1": "220KV SADASIVAPET-1",
        "220 KV Sadasivapet-2": "220KV SADASIVAPET-2",
    }

    FEEDER_ORDER = [
        "400KV MAHESHWARAM-2",
        "400KV MAHESHWARAM-1",
        "400KV NARSAPUR-1",
        "400KV NARSAPUR-2",
        "400KV KETHIREDDYPALLY-1",
        "400KV KETHIREDDYPALLY-2",
        "400KV NIZAMABAD-1 & 2",
        "220KV PARIGI-1",
        "220KV PARIGI-2",
        "220KV THANDUR",
        "220KV GACHIBOWLI-1",
        "220KV GACHIBOWLI-2",
        "220KV KETHIREDDYPALLY",
        "220KV YEDDUMAILARAM-1",
        "220KV YEDDUMAILARAM-2",
        "220KV SADASIVAPET-1",
        "220KV SADASIVAPET-2",
    ]

    target_feeders: List[Dict[str, Any]] = []
    for f in all_feeders:
        if f.get("name") in FEEDER_MAPPING:
            f["display_name"] = FEEDER_MAPPING[f["name"]]
            target_feeders.append(f)

    target_feeders.sort(
        key=lambda x: FEEDER_ORDER.index(x["display_name"])
        if x["display_name"] in FEEDER_ORDER
        else 999
    )

    entries = await db.entries.find(
        {"date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0},
    ).to_list(5000)

    entries_by_feeder: Dict[str, List[Dict[str, Any]]] = {}
    for e in entries:
        fid = e.get("feeder_id")
        if not fid:
            continue
        if fid not in entries_by_feeder:
            entries_by_feeder[fid] = []
        entries_by_feeder[fid].append(e)

    report_data: List[Dict[str, Any]] = []
    for idx, f in enumerate(target_feeders):
        f_entries = entries_by_feeder.get(f["id"], [])
        f_entries.sort(key=lambda x: x["date"])

        data: Dict[str, Any] = {
            "sl_no": idx + 1,
            "feeder_name": f["display_name"],
            "shankarpally": {"import": {}, "export": {}},
            "other_end": {"import": {}, "export": {}},
            "stats": {},
        }

        if f_entries:
            first = f_entries[0]
            last = f_entries[-1]

            s_imp_init = first.get("end1_import_initial", 0)
            s_imp_final = last.get("end1_import_final", 0)
            s_imp_mf = f.get("end1_import_mf", 1)
            s_imp_cons = (s_imp_final - s_imp_init) * s_imp_mf

            data["shankarpally"]["import"] = {
                "initial": s_imp_init,
                "final": s_imp_final,
                "mf": s_imp_mf,
                "consumption": s_imp_cons,
            }

            s_exp_init = first.get("end1_export_initial", 0)
            s_exp_final = last.get("end1_export_final", 0)
            s_exp_mf = f.get("end1_export_mf", 1)
            s_exp_cons = (s_exp_final - s_exp_init) * s_exp_mf

            data["shankarpally"]["export"] = {
                "initial": s_exp_init,
                "final": s_exp_final,
                "mf": s_exp_mf,
                "consumption": s_exp_cons,
            }

            o_imp_init = first.get("end2_import_initial", 0)
            o_imp_final = last.get("end2_import_final", 0)
            o_imp_mf = f.get("end2_import_mf", 1)
            o_imp_cons = (o_imp_final - o_imp_init) * o_imp_mf

            data["other_end"]["import"] = {
                "initial": o_imp_init,
                "final": o_imp_final,
                "mf": o_imp_mf,
                "consumption": o_imp_cons,
            }

            o_exp_init = first.get("end2_export_initial", 0)
            o_exp_final = last.get("end2_export_final", 0)
            o_exp_mf = f.get("end2_export_mf", 1)
            o_exp_cons = (o_exp_final - o_exp_init) * o_exp_mf

            data["other_end"]["export"] = {
                "initial": o_exp_init,
                "final": o_exp_final,
                "mf": o_exp_mf,
                "consumption": o_exp_cons,
            }

            D = s_imp_cons
            L = o_imp_cons
            H = s_exp_cons
            P = o_exp_cons
            numerator = (D + L) - (H + P)
            denominator = D + L
            pct_loss = (numerator / denominator * 100) if denominator != 0 else 0
            data["stats"]["pct_loss"] = pct_loss
        else:
            for end in ["shankarpally", "other_end"]:
                for type_ in ["import", "export"]:
                    data[end][type_] = {
                        "initial": 0,
                        "final": 0,
                        "mf": 0,
                        "consumption": 0,
                    }
            data["stats"]["pct_loss"] = 0

        report_data.append(data)

    return report_data


@api_router.get("/reports/line-losses/preview/{year}/{month}")
async def get_line_losses_report_preview(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user),
):
    try:
        data = await _get_line_losses_report_data(year, month)
        return data
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        with open("preview_error_log.txt", "w") as f:
            f.write(error_msg)
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": str(e)})

@api_router.get("/reports/line-losses/daily-preview/{date_str}")
async def get_line_losses_daily_report_preview(
    date_str: str,
    current_user: User = Depends(get_current_user)
):
    try:
        # 1. Get Feeders (Line Losses Module)
        all_feeders = await db.feeders.find({}, {"_id": 0}).to_list(100)
        
        # Mapping DB Names to Report Names
        FEEDER_MAPPING = {
            "400 KV Shankarpally-MHRM-2": "400KV MAHESHWARAM-2",
            "400 KV Shankarpally-MHRM-1": "400KV MAHESHWARAM-1",
            "400 KV Shankarpally-Narsapur-1": "400KV NARSAPUR-1",
            "400 KV Shankarpally-Narsapur-2": "400KV NARSAPUR-2",
            "400 KV KethiReddyPally-1": "400KV KETHIREDDYPALLY-1",
            "400 KV KethiReddyPally-2": "400KV KETHIREDDYPALLY-2",
            "400 KV Nizamabad-1&2": "400KV NIZAMABAD-1 & 2",
            "220 KV Parigi-1": "220KV PARIGI-1",
            "220 KV Parigi-2": "220KV PARIGI-2",
            "220 KV Tandur": "220KV THANDUR",
            "220 KV Gachibowli-1": "220KV GACHIBOWLI-1",
            "220 KV Gachibowli-2": "220KV GACHIBOWLI-2",
            "220 KV KethiReddyPally": "220KV KETHIREDDYPALLY",
            "220 KV Yeddumailaram-1": "220KV YEDDUMAILARAM-1",
            "220 KV Yeddumailaram-2": "220KV YEDDUMAILARAM-2",
            "220 KV Sadasivapet-1": "220KV SADASIVAPET-1",
            "220 KV Sadasivapet-2": "220KV SADASIVAPET-2"
        }

        # Filter and Sort
        FEEDER_ORDER = [
            "400KV MAHESHWARAM-2",
            "400KV MAHESHWARAM-1",
            "400KV NARSAPUR-1",
            "400KV NARSAPUR-2",
            "400KV KETHIREDDYPALLY-1",
            "400KV KETHIREDDYPALLY-2",
            "400KV NIZAMABAD-1 & 2",
            "220KV PARIGI-1",
            "220KV PARIGI-2",
            "220KV THANDUR",
            "220KV GACHIBOWLI-1",
            "220KV GACHIBOWLI-2",
            "220KV KETHIREDDYPALLY",
            "220KV YEDDUMAILARAM-1",
            "220KV YEDDUMAILARAM-2",
            "220KV SADASIVAPET-1",
            "220KV SADASIVAPET-2"
        ]
        
        target_feeders = []
        for f in all_feeders:
            if f['name'] in FEEDER_MAPPING:
                f['display_name'] = FEEDER_MAPPING[f['name']]
                target_feeders.append(f)
        
        target_feeders.sort(key=lambda x: FEEDER_ORDER.index(x['display_name']) if x['display_name'] in FEEDER_ORDER else 999)
        
        # 3. Get Entries
        entries = await db.entries.find(
            {"date": date_str},
            {"_id": 0}
        ).to_list(5000)
        
        entries_by_feeder = {}
        for e in entries:
            if e['feeder_id'] not in entries_by_feeder:
                entries_by_feeder[e['feeder_id']] = []
            entries_by_feeder[e['feeder_id']].append(e)
            
        # 4. Build Data
        report_data = []
        for idx, f in enumerate(target_feeders):
            f_entries = entries_by_feeder.get(f['id'], [])
            
            data = {
                "sl_no": idx + 1,
                "feeder_name": f['display_name'],
                "shankarpally": {"import": {}, "export": {}},
                "other_end": {"import": {}, "export": {}},
                "stats": {}
            }
            
            if f_entries:
                # For daily report, we only have one entry per feeder (if it exists)
                # But we treat it as first/last for consistency with calculation logic
                entry = f_entries[0]
                
                # Shankarpally (End 1)
                # Import
                s_imp_init = entry.get('end1_import_initial', 0)
                s_imp_final = entry.get('end1_import_final', 0)
                s_imp_mf = f.get('end1_import_mf', 1)
                s_imp_cons = (s_imp_final - s_imp_init) * s_imp_mf
                
                data["shankarpally"]["import"] = {
                    "initial": s_imp_init,
                    "final": s_imp_final,
                    "mf": s_imp_mf,
                    "consumption": s_imp_cons
                }
                
                # Export
                s_exp_init = entry.get('end1_export_initial', 0)
                s_exp_final = entry.get('end1_export_final', 0)
                s_exp_mf = f.get('end1_export_mf', 1)
                s_exp_cons = (s_exp_final - s_exp_init) * s_exp_mf
                
                data["shankarpally"]["export"] = {
                    "initial": s_exp_init,
                    "final": s_exp_final,
                    "mf": s_exp_mf,
                    "consumption": s_exp_cons
                }
                
                # Other End (End 2)
                # Import
                o_imp_init = entry.get('end2_import_initial', 0)
                o_imp_final = entry.get('end2_import_final', 0)
                o_imp_mf = f.get('end2_import_mf', 1)
                o_imp_cons = (o_imp_final - o_imp_init) * o_imp_mf
                
                data["other_end"]["import"] = {
                    "initial": o_imp_init,
                    "final": o_imp_final,
                    "mf": o_imp_mf,
                    "consumption": o_imp_cons
                }
                
                # Export
                o_exp_init = entry.get('end2_export_initial', 0)
                o_exp_final = entry.get('end2_export_final', 0)
                o_exp_mf = f.get('end2_export_mf', 1)
                o_exp_cons = (o_exp_final - o_exp_init) * o_exp_mf
                
                data["other_end"]["export"] = {
                    "initial": o_exp_init,
                    "final": o_exp_final,
                    "mf": o_exp_mf,
                    "consumption": o_exp_cons
                }
                
                # Calculate Losses
                # Q = (D+L-H-P)/(D+L)
                D = s_imp_cons
                L = o_imp_cons
                H = s_exp_cons
                P = o_exp_cons
                
                numerator = (D + L) - (H + P)
                denominator = (D + L)
                
                pct_loss = (numerator / denominator * 100) if denominator != 0 else 0
                
                data["stats"]["pct_loss"] = pct_loss
                
            else:
                # No data
                for end in ["shankarpally", "other_end"]:
                    for type_ in ["import", "export"]:
                        data[end][type_] = {"initial": 0, "final": 0, "mf": 0, "consumption": 0}
                data["stats"]["pct_loss"] = 0
                
            report_data.append(data)
            
        return report_data
        
    except Exception as e:
        print(f"Error in daily report preview: {str(e)}")
        return JSONResponse(status_code=500, content={"detail": str(e)})

async def _generate_line_losses_report_wb(year: int, month: int):
    import calendar
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter
    
    # --- Common Styles ---
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(bold=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Line Losses"
    
    month_name = calendar.month_name[month]
    
    # 1. Headers
    # Row 1: Title
    ws.merge_cells('A1:T1')
    c = ws.cell(row=1, column=1, value=f"400KV SHANKARAPALLY SS- ENERGY LOSSES FOR THE MONTH {month_name}-{year}")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    # Row 2: Main Headers
    ws.merge_cells('A2:A3')
    c = ws.cell(row=2, column=1, value="Sl.\nNo.")
    c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('B2:B3')
    c = ws.cell(row=2, column=2, value="Name of the Feeder")
    c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('C2:J2')
    c = ws.cell(row=2, column=3, value="Shankarapally End")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('K2:R2')
    c = ws.cell(row=2, column=11, value="Other End")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('S2:S3')
    c = ws.cell(row=2, column=19, value="% Losses/\nCumulative\nLosses")
    c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('T2:T3')
    c = ws.cell(row=2, column=20, value="Remarks")
    c.alignment = center_align; c.border = thin_border
    
    # Row 3: Sub Headers (Import/Export)
    # Shankarpally
    ws.merge_cells('C3:F3')
    c = ws.cell(row=3, column=3, value="Import")
    c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('G3:J3')
    c = ws.cell(row=3, column=7, value="Export")
    c.alignment = center_align; c.border = thin_border
    
    # Other End
    ws.merge_cells('K3:N3')
    c = ws.cell(row=3, column=11, value="Import")
    c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('O3:R3')
    c = ws.cell(row=3, column=15, value="Export")
    c.alignment = center_align; c.border = thin_border
    
    # Row 4: Detailed Headers
    sub_headers = [
        "Initial Reading\nin MWH", "Final Reading in\nMWH", "MF", "Consumption",
        "Initial Reading\nin MWH", "Final Reading in\nMWH", "MF", "Consumption",
        "Initial Reading\nin MWH", "Final Reading in\nMWH", "MF", "Consumption",
        "Initial Reading\nin MWH", "Final Reading in\nMWH", "MF", "Consumption"
    ]
    
    for i, h in enumerate(sub_headers):
        cell = ws.cell(row=4, column=i+3, value=h)
        cell.alignment = center_align; cell.border = thin_border; cell.font = bold_font
        
    # Row 5: Column Letters/Formulas
    letters = ["A", "B", "C", "D=(B-A)*C", "E", "F", "G", "H=(F-E)*G", 
               "I", "J", "K", "L=(J-I)*K", "M", "N", "O", "P=(N-M)*O"]
    
    for i, l in enumerate(letters):
        cell = ws.cell(row=5, column=i+3, value=l)
        cell.alignment = center_align; cell.border = thin_border; cell.font = bold_font
        
    ws.cell(row=5, column=19, value="Q=(D+L-H-P)/(D+L)").alignment = center_align
    ws.cell(row=5, column=19).border = thin_border
    ws.cell(row=5, column=20, value="").border = thin_border # Remarks
    
    # 2. Fetch Data
    all_feeders = await db.feeders.find({}, {"_id": 0}).to_list(100)
    
    FEEDER_MAPPING = {
        "400 KV Shankarpally-MHRM-2": "400KV MAHESHWARAM-2",
        "400 KV Shankarpally-MHRM-1": "400KV MAHESHWARAM-1",
        "400 KV Shankarpally-Narsapur-1": "400KV NARSAPUR-1",
        "400 KV Shankarpally-Narsapur-2": "400KV NARSAPUR-2",
        "400 KV KethiReddyPally-1": "400KV KETHIREDDYPALLY-1",
        "400 KV KethiReddyPally-2": "400KV KETHIREDDYPALLY-2",
        "400 KV Nizamabad-1&2": "400KV NIZAMABAD-1 & 2",
        "220 KV Parigi-1": "220KV PARIGI-1",
        "220 KV Parigi-2": "220KV PARIGI-2",
        "220 KV Tandur": "220KV THANDUR",
        "220 KV Gachibowli-1": "220KV GACHIBOWLI-1",
        "220 KV Gachibowli-2": "220KV GACHIBOWLI-2",
        "220 KV KethiReddyPally": "220KV KETHIREDDYPALLY",
        "220 KV Yeddumailaram-1": "220KV YEDDUMAILARAM-1",
        "220 KV Yeddumailaram-2": "220KV YEDDUMAILARAM-2",
        "220 KV Sadasivapet-1": "220KV SADASIVAPET-1",
        "220 KV Sadasivapet-2": "220KV SADASIVAPET-2"
    }

    FEEDER_ORDER = [
        "400KV MAHESHWARAM-2", "400KV MAHESHWARAM-1", "400KV NARSAPUR-1", "400KV NARSAPUR-2",
        "400KV KETHIREDDYPALLY-1", "400KV KETHIREDDYPALLY-2", "400KV NIZAMABAD-1 & 2", "220KV PARIGI-1", "220KV PARIGI-2",
        "220KV THANDUR", "220KV GACHIBOWLI-1", "220KV GACHIBOWLI-2", "220KV KETHIREDDYPALLY",
        "220KV YEDDUMAILARAM-1", "220KV YEDDUMAILARAM-2", "220KV SADASIVAPET-1", "220KV SADASIVAPET-2"
    ]
    
    target_feeders = []
    for f in all_feeders:
        if f['name'] in FEEDER_MAPPING:
            f['display_name'] = FEEDER_MAPPING[f['name']]
            target_feeders.append(f)

    target_feeders.sort(key=lambda x: FEEDER_ORDER.index(x['display_name']) if x['display_name'] in FEEDER_ORDER else 999)
    
    start_date = f"{year}-{month:02d}-01"
    if month == 12: end_date = f"{year + 1}-01-01"
    else: end_date = f"{year}-{month + 1:02d}-01"
        
    entries = await db.entries.find(
        {"date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0}
    ).to_list(5000)
    
    entries_by_feeder = {}
    for e in entries:
        if e['feeder_id'] not in entries_by_feeder:
            entries_by_feeder[e['feeder_id']] = []
        entries_by_feeder[e['feeder_id']].append(e)
        
    row_idx = 6
    for idx, f in enumerate(target_feeders):
        f_entries = entries_by_feeder.get(f['id'], [])
        f_entries.sort(key=lambda x: x['date'])
        
        vals = [0] * 16 # 4 groups of 4
        pct_loss = 0
        
        if f_entries:
            first = f_entries[0]
            last = f_entries[-1]
            
            # S-Imp
            si_init = first.get('end1_import_initial', 0) or 0
            si_final = last.get('end1_import_final', 0) or 0
            si_mf = f.get('end1_import_mf', 1) or 1
            si_cons = (si_final - si_init) * si_mf
            
            # S-Exp
            se_init = first.get('end1_export_initial', 0) or 0
            se_final = last.get('end1_export_final', 0) or 0
            se_mf = f.get('end1_export_mf', 1) or 1
            se_cons = (se_final - se_init) * se_mf
            
            # O-Imp
            oi_init = first.get('end2_import_initial', 0) or 0
            oi_final = last.get('end2_import_final', 0) or 0
            oi_mf = f.get('end2_import_mf', 1) or 1
            oi_cons = (oi_final - oi_init) * oi_mf
            
            # O-Exp
            oe_init = first.get('end2_export_initial', 0) or 0
            oe_final = last.get('end2_export_final', 0) or 0
            oe_mf = f.get('end2_export_mf', 1) or 1
            oe_cons = (oe_final - oe_init) * oe_mf
            
            vals = [
                si_init, si_final, si_mf, si_cons,
                se_init, se_final, se_mf, se_cons,
                oi_init, oi_final, oi_mf, oi_cons,
                oe_init, oe_final, oe_mf, oe_cons
            ]
            
            D = si_cons; L = oi_cons; H = se_cons; P = oe_cons
            numerator = (D + L) - (H + P)
            denominator = (D + L)
            pct_loss = (numerator / denominator * 100) if denominator != 0 else 0

        # Write Row
        ws.cell(row=row_idx, column=1, value=idx+1).border = thin_border
        ws.cell(row=row_idx, column=2, value=f['display_name']).border = thin_border
        
        for i, v in enumerate(vals):
            cell = ws.cell(row=row_idx, column=i+3, value=v)
            cell.number_format = '0.00'
            cell.border = thin_border
            cell.alignment = center_align
            
        cell = ws.cell(row=row_idx, column=19, value=pct_loss)
        cell.number_format = '0.00'
        cell.border = thin_border
        cell.alignment = center_align
        
        ws.cell(row=row_idx, column=20, value="-").border = thin_border
        
        row_idx += 1
        
    # Auto-fit
    for i in range(1, ws.max_column + 1):
        col_letter = get_column_letter(i)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.row <= 5: continue
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except: pass
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)
    
    return wb


async def _get_new_line_losses_report_data(year: int, month: int) -> Dict[str, Any]:
    import calendar
    from datetime import date

    base_rows = await _get_line_losses_report_data(year, month)

    if month == 1:
        prev_year = year - 1
        prev_month = 12
    else:
        prev_year = year
        prev_month = month - 1

    prev_last_day = calendar.monthrange(prev_year, prev_month)[1]
    curr_last_day = calendar.monthrange(year, month)[1]
    prev_end = date(prev_year, prev_month, prev_last_day)
    curr_end = date(year, month, curr_last_day)

    prev_label = prev_end.strftime("%d-%m-%Y") + " (12:00 Hrs)"
    curr_label = curr_end.strftime("%d-%m-%Y") + " (12:00 Hrs)"

    header_text = (
        "Line loss report of 400KV & 220kV feeders at 400 KV Shankarpally "
        f"From {prev_label} to {curr_label}"
    )

    rows: List[Dict[str, Any]] = []
    sl = 1
    ss_name = "400KV SHANKARPALLY"

    for base in base_rows:
        feeder_name = (base.get("feeder_name") or "").strip()

        if feeder_name.upper() == "400KV NIZAMABAD-1 & 2":
            continue
        shank = base.get("shankarpally") or {}
        other = base.get("other_end") or {}

        s_exp = shank.get("export") or {}
        o_imp = other.get("import") or {}
        s_exp_cons = float(s_exp.get("consumption") or 0)
        o_imp_cons = float(o_imp.get("consumption") or 0)
        loss_exp = s_exp_cons - o_imp_cons
        pct_exp = (loss_exp / s_exp_cons * 100.0) if s_exp_cons != 0 else None

        rows.append(
            {
                "sl_no": sl,
                "ss_name": ss_name,
                "feeder_name": feeder_name,
                "flow_type": "Export",
                "sh_initial": s_exp.get("initial", 0),
                "sh_final": s_exp.get("final", 0),
                "sh_mf": s_exp.get("mf", 0),
                "sh_consumption": s_exp_cons,
                "other_initial": o_imp.get("initial", 0),
                "other_final": o_imp.get("final", 0),
                "other_mf": o_imp.get("mf", 0),
                "other_consumption": o_imp_cons,
                "losses": loss_exp,
                "pct_losses": pct_exp,
                "remarks": "",
            }
        )
        sl += 1

        s_imp = shank.get("import") or {}
        o_exp = other.get("export") or {}
        s_imp_cons = float(s_imp.get("consumption") or 0)
        o_exp_cons = float(o_exp.get("consumption") or 0)
        loss_imp = s_imp_cons - o_exp_cons
        pct_imp = (loss_imp / s_imp_cons * 100.0) if s_imp_cons != 0 else None

        rows.append(
            {
                "sl_no": sl,
                "ss_name": ss_name,
                "feeder_name": feeder_name,
                "flow_type": "Import",
                "sh_initial": s_imp.get("initial", 0),
                "sh_final": s_imp.get("final", 0),
                "sh_mf": s_imp.get("mf", 0),
                "sh_consumption": s_imp_cons,
                "other_initial": o_exp.get("initial", 0),
                "other_final": o_exp.get("final", 0),
                "other_mf": o_exp.get("mf", 0),
                "other_consumption": o_exp_cons,
                "losses": loss_imp,
                "pct_losses": pct_imp,
                "remarks": "",
            }
        )
        sl += 1

    return {
        "header": header_text,
        "rows": rows,
    }


async def _generate_new_line_losses_report_wb(year: int, month: int):
    from openpyxl.styles import Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter

    data = await _get_new_line_losses_report_data(year, month)
    header_text = data.get("header", "")
    rows = data.get("rows") or []

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    bold = Font(bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "New Line Losses"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
    c = ws.cell(row=1, column=1, value=header_text)
    c.font = bold
    c.alignment = center
    c.border = thin_border

    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.cell(row=2, column=1, value="Sl.No").alignment = center
    ws.cell(row=2, column=1).border = thin_border

    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    ws.cell(row=2, column=2, value="Name of the SS").alignment = center
    ws.cell(row=2, column=2).border = thin_border

    ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)
    ws.cell(row=2, column=3, value="Name of the Feeder").alignment = center
    ws.cell(row=2, column=3).border = thin_border

    ws.merge_cells(start_row=2, start_column=4, end_row=3, end_column=4)
    ws.cell(row=2, column=4, value="Type of flow wrt SS").alignment = center
    ws.cell(row=2, column=4).border = thin_border

    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=8)
    ws.cell(row=2, column=5, value="Shankarpally End").alignment = center
    ws.cell(row=2, column=5).font = bold
    ws.cell(row=2, column=5).border = thin_border

    ws.merge_cells(start_row=2, start_column=9, end_row=2, end_column=12)
    ws.cell(row=2, column=9, value="Other End").alignment = center
    ws.cell(row=2, column=9).font = bold
    ws.cell(row=2, column=9).border = thin_border

    ws.merge_cells(start_row=2, start_column=13, end_row=3, end_column=13)
    ws.cell(row=2, column=13, value="Losses").alignment = center
    ws.cell(row=2, column=13).border = thin_border

    ws.merge_cells(start_row=2, start_column=14, end_row=3, end_column=14)
    ws.cell(row=2, column=14, value="% of Losses").alignment = center
    ws.cell(row=2, column=14).border = thin_border

    ws.merge_cells(start_row=2, start_column=15, end_row=3, end_column=15)
    ws.cell(row=2, column=15, value="Remarks").alignment = center
    ws.cell(row=2, column=15).border = thin_border

    headers = [
        "Initial Reading",
        "Final Reading",
        "MF",
        "Consumption in MWH",
        "Initial Reading",
        "Final Reading",
        "MF",
        "Consumption in MWH",
    ]
    col = 5
    for h in headers:
        cell = ws.cell(row=3, column=col, value=h)
        cell.alignment = center
        cell.font = bold
        cell.border = thin_border
        col += 1

    row_idx = 4
    for r in rows:
        ws.cell(row=row_idx, column=1, value=r.get("sl_no")).alignment = center
        ws.cell(row=row_idx, column=1).border = thin_border

        ws.cell(row=row_idx, column=2, value=r.get("ss_name")).alignment = center
        ws.cell(row=row_idx, column=2).border = thin_border

        ws.cell(row=row_idx, column=3, value=r.get("feeder_name")).alignment = center
        ws.cell(row=row_idx, column=3).border = thin_border

        ws.cell(row=row_idx, column=4, value=r.get("flow_type")).alignment = center
        ws.cell(row=row_idx, column=4).border = thin_border

        numeric_fields = [
            ("sh_initial", 5),
            ("sh_final", 6),
            ("sh_mf", 7),
            ("sh_consumption", 8),
            ("other_initial", 9),
            ("other_final", 10),
            ("other_mf", 11),
            ("other_consumption", 12),
            ("losses", 13),
            ("pct_losses", 14),
        ]
        for key, col_idx in numeric_fields:
            val = r.get(key)
            display_val = "-" if key == "pct_losses" and val is None else val
            cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
            cell.alignment = right_align
            cell.border = thin_border

        ws.cell(row=row_idx, column=15, value=r.get("remarks", "")).border = thin_border
        row_idx += 1

    if rows:
        first_data_row = 4
        last_data_row = first_data_row + len(rows) - 1

        ws.merge_cells(
            start_row=first_data_row,
            start_column=2,
            end_row=last_data_row,
            end_column=2,
        )
        ss_cell = ws.cell(row=first_data_row, column=2)
        ss_cell.alignment = center
        ss_cell.border = thin_border

        total_rows = len(rows)
        for i in range(0, total_rows, 2):
            r1 = rows[i]
            r2 = rows[i + 1] if i + 1 < total_rows else None
            if not r2 or r1.get("feeder_name") != r2.get("feeder_name"):
                continue
            row_start = first_data_row + i
            row_end = row_start + 1
            ws.merge_cells(
                start_row=row_start,
                start_column=3,
                end_row=row_end,
                end_column=3,
            )
            feeder_cell = ws.cell(row=row_start, column=3)
            feeder_cell.alignment = center
            feeder_cell.border = thin_border

    for col_idx in range(1, 16):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            val = cell.value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = max(max_len + 2, 10)

    return wb

@api_router.get("/reports/line-losses/export/{year}/{month}")
async def export_line_losses_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        import calendar
        wb = await _generate_line_losses_report_wb(year, month)
        month_name = calendar.month_name[month]
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Line_Losses_{month_name}_{year}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        with open("line_losses_export_error.txt", "w") as f:
            f.write(error_msg)
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": str(e)})


@api_router.get("/reports/new-line-losses/preview/{year}/{month}")
async def get_new_line_losses_report_preview(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user),
):
    try:
        return await _get_new_line_losses_report_data(year, month)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/reports/new-line-losses/export/{year}/{month}")
async def export_new_line_losses_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user),
):
    wb = await _generate_new_line_losses_report_wb(year, month)
    try:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"New_Line_Losses_{year}_{month:02d}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        wb.close()

# --- PTR Max-Min (Format-1) Endpoints ---

@api_router.get("/reports/ptr-max-min-format1/preview/{year}/{month}")
async def get_ptr_max_min_preview(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"

    # Fetch ICT feeders
    ict_feeders = await db.max_min_feeders.find({"type": "ict_feeder"}, {"_id": 0}).to_list(100)
    
    # Sort ICT feeders
    ICT_ORDER = ["ICT-1 (315MVA)", "ICT-2 (315MVA)", "ICT-3 (315MVA)", "ICT-4 (500MVA)"]
    ict_feeders.sort(key=lambda x: ICT_ORDER.index(x['name']) if x['name'] in ICT_ORDER else 999)

    MD_SO_FAR_BASELINE_PTR = {
        "ICT-1 (315MVA)": 226,
        "ICT-2 (315MVA)": 225,
        "ICT-3 (315MVA)": 226,
        "ICT-4 (500MVA)": 387,
    }

    MD_YEAR_PTR = 2026
    md_2026_map: dict[str, float] = {}

    if year == MD_YEAR_PTR and ict_feeders:
        md_start = f"{MD_YEAR_PTR}-01-01"
        if month == 12:
            md_end = f"{MD_YEAR_PTR + 1}-01-01"
        else:
            md_end = f"{MD_YEAR_PTR}-{month + 1:02d}-01"

        ict_ids = [f["id"] for f in ict_feeders if "id" in f]

        if ict_ids:
            md_entries = await db.max_min_entries.find(
                {
                    "feeder_id": {"$in": ict_ids},
                    "date": {"$gte": md_start, "$lt": md_end},
                },
                {"_id": 0},
            ).to_list(10000)

            for e in md_entries:
                feeder_id = e.get("feeder_id")
                if not feeder_id:
                    continue
                d = e.get("data", {})
                mw_val = get_float_safe(d.get("max", {}).get("mw"))
                if mw_val is None:
                    continue
                prev = md_2026_map.get(feeder_id)
                if prev is None or mw_val > prev:
                    md_2026_map[feeder_id] = mw_val
    
    data = []
    
    for feeder in ict_feeders:
        entries = await db.max_min_entries.find(
            {"feeder_id": feeder['id'], "date": {"$gte": start_date, "$lt": end_date}},
            {"_id": 0}
        ).to_list(1000)
        
        if not entries:
            # Placeholder for missing data
            rating = "315"
            if "500" in feeder['name']:
                rating = "500"

            md_2026_val = md_2026_map.get(feeder["id"]) if year == MD_YEAR_PTR else None
            baseline = MD_SO_FAR_BASELINE_PTR.get(feeder["name"])
            final_md_so_far = baseline
            if year == MD_YEAR_PTR and md_2026_val is not None and baseline is not None and md_2026_val > baseline:
                final_md_so_far = md_2026_val

            data.append({
                "district": "Rangareddy",
                "substation": "400/220 KV SHANKARPALLY",
                "ptr_kv": "400/220",
                "rating": rating,
                "general": {"mw": 0, "mvar": 0},
                "max": None,
                "min": None,
                "md_2026": md_2026_val,
                "md_so_far": final_md_so_far
            })
            continue
            
        # Stats
        max_mw = -1.0
        max_entry = None
        min_mw = float('inf')
        min_entry = None
        
        total_mw = 0.0
        total_mvar_avg = 0.0 # Derived from daily max MVAR
        count = 0
        
        for e in entries:
            d = e.get('data', {})
            
            # Max MW logic (Find entry with highest Max MW)
            try:
                curr_max_mw = float(d.get('max', {}).get('mw', 0) or 0)
            except:
                curr_max_mw = 0.0
                
            if curr_max_mw > max_mw:
                max_mw = curr_max_mw
                max_entry = e
            
            # Min MW logic (Find entry with lowest Min MW)
            try:
                curr_min_mw = float(d.get('min', {}).get('mw', 0) or 0)
            except:
                curr_min_mw = 0.0
            
            # Use strict comparison, assuming data is valid.
            if curr_min_mw < min_mw:
                min_mw = curr_min_mw
                min_entry = e
                
            # Avg Calculation
            try:
                day_avg_mw = float(d.get('avg', {}).get('mw', 0) or 0)
            except:
                day_avg_mw = 0.0
                
            try:
                day_max_mvar = float(d.get('max', {}).get('mvar', 0) or 0)
                # day_min_mvar = float(d.get('min', {}).get('mvar', 0) or 0)
                day_avg_mvar = day_max_mvar # User requested average of full-month MAX MVAR values
            except:
                day_avg_mvar = 0.0
                
            total_mw += day_avg_mw
            total_mvar_avg += day_avg_mvar
            count += 1
            
        avg_mw = total_mw / count if count > 0 else 0
        avg_mvar = total_mvar_avg / count if count > 0 else 0
        
        # Extract Max Details
        max_details = {}
        if max_entry:
            md = max_entry.get('data', {}).get('max', {})
            mw_val = float(md.get('mw', 0) or 0)
            mvar_val = float(md.get('mvar', 0) or 0)
            max_details = {
                "date": max_entry['date'],
                "time": format_time(md.get('time', '')),
                "mw": mw_val,
                "mvar": mvar_val,
                "mva": (mw_val**2 + mvar_val**2)**0.5
            }
        else:
             max_details = {"date": "-", "time": "-", "mw": 0, "mvar": 0, "mva": 0}
        
        # Extract Min Details
        min_details = {}
        if min_entry and min_mw != float('inf'):
            md = min_entry.get('data', {}).get('min', {})
            mw_val = float(md.get('mw', 0) or 0)
            mvar_val = float(md.get('mvar', 0) or 0)
            min_details = {
                "date": min_entry['date'],
                "time": format_time(md.get('time', '')),
                "mw": mw_val,
                "mvar": mvar_val,
                "mva": (mw_val**2 + mvar_val**2)**0.5
            }
        else:
            min_details = {"date": "-", "time": "-", "mw": 0, "mvar": 0, "mva": 0}
            
        rating = "315"
        if "500" in feeder['name']:
            rating = "500"

        md_2026_val = md_2026_map.get(feeder["id"]) if year == MD_YEAR_PTR else None
        baseline = MD_SO_FAR_BASELINE_PTR.get(feeder["name"])
        final_md_so_far = baseline
        if year == MD_YEAR_PTR and md_2026_val is not None and baseline is not None and md_2026_val > baseline:
            final_md_so_far = md_2026_val

        data.append({
            "district": "Rangareddy",
            "substation": "400/220 KV SHANKARPALLY",
            "ptr_kv": "400/220",
            "rating": rating,
            "general": {
                "mw": avg_mw,
                "mvar": avg_mvar
            },
            "max": max_details,
            "min": min_details,
            "md_2026": md_2026_val,
            "md_so_far": final_md_so_far
        })
        
    return data

async def _generate_ptr_max_min_report_wb(year: int, month: int, current_user: User):
    data = await get_ptr_max_min_preview(year, month, current_user)
    
    month_name = calendar.month_name[month]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "PTR Max-Min Format-1"
    
    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    ws.merge_cells('A1:T1')
    c = ws.cell(row=1, column=1, value="FORMAT-I")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('A2:A3')
    c = ws.cell(row=2, column=1, value="ZONE: CE/METRO/HYD")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('B2:R2')
    c = ws.cell(row=2, column=2, value=f"NORMAL & MAXIMUM / MINIMUM LOADING ON PTRs FOR THE MONTH OF {month_name}-{year} in SE/OMC/Metro-West Circle")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('S2:S4')
    c = ws.cell(row=2, column=19, value=f"MD\nreached\nso far in\n{year}")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('T2:T4')
    c = ws.cell(row=2, column=20, value="MD\nreached\nso far")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    # Row 3
    ws.merge_cells('E3:F3')
    c = ws.cell(row=3, column=5, value="GENERAL\nLOADING")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('G3:L3')
    c = ws.cell(row=3, column=7, value="MAXIMUM LOAD DETAILS")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('M3:R3')
    c = ws.cell(row=3, column=13, value="MINIMUM LOAD DETAILS")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    # Row 4
    headers_r4 = [
        "DISTRICT\nNAME", "SUB-STATION\nNAME", "PTR\nHV/LV in\nKV", "PTR DETAILS/\nRATING IN\nMVA",
        "MW", "MVAR",
        "DATE", "TIME", "MW", "MVAR", "MVA", "BUS\nVOLTAGE IN\nKV @MAX",
        "DATE", "TIME", "MW", "MVAR", "MVA", "BUS VOLTAGE\nIN KV @MIN\nLOADING"
    ]
    
    for i, h in enumerate(headers_r4):
        c = ws.cell(row=4, column=1+i, value=h)
        c.font = bold_font; c.alignment = center_align; c.border = thin_border
        
    # Data
    row_idx = 5
    
    def set_style(c):
        c.border = thin_border
        c.alignment = center_align
        return c
        
    for item in data:
        set_style(ws.cell(row=row_idx, column=1, value=item['district']))
        set_style(ws.cell(row=row_idx, column=2, value=item['substation']))
        set_style(ws.cell(row=row_idx, column=3, value=item['ptr_kv']))
        set_style(ws.cell(row=row_idx, column=4, value=int(item['rating'])))
        
        set_style(ws.cell(row=row_idx, column=5, value=f"{item['general']['mw']:.2f}"))
        set_style(ws.cell(row=row_idx, column=6, value=f"{item['general']['mvar']:.2f}"))
        
        m = item['max']
        if m:
            set_style(ws.cell(row=row_idx, column=7, value=format_date(m['date'])))
            set_style(ws.cell(row=row_idx, column=8, value=m['time']))
            set_style(ws.cell(row=row_idx, column=9, value=f"{m['mw']:.2f}"))
            set_style(ws.cell(row=row_idx, column=10, value=f"{m['mvar']:.2f}"))
            set_style(ws.cell(row=row_idx, column=11, value=f"{m['mva']:.2f}"))
        else:
            for c in range(7, 12): set_style(ws.cell(row=row_idx, column=c, value="-"))

        set_style(ws.cell(row=row_idx, column=12, value=""))
        
        m = item['min']
        if m:
            set_style(ws.cell(row=row_idx, column=13, value=format_date(m['date'])))
            set_style(ws.cell(row=row_idx, column=14, value=m['time']))
            set_style(ws.cell(row=row_idx, column=15, value=f"{m['mw']:.2f}"))
            set_style(ws.cell(row=row_idx, column=16, value=f"{m['mvar']:.2f}"))
            set_style(ws.cell(row=row_idx, column=17, value=f"{m['mva']:.2f}"))
        else:
            for c in range(13, 18): set_style(ws.cell(row=row_idx, column=c, value="-"))

        set_style(ws.cell(row=row_idx, column=18, value=""))
        
        md_2026_val = item.get("md_2026")
        md_so_far_val = item.get("md_so_far")

        set_style(ws.cell(row=row_idx, column=19, value=md_2026_val if md_2026_val is not None else ""))
        set_style(ws.cell(row=row_idx, column=20, value=md_so_far_val if md_so_far_val is not None else ""))
        
        row_idx += 1
        
    # Merge District and Substation
    if row_idx > 5:
        ws.merge_cells(f'A5:A{row_idx-1}')
        ws.cell(row=5, column=1).alignment = center_align
        ws.merge_cells(f'B5:B{row_idx-1}')
        ws.cell(row=5, column=2).alignment = center_align
        
    # Auto-width
    for col in range(1, 21):
        ws.column_dimensions[get_column_letter(col)].width = 12
        
    return wb

@api_router.get("/reports/ptr-max-min-format1/export/{year}/{month}")
async def export_ptr_max_min_report(
    year: str,
    month: str,
    current_user: User = Depends(get_current_user)
):
    try:
        try:
            year_int = int(str(year).split('.')[0])
            month_int = int(str(month).split('.')[0])
        except ValueError:
            return JSONResponse(status_code=422, content={"detail": f"Invalid year or month format. Received: year={year}, month={month}"})

        wb = await _generate_ptr_max_min_report_wb(year_int, month_int, current_user)
        month_name = calendar.month_name[month_int]
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"PTR_Max_Min_{month_name}_{year_int}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"Export error traceback: {error_msg}")
        logger.error(f"Export error: {error_msg}")
        return JSONResponse(status_code=500, content={"detail": error_msg})

@api_router.get("/reports/tl-max-loading-format4/preview/{year}/{month}")
async def get_tl_max_loading_preview(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    start_date = f"{year}-{month:02d}-01"
    if month == 12:
        end_date = f"{year + 1}-01-01"
    else:
        end_date = f"{year}-{month + 1:02d}-01"

    # Define Feeder Order and Mapping
    TL_ORDER = [
        # 400KV
        "400KV MAHESHWARAM-2", "400KV MAHESHWARAM-1", 
        "400KV NARSAPUR-1", "400KV NARSAPUR-2",
        "400KV KETHIREDDYPALLY-1", "400KV KETHIREDDYPALLY-2",
        "400KV NIZAMABAD-1", "400KV NIZAMABAD-2",
        # 220KV
        "220KV PARIGI-1", "220KV PARIGI-2",
        "220KV THANDUR",
        "220KV GACHIBOWLI-1", "220KV GACHIBOWLI-2",
        "220KV KETHIREDDYPALLY",
        "220KV YEDDUMAILARAM-1", "220KV YEDDUMAILARAM-2",
        "220KV SADASIVAPET-1", "220KV SADASIVAPET-2"
    ]
    
    DISPLAY_NAMES = {
        "400KV MAHESHWARAM-2": "Shankraplly - Maheshwaram - II",
        "400KV MAHESHWARAM-1": "Shankraplly - Maheshwaram - I",
        "400KV NARSAPUR-1": "Shankraplly - Narsapur - I",
        "400KV NARSAPUR-2": "Shankraplly - Narsapur - II",
        "400KV KETHIREDDYPALLY-1": "Shankraplly -Kethireddypally-1",
        "400KV KETHIREDDYPALLY-2": "Shankraplly - Kethireddypally-2",
        "400KV NIZAMABAD-1": "Shankraplly - Nizamabad - I",
        "400KV NIZAMABAD-2": "Shankraplly - Nizamabad - II",
        "220KV PARIGI-1": "Shankarpally - Parigi-1",
        "220KV PARIGI-2": "Shankarpally - Parigi-2",
        "220KV THANDUR": "Shankarpally - Thandur",
        "220KV GACHIBOWLI-1": "Shankarpally - Gachibowli-1",
        "220KV GACHIBOWLI-2": "Shnakarpally - Gachibowli-2",
        "220KV KETHIREDDYPALLY": "Shankarpally -Kethireddypally",
        "220KV YEDDUMAILARAM-1": "Shankarpally - Yeddumailaram-1",
        "220KV YEDDUMAILARAM-2": "Shankarpally - Yeddumailaram-2",
        "220KV SADASIVAPET-1": "Shankarpally - Sadasivapet-1",
        "220KV SADASIVAPET-2": "Shankarpally - Sadasivapet-2"
    }

    MD_SO_FAR_BASELINE = {
        "400KV MAHESHWARAM-2": 426,
        "400KV MAHESHWARAM-1": 425,
        "400KV NARSAPUR-1": 564,
        "400KV NARSAPUR-2": 561,
        "400KV KETHIREDDYPALLY-1": 420,
        "400KV KETHIREDDYPALLY-2": 418,
        "400KV NIZAMABAD-1": 484,
        "400KV NIZAMABAD-2": 584,
        "220KV PARIGI-1": 170,
        "220KV PARIGI-2": 170,
        "220KV THANDUR": 208,
        "220KV GACHIBOWLI-1": 306,
        "220KV GACHIBOWLI-2": 376,
        "220KV KETHIREDDYPALLY": 140,
        "220KV YEDDUMAILARAM-1": 179,
        "220KV YEDDUMAILARAM-2": 173,
        "220KV SADASIVAPET-1": 116,
        "220KV SADASIVAPET-2": 115,
    }

    # Fetch all feeders (needed for group logic)
    all_db_feeders = await db.max_min_feeders.find({}, {"_id": 0}).to_list(1000)
    all_feeder_map = {f['name']: f for f in all_db_feeders}
    all_feeder_id_map = {f['id']: f for f in all_db_feeders}

    MD_YEAR = 2026
    md_2026_map: dict[str, float | None] = {}

    if year == MD_YEAR:
        md_start = f"{MD_YEAR}-01-01"
        if month == 12:
            md_end = f"{MD_YEAR + 1}-01-01"
        else:
            md_end = f"{MD_YEAR}-{month + 1:02d}-01"

        md_entries = await db.max_min_entries.find(
            {
                "date": {"$gte": md_start, "$lt": md_end},
            },
            {"_id": 0},
        ).to_list(100000)

        entries_2026_by_feeder: dict[str, list[dict]] = {}
        for e in md_entries:
            fid = e.get("feeder_id")
            if not fid:
                continue
            if fid not in entries_2026_by_feeder:
                entries_2026_by_feeder[fid] = []
            entries_2026_by_feeder[fid].append(e)

        for feeder_name in TL_ORDER:
            feeder = all_feeder_map.get(feeder_name)
            if not feeder:
                continue

            feeder_id = feeder["id"]
            feeder_entries = entries_2026_by_feeder.get(feeder_id, [])

            is_special, partner_names = get_feeder_group_info(feeder["name"])
            if is_special:
                p_group_map: dict[str, list[dict]] = {feeder_id: feeder_entries}
                for pname in partner_names:
                    p_feeder = all_feeder_map.get(pname)
                    if not p_feeder:
                        continue
                    p_id = p_feeder["id"]
                    p_group_map[p_id] = entries_2026_by_feeder.get(p_id, [])

                leader_id = determine_leader(p_group_map)
                if not leader_id:
                    leader_id = feeder_id

                leader_entries = p_group_map.get(leader_id, [])
                ftype = feeder.get("type", "feeder")
                stats_year = calculate_coincident_stats(leader_entries, feeder_id, p_group_map, ftype)
            else:
                ftype = feeder.get("type", "feeder")
                stats_year = calculate_standard_stats(feeder_entries, ftype)

            max_mw_year = stats_year.get("max_mw", "-")
            max_mw_year_val = get_float_safe(max_mw_year)
            if max_mw_year_val is not None:
                md_2026_map[feeder_name] = max_mw_year_val

    # Fetch all entries for the month
    all_entries = await db.max_min_entries.find(
        {"date": {"$gte": start_date, "$lt": end_date}},
        {"_id": 0}
    ).to_list(10000)

    # Group entries by feeder_id
    entries_by_feeder = {}
    for e in all_entries:
        fid = e.get('feeder_id')
        if fid:
            if fid not in entries_by_feeder:
                entries_by_feeder[fid] = []
            entries_by_feeder[fid].append(e)
    
    data = []
    
    for idx, feeder_name in enumerate(TL_ORDER):
        feeder = all_feeder_map.get(feeder_name)
        
        mw_val = ""
        mvar_val = ""
        date_val = ""
        time_val = ""
        md_2026_val_str = ""
        md_so_far_val_str = ""
        
        if feeder:
            feeder_entries = entries_by_feeder.get(feeder['id'], [])
            
            # Check for Grouping Logic (Pair Feeder Logic)
            is_special, partner_names = get_feeder_group_info(feeder['name'])
            
            if is_special:
                p_group_map = {}
                p_group_map[feeder['id']] = feeder_entries
                
                # Gather partner entries
                for pname in partner_names:
                    p_feeder = all_feeder_map.get(pname)
                    if p_feeder:
                         p_group_map[p_feeder['id']] = entries_by_feeder.get(p_feeder['id'], [])
                
                # Determine Leader for the Full Month
                leader_id = determine_leader(p_group_map)
                if not leader_id: leader_id = feeder['id']
                
                leader_entries = p_group_map.get(leader_id, [])
                
                # Calculate Stats
                # Note: 'type' might be missing in feeder obj if not fetched properly, default to 'feeder'
                ftype = feeder.get('type', 'feeder') 
                stats = calculate_coincident_stats(leader_entries, feeder['id'], p_group_map, ftype)
                
                # Extract values
                try:
                    max_mw = stats.get('max_mw', '-')
                    if max_mw != '-' and max_mw is not None:
                        mw_val = f"{float(max_mw):.2f}"
                        date_val = format_date(stats.get('max_mw_date', ''))
                        time_val = format_time(stats.get('max_mw_time', ''))
                except:
                    pass
            else:
                # Standard Logic (Individual Max)
                ftype = feeder.get('type', 'feeder')
                stats = calculate_standard_stats(feeder_entries, ftype)
                
                try:
                    max_mw = stats.get('max_mw', '-')
                    if max_mw != '-' and max_mw is not None:
                        mw_val = f"{float(max_mw):.2f}"
                        date_val = format_date(stats.get('max_mw_date', ''))
                        time_val = format_time(stats.get('max_mw_time', ''))
                except:
                    pass

            md_2026_num = md_2026_map.get(feeder_name)
            if md_2026_num is not None:
                md_2026_val_str = f"{md_2026_num:.0f}"

            baseline = MD_SO_FAR_BASELINE.get(feeder_name)
            final_md_so_far = baseline
            if md_2026_num is not None and baseline is not None and md_2026_num > baseline:
                final_md_so_far = md_2026_num
            if final_md_so_far is not None:
                md_so_far_val_str = f"{final_md_so_far:.0f}"

        display_name = DISPLAY_NAMES.get(feeder_name, feeder_name)
        voltage = "400 KV" if "400" in feeder_name else "220KV"
        
        data.append({
            "sl_no": idx + 1,
            "district": "Ranga Reddy",
            "voltage": voltage,
            "substation": "Shankarpally",
            "line_name": display_name,
            "mw": mw_val,
            "mvar": mvar_val,
            "date": date_val,
            "time": time_val,
            "md_2026": md_2026_val_str,
            "md_so_far": md_so_far_val_str,
            "remarks": ""
        })
            
    return data

async def _generate_tl_max_loading_report_wb(year: int, month: int, current_user: User):
    print(f"Exporting TL Max Loading for {year}-{month}")
    data = await get_tl_max_loading_preview(year, month, current_user)
    print(f"Data fetched: {len(data)} rows")
    month_name = calendar.month_name[month]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "TL Max Loading Format-4"
    
    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Header Row 1
    ws.merge_cells('A1:L1')
    c = ws.cell(row=1, column=1, value=f"FORMAT-IV - SE/OMC/Metro-West Circle for the Month of {month_name}-{year}")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    # Header Row 2
    ws.merge_cells('A2:D2')
    c = ws.cell(row=2, column=1, value="ZONE CE/METRO/HYD")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    ws.merge_cells('F2:I2')
    c = ws.cell(row=2, column=6, value="MAXIMUM")
    c.font = bold_font; c.alignment = center_align; c.border = thin_border
    
    headers = [
        "SL.No", "DISTRICT\nNAME", "VOLTAGE\nLEVEL IN KV", "SUBSTATION\nNAME",
        "TRANSMISSION LINE NAME",
        "MW", "MVAR", "DATE", "TIME",
        "MD\nreached\nin 2026", "MD\nreached\nso far", "REMARKS"
    ]
    
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=1+i, value=h)
        c.font = bold_font; c.alignment = center_align; c.border = thin_border
        
    # Data Rows
    row_idx = 4
    
    def set_style(c):
        c.border = thin_border
        c.alignment = center_align
        return c
        
    for item in data:
        set_style(ws.cell(row=row_idx, column=1, value=item['sl_no']))
        set_style(ws.cell(row=row_idx, column=2, value=item['district']))
        set_style(ws.cell(row=row_idx, column=3, value=item['voltage']))
        set_style(ws.cell(row=row_idx, column=4, value=item['substation']))
        set_style(ws.cell(row=row_idx, column=5, value=item['line_name']))
        set_style(ws.cell(row=row_idx, column=6, value=item['mw']))
        set_style(ws.cell(row=row_idx, column=7, value=item['mvar']))
        set_style(ws.cell(row=row_idx, column=8, value=item['date']))
        set_style(ws.cell(row=row_idx, column=9, value=item['time']))
        set_style(ws.cell(row=row_idx, column=10, value=item['md_2026']))
        set_style(ws.cell(row=row_idx, column=11, value=item['md_so_far']))
        set_style(ws.cell(row=row_idx, column=12, value=item['remarks']))
        row_idx += 1
        
    # Merging Logic
    # District (Col 2), Substation (Col 4) -> Merge for all rows
    if row_idx > 4:
        # District
        ws.merge_cells(f'B4:B{row_idx-1}')
        
        # Substation
        ws.merge_cells(f'D4:D{row_idx-1}')
        
        # Voltage (Col 3) - Split into 400KV and 220KV blocks
        num_400 = 8
        num_220 = 10
        
        start_row = 4
        if len(data) >= num_400:
            ws.merge_cells(f'C{start_row}:C{start_row+num_400-1}')
        
        if len(data) >= num_400 + num_220:
            ws.merge_cells(f'C{start_row+num_400}:C{start_row+num_400+num_220-1}')
            
    # Column Widths
    ws.column_dimensions['E'].width = 30 # Line Name
    for col in [1, 2, 3, 4, 6, 7, 8, 9, 10, 11, 12]:
         ws.column_dimensions[get_column_letter(col)].width = 15

    return wb

@api_router.get("/reports/tl-max-loading-format4/export/{year}/{month}")
async def export_tl_max_loading_report(
    year: str,
    month: str,
    current_user: User = Depends(get_current_user)
):
    try:
        try:
            year_int = int(str(year).split('.')[0])
            month_int = int(str(month).split('.')[0])
        except ValueError:
            return JSONResponse(status_code=422, content={"detail": f"Invalid year or month format. Received: year={year}, month={month}"})

        wb = await _generate_tl_max_loading_report_wb(year_int, month_int, current_user)
        month_name = calendar.month_name[month_int]
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"TL_Max_Loading_{month_name}_{year_int}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"Export error traceback: {error_msg}")
        logger.error(f"Export error: {error_msg}")
        # Write to file for debugging
        try:
            with open("export_error.log", "w") as f:
                f.write(error_msg)
        except:
            pass
        return JSONResponse(status_code=500, content={"detail": error_msg})

@api_router.get("/reports/interruptions/export/{year}/{month}")
async def export_interruptions_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        import calendar
        wb = await _generate_interruptions_report_wb(year, month)
        month_name = calendar.month_name[month]

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Interruptions_Report_{month_name}_{year}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": str(e)})


@api_router.get("/reports/mis-interruptions/preview/{year}/{month}")
async def preview_mis_interruptions_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        data = await _get_mis_interruptions_report_data(year, month)
        return data
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": str(e)})


@api_router.get("/reports/mis-interruptions/export/{year}/{month}")
async def export_mis_interruptions_report(
    year: int,
    month: int,
    current_user: User = Depends(get_current_user)
):
    try:
        import calendar
        wb = await _generate_mis_interruptions_report_wb(year, month)
        month_name = calendar.month_name[month]

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"MIS_Interruption_Details_{month_name}_{year}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": str(e)})

@api_router.post("/reports/send-mail")
async def send_reports_email_endpoint(
    request: EmailReportRequest,
    current_user: User = Depends(get_current_user)
):
    try:
        year = request.year
        month = request.month
        recipient_email = request.email
        month_name = calendar.month_name[month]
        report_ids = request.report_ids
        
        attachments = []
        errors = []
        
        # 1. Fortnight Report
        if not report_ids or 'fortnight' in report_ids:
            try:
                wb = await _generate_fortnight_report_wb(year, month)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"Fortnight_Report_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating Fortnight Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)
            
        # 2. Energy Consumption (Optional, not currently in frontend list)
        if report_ids and 'energy-consumption' in report_ids:
            try:
                wb = await _generate_energy_export_wb(year, month)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"Energy_Consumption_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating Energy Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)

        # 3. Boundary Meter Report
        if not report_ids or 'boundary-meter' in report_ids:
            try:
                wb = await _generate_boundary_meter_wb(year, month)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"Boundary_Meter_Report_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating Boundary Meter Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)
             
        # 4. KPI Report
        if not report_ids or 'kpi' in report_ids:
            try:
                wb = await _generate_kpi_report_wb(year, month)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"KPI_Report_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating KPI Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)
             
        # 5. Line Losses
        if not report_ids or 'line-losses' in report_ids:
            try:
                wb = await _generate_line_losses_report_wb(year, month)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"Line_Losses_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating Line Losses Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)
             
        # 5a. New Line Losses (optional - only when explicitly requested)
        if report_ids and 'new-line-losses' in report_ids:
            try:
                wb = await _generate_new_line_losses_report_wb(year, month)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"New_Line_Losses_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating New Line Losses Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)
             
        # 6. Daily Max MVA
        if not report_ids or 'daily-max-mva' in report_ids:
            try:
                wb = await _generate_daily_max_mva_wb(year, month)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"Daily_Max_MVA_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating Daily Max MVA Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)
             
        # 7. PTR Max Min
        if not report_ids or 'ptr-max-min' in report_ids:
            try:
                wb = await _generate_ptr_max_min_report_wb(year, month, current_user)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"PTR_Max_Min_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating PTR Max Min Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)

        # 8. TL Max Loading
        if not report_ids or 'tl-max-loading' in report_ids:
            try:
                wb = await _generate_tl_max_loading_report_wb(year, month, current_user)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"TL_Max_Loading_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating TL Max Loading Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)

        # 9. Interruptions Monthly Report
        if not report_ids or 'interruptions' in report_ids:
            try:
                wb = await _generate_interruptions_report_wb(year, month)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"Interruptions_Report_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating Interruptions Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)
        
        # 10. MIS Interruption Details (optional - only when explicitly requested)
        if report_ids and 'mis-interruptions' in report_ids:
            try:
                wb = await _generate_mis_interruptions_report_wb(year, month)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                attachments.append((f"MIS_Interruption_Details_{month_name}_{year}.xlsx", output.read()))
            except Exception as e:
                import traceback
                error_msg = f"Error generating MIS Interruption Details Report: {str(e)}\n{traceback.format_exc()}\n"
                print(error_msg)
                errors.append(error_msg)
        
        if errors:
            error_content = "\n".join(errors)
            attachments.append(("generation_errors.txt", error_content.encode('utf-8')))
             
        if not attachments:
            raise HTTPException(status_code=400, detail="No reports could be generated for this period.")
            
        # Determine Subject Logic
        from datetime import date
        short_month_year = date(year, month, 1).strftime("%b-%Y")
        
        # Mapping of report IDs to readable names
        report_names_map = {
            'fortnight': 'Fortnight',
            'energy-consumption': 'Energy Consumption',
            'boundary-meter': 'Boundary Meter reading', # Based on user example "Boundary Meter reading"
            'kpi': 'KPI',
            'line-losses': 'Line Losses',
            'new-line-losses': 'New Line Losses',
            'daily-max-mva': 'Daily Max MVA',
            'ptr-max-min': 'PTR Max Min',
            'tl-max-loading': 'TL Max Loading',
            'interruptions': 'Interruptions',
            'mis-interruptions': 'MIS Interruption Details'
        }

        # Identify which reports were actually generated (or attempted)
        # Note: 'report_ids' might be None/Empty (meaning ALL default), or a list.
        # We need to reconstruct the list of report names that were part of this request.
        
        generated_report_names = []
        
        # List of all available report keys (excluding optional ones if not requested)
        all_default_keys = [
            'fortnight', 'boundary-meter', 'kpi', 'line-losses',
            'daily-max-mva', 'ptr-max-min', 'tl-max-loading', 'interruptions'
        ]
        
        # If report_ids is empty, it means all default reports
        current_report_keys = report_ids if report_ids else all_default_keys
        
        # Filter out keys that might be in report_ids but are optional and not in map if any
        # And ensure we use the readable names
        for key in current_report_keys:
             if key in report_names_map:
                 generated_report_names.append(report_names_map[key])
        
        # Determine if it is "All Reports"
        # Condition: report_ids is empty OR it contains all default keys (and maybe optional ones too)
        # Simplify: If report_ids is empty -> All Reports
        # If report_ids has all the default keys -> All Reports
        
        is_all_reports = False
        if not report_ids:
            is_all_reports = True
        elif set(all_default_keys).issubset(set(report_ids)):
            is_all_reports = True
            
        if is_all_reports:
             subject = f"MIS Reports of {month_name} {year}"
        elif len(generated_report_names) == 1:
             subject = f"MIS Report of {generated_report_names[0]} of {short_month_year}"
        else:
             # Less than all, but more than 1 (or 0, but we checked attachments)
             reports_str = ", ".join(generated_report_names)
             subject = f"MIS Reports of {reports_str} of {short_month_year}"

        # Send Email in background
        import asyncio
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, send_reports_email, recipient_email, attachments, subject)
        
        return {"message": f"Reports sent successfully to {recipient_email}"}
        
    except HTTPException as he:
        raise he
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(f"Send mail error: {error_msg}")
        return JSONResponse(status_code=500, content={"detail": str(e)})

@api_router.get("/admin/analytics/energy")
async def admin_energy_analytics(
    sheet_ids: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_admin: User = Depends(get_current_admin),
):
    query: dict[str, Any] = {}
    if sheet_ids:
        ids = [s for s in sheet_ids.split(",") if s]
        if ids:
            query["sheet_id"] = {"$in": ids}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    entries = await db.energy_entries.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    sheets = await db.energy_sheets.find({}, {"_id": 0}).to_list(1000)
    return {"entries": entries, "sheets": sheets}


@api_router.get("/admin/analytics/line-losses")
async def admin_line_losses_analytics(
    feeder_ids: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_admin: User = Depends(get_current_admin),
):
    query: dict[str, Any] = {}
    if feeder_ids:
        ids = [f for f in feeder_ids.split(",") if f]
        if ids:
            query["feeder_id"] = {"$in": ids}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    projection = {
        "_id": 0,
        "feeder_id": 1,
        "date": 1,
        "end1_import_consumption": 1,
        "end2_import_consumption": 1,
        "loss_percent": 1,
    }
    entries = await db.entries.find(query, projection).sort("date", 1).to_list(10000)
    return {"entries": entries}


@api_router.get("/admin/analytics/max-min")
async def admin_max_min_analytics(
    feeder_ids: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_admin: User = Depends(get_current_admin),
):
    query: dict[str, Any] = {}
    if feeder_ids:
        ids = [f for f in feeder_ids.split(",") if f]
        if ids:
            query["feeder_id"] = {"$in": ids}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    entries = await db.max_min_entries.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    feeders = await db.max_min_feeders.find({}, {"_id": 0}).to_list(1000)
    return {"entries": entries, "feeders": feeders}


@api_router.get("/admin/analytics/max-min/export")
async def export_admin_max_min_analytics(
    feeder_ids: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    columns: Optional[str] = None,
    current_admin: User = Depends(get_current_admin),
):
    try:
        query: dict[str, Any] = {}
        if feeder_ids:
            ids = [f for f in feeder_ids.split(",") if f]
            if ids:
                query["feeder_id"] = {"$in": ids}
        if start_date and end_date:
            query["date"] = {"$gte": start_date, "$lte": end_date}
        elif start_date:
            query["date"] = {"$gte": start_date}
        elif end_date:
            query["date"] = {"$lte": end_date}

        entries = await db.max_min_entries.find(query, {"_id": 0}).sort("date", 1).to_list(20000)
        feeders = await db.max_min_feeders.find({}, {"_id": 0}).to_list(1000)
        feeder_map: dict[str, dict[str, Any]] = {f["id"]: f for f in feeders if f.get("id")}

        wb = Workbook()
        ws = wb.active
        ws.title = "Max-Min Analytics"

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        all_columns = [
            ("feeder", "Feeder"),
            ("maxAmps", "Max Amps"),
            ("maxMw", "Max MW"),
            ("date", "Date"),
            ("maxTime", "Max Time"),
            ("maxMvar", "Max MVAR"),
            ("minAmps", "Min Amps"),
            ("minMw", "Min MW"),
            ("minMvar", "Min MVAR"),
            ("minDate", "Min Date"),
            ("minTime", "Min Time"),
            ("avgAmps", "Avg Amps"),
            ("avgMw", "Avg MW"),
        ]

        if columns:
            requested = [c.strip() for c in columns.split(",") if c.strip()]
            selected_keys = [key for key, _ in all_columns if key in requested]
        else:
            selected_keys = [key for key, _ in all_columns]

        if not selected_keys:
            selected_keys = [key for key, _ in all_columns]

        headers = [label for key, label in all_columns if key in selected_keys]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        for e in entries:
            data = e.get("data") or {}
            max_data = data.get("max") or {}
            min_data = data.get("min") or {}
            avg_data = data.get("avg") or {}
            feeder_id = e.get("feeder_id")
            feeder = feeder_map.get(feeder_id) or {}

            row: list[Any] = []
            for key in selected_keys:
                if key == "date":
                    row.append(format_date(e.get("date")))
                elif key == "feeder":
                    row.append(feeder.get("name") or feeder_id or "")
                elif key == "maxTime":
                    row.append(format_time(max_data.get("time")))
                elif key == "maxAmps":
                    row.append(max_data.get("amps") or "")
                elif key == "maxMw":
                    row.append(max_data.get("mw") or "")
                elif key == "maxMvar":
                    row.append(max_data.get("mvar") or "")
                elif key == "minAmps":
                    row.append(min_data.get("amps") or "")
                elif key == "minMw":
                    row.append(min_data.get("mw") or "")
                elif key == "minMvar":
                    row.append(min_data.get("mvar") or "")
                elif key == "minDate":
                    row.append(format_date(e.get("date")))
                elif key == "minTime":
                    row.append(format_time(min_data.get("time")))
                elif key == "avgAmps":
                    row.append(avg_data.get("amps") or "")
                elif key == "avgMw":
                    row.append(avg_data.get("mw") or "")
            ws.append(row)

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                try:
                    cell_val = str(cell.value) if cell.value is not None else ""
                    if len(cell_val) > max_length:
                        max_length = len(cell_val)
                except Exception:
                    continue
            ws.column_dimensions[col_letter].width = max_length + 2

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename_parts = ["Admin_MaxMin_Analytics"]
        if start_date:
            filename_parts.append(start_date)
        if end_date:
            filename_parts.append(end_date)
        filename = "_".join(filename_parts) + ".xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": str(e)})


@api_router.post("/admin/analytics/max-min/export-view")
async def export_admin_max_min_view(
    payload: AdminMaxMinExportViewPayload,
    current_admin: User = Depends(get_current_admin),
):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Max-Min Analytics View"

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = [col.label for col in payload.columns]
        fields = [col.field for col in payload.columns]

        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        for row_data in payload.rows:
            excel_row: list[Any] = []
            for field in fields:
                value = row_data.get(field, "")
                excel_row.append(value if value is not None else "")
            ws.append(excel_row)

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                try:
                    cell_val = str(cell.value) if cell.value is not None else ""
                    if len(cell_val) > max_length:
                        max_length = len(cell_val)
                except Exception:
                    continue
            ws.column_dimensions[col_letter].width = max_length + 2

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "Admin_MaxMin_Analytics_View.xlsx"
        meta = payload.meta or {}
        start_date = meta.get("startDate") or meta.get("start_date")
        end_date = meta.get("endDate") or meta.get("end_date")
        if start_date or end_date:
            parts = ["Admin_MaxMin_Analytics"]
            if start_date:
                parts.append(str(start_date))
            if end_date:
                parts.append(str(end_date))
            filename = "_".join(parts) + ".xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        print(error_msg)
        return JSONResponse(status_code=500, content={"detail": str(e)})

@api_router.get("/admin/analytics/station-load")
async def admin_station_load_analytics(
    feeder_ids: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    mode: str = "day",
    current_admin: User = Depends(get_current_admin),
):
    mode_normalized = (mode or "day").lower()
    if mode_normalized not in {"day", "month"}:
        raise HTTPException(status_code=400, detail="mode must be 'day' or 'month'")
    ict_feeders = await db.max_min_feeders.find({"type": "ict_feeder"}, {"_id": 0}).to_list(1000)
    feeder_map: dict[str, dict[str, Any]] = {f["id"]: f for f in ict_feeders if f.get("id")}
    if feeder_ids:
        selected_ids = [f for f in feeder_ids.split(",") if f]
        selected_ids = [fid for fid in selected_ids if fid in feeder_map]
    else:
        selected_ids = list(feeder_map.keys())
    if not selected_ids:
        return {"mode": mode_normalized, "feeders": ict_feeders, "rows": []}
    query: dict[str, Any] = {"feeder_id": {"$in": selected_ids}}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    projection = {"_id": 0, "feeder_id": 1, "date": 1, "data": 1}
    raw_entries = await db.max_min_entries.find(query, projection).sort("date", 1).to_list(20000)

    def to_float(value: Any) -> float | None:
        if value is None:
            return None
        s = str(value).strip()
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            return None

    rows_map: dict[str, dict[str, Any]] = {}
    if mode_normalized == "day":
        for e in raw_entries:
            date_str = e.get("date")
            feeder_id = e.get("feeder_id")
            if not date_str or feeder_id not in feeder_map:
                continue
            data = e.get("data") or {}
            max_data = data.get("max") or {}
            amps = to_float(max_data.get("amps"))
            mw = to_float(max_data.get("mw"))
            mvar = to_float(max_data.get("mvar"))
            time_str = max_data.get("time")
            key = date_str
            if key not in rows_map:
                rows_map[key] = {"period": key, "per_ict": {}, "station_mw": 0.0, "station_mvar": 0.0}
            per_ict = rows_map[key]["per_ict"]
            per_ict[feeder_id] = {
                "feeder_id": feeder_id,
                "feeder_name": feeder_map[feeder_id].get("name"),
                "date": date_str,
                "time": time_str,
                "amps": amps,
                "mw": mw,
                "mvar": mvar,
            }
            if mw is not None:
                rows_map[key]["station_mw"] += mw
            if mvar is not None:
                rows_map[key]["station_mvar"] += mvar
    else:
        for e in raw_entries:
            date_str = e.get("date")
            feeder_id = e.get("feeder_id")
            if not date_str or feeder_id not in feeder_map:
                continue
            data = e.get("data") or {}
            max_data = data.get("max") or {}
            amps = to_float(max_data.get("amps"))
            mw = to_float(max_data.get("mw"))
            mvar = to_float(max_data.get("mvar"))
            time_str = max_data.get("time")
            if not mw and not mvar and not amps:
                continue
            month_key = date_str[:7]
            if month_key not in rows_map:
                rows_map[month_key] = {"period": month_key, "per_ict": {}}
            per_ict = rows_map[month_key]["per_ict"]
            existing = per_ict.get(feeder_id)
            replace = False
            mw_val = mw if mw is not None else 0.0
            if existing is None:
                replace = True
            else:
                existing_mw = existing.get("mw")
                existing_mw_val = existing_mw if isinstance(existing_mw, (int, float)) else 0.0
                if mw_val > existing_mw_val:
                    replace = True
            if replace:
                per_ict[feeder_id] = {
                    "feeder_id": feeder_id,
                    "feeder_name": feeder_map[feeder_id].get("name"),
                    "date": date_str,
                    "time": time_str,
                    "amps": amps,
                    "mw": mw,
                    "mvar": mvar,
                }
    rows: list[dict[str, Any]] = []
    for key in sorted(rows_map.keys()):
        row = rows_map[key]
        per_ict = row.get("per_ict") or {}
        if mode_normalized == "day":
            station_mw = row.get("station_mw", 0.0)
            station_mvar = row.get("station_mvar", 0.0)
        else:
            station_mw = 0.0
            station_mvar = 0.0
            for v in per_ict.values():
                mw = v.get("mw")
                mvar = v.get("mvar")
                if isinstance(mw, (int, float)):
                    station_mw += mw
                if isinstance(mvar, (int, float)):
                    station_mvar += mvar
        station_mva = (station_mw ** 2 + station_mvar ** 2) ** 0.5 if station_mw or station_mvar else None
        ict_values: dict[str, Any] = {}
        for fid, v in per_ict.items():
            mw = v.get("mw")
            mvar = v.get("mvar")
            if isinstance(mw, (int, float)) or isinstance(mvar, (int, float)):
                ict_mva = None
                if isinstance(mw, (int, float)) and isinstance(mvar, (int, float)):
                    ict_mva = (mw ** 2 + mvar ** 2) ** 0.5
            else:
                ict_mva = None
            ict_values[fid] = {**v, "mva": ict_mva}
        rows.append(
            {
                "period": row["period"],
                "per_ict": ict_values,
                "station": {"mw": station_mw, "mvar": station_mvar, "mva": station_mva},
            }
        )
    return {"mode": mode_normalized, "feeders": [feeder_map[fid] for fid in selected_ids], "rows": rows}


@api_router.get("/admin/analytics/station-load/export")
async def admin_station_load_export(
    feeder_ids: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    mode: str = "day",
    metrics: Optional[str] = None,
    include_station: bool = True,
    current_admin: User = Depends(get_current_admin),
):
    mode_normalized = (mode or "day").lower()
    if mode_normalized not in {"day", "month"}:
        raise HTTPException(status_code=400, detail="mode must be 'day' or 'month'")
    ict_feeders = await db.max_min_feeders.find({"type": "ict_feeder"}, {"_id": 0}).to_list(1000)
    feeder_map: dict[str, dict[str, Any]] = {f["id"]: f for f in ict_feeders if f.get("id")}
    if feeder_ids:
        selected_ids = [f for f in feeder_ids.split(",") if f]
        selected_ids = [fid for fid in selected_ids if fid in feeder_map]
    else:
        selected_ids = list(feeder_map.keys())
    if not selected_ids:
        raise HTTPException(status_code=400, detail="No valid ICT feeders selected")
    query: dict[str, Any] = {"feeder_id": {"$in": selected_ids}}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    projection = {"_id": 0, "feeder_id": 1, "date": 1, "data": 1}
    raw_entries = await db.max_min_entries.find(query, projection).sort("date", 1).to_list(20000)

    def to_float(value: Any) -> float | None:
        if value is None:
            return None
        s = str(value).strip()
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            return None

    rows_map: dict[str, dict[str, Any]] = {}
    if mode_normalized == "day":
        for e in raw_entries:
            date_str = e.get("date")
            feeder_id = e.get("feeder_id")
            if not date_str or feeder_id not in feeder_map:
                continue
            data = e.get("data") or {}
            max_data = data.get("max") or {}
            amps = to_float(max_data.get("amps"))
            mw = to_float(max_data.get("mw"))
            mvar = to_float(max_data.get("mvar"))
            time_str = max_data.get("time")
            key = date_str
            if key not in rows_map:
                rows_map[key] = {"period": key, "per_ict": {}, "station_mw": 0.0, "station_mvar": 0.0}
            per_ict = rows_map[key]["per_ict"]
            per_ict[feeder_id] = {
                "feeder_id": feeder_id,
                "feeder_name": feeder_map[feeder_id].get("name"),
                "date": date_str,
                "time": time_str,
                "amps": amps,
                "mw": mw,
                "mvar": mvar,
            }
            if mw is not None:
                rows_map[key]["station_mw"] += mw
            if mvar is not None:
                rows_map[key]["station_mvar"] += mvar
    else:
        for e in raw_entries:
            date_str = e.get("date")
            feeder_id = e.get("feeder_id")
            if not date_str or feeder_id not in feeder_map:
                continue
            data = e.get("data") or {}
            max_data = data.get("max") or {}
            amps = to_float(max_data.get("amps"))
            mw = to_float(max_data.get("mw"))
            mvar = to_float(max_data.get("mvar"))
            time_str = max_data.get("time")
            if not mw and not mvar and not amps:
                continue
            month_key = date_str[:7]
            if month_key not in rows_map:
                rows_map[month_key] = {"period": month_key, "per_ict": {}}
            per_ict = rows_map[month_key]["per_ict"]
            existing = per_ict.get(feeder_id)
            replace = False
            mw_val = mw if mw is not None else 0.0
            if existing is None:
                replace = True
            else:
                existing_mw = existing.get("mw")
                existing_mw_val = existing_mw if isinstance(existing_mw, (int, float)) else 0.0
                if mw_val > existing_mw_val:
                    replace = True
            if replace:
                per_ict[feeder_id] = {
                    "feeder_id": feeder_id,
                    "feeder_name": feeder_map[feeder_id].get("name"),
                    "date": date_str,
                    "time": time_str,
                    "amps": amps,
                    "mw": mw,
                    "mvar": mvar,
                }

    metric_order = ["amps", "mw", "mvar", "mva"]
    station_order = ["mva", "mw", "mvar"]
    if metrics:
        requested = [m.strip().lower() for m in metrics.split(",") if m.strip()]
        active_metric_keys = [m for m in metric_order if m in requested]
    else:
        active_metric_keys = metric_order.copy()
    if not active_metric_keys:
        raise HTTPException(status_code=400, detail="At least one metric must be selected")
    if include_station:
        station_metric_keys = [m for m in station_order if m in active_metric_keys]
    else:
        station_metric_keys = []

    # Sort ICTs by name to match frontend table order
    sorted_ids = sorted(
        selected_ids,
        key=lambda fid: (feeder_map[fid].get("name") or "").lower(),
    )

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = wb.create_sheet("Station Load")

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    metric_label_map = {"amps": "Amps", "mw": "MW", "mvar": "MVAR", "mva": "MVA"}

    # Header row 1 and 2 to mirror on-screen table:
    # Row 1: Date, Time, ICT group headers, Station Load group header
    # Row 2: sub-headers Amps/MW/MVAR/MVA under each group

    # Date / Time with vertical merge (row span 2)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    c_date = ws.cell(row=1, column=1, value="Date")
    c_date.fill = header_fill
    c_date.font = header_font
    c_date.alignment = header_align

    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    c_time = ws.cell(row=1, column=2, value="Time")
    c_time.fill = header_fill
    c_time.font = header_font
    c_time.alignment = header_align

    col_idx = 3

    # ICT group headers and sub-headers
    for fid in sorted_ids:
        feeder = feeder_map[fid]
        name = feeder.get("name", fid)
        span = len(active_metric_keys) or 1
        if span > 1:
            ws.merge_cells(
                start_row=1,
                start_column=col_idx,
                end_row=1,
                end_column=col_idx + span - 1,
            )
        gcell = ws.cell(row=1, column=col_idx, value=name)
        gcell.fill = header_fill
        gcell.font = header_font
        gcell.alignment = header_align

        for m in active_metric_keys:
            subcell = ws.cell(row=2, column=col_idx, value=metric_label_map[m])
            subcell.fill = header_fill
            subcell.font = header_font
            subcell.alignment = header_align
            col_idx += 1

    # Station Load group headers and sub-headers (if included)
    if station_metric_keys:
        span = len(station_metric_keys)
        if span > 1:
            ws.merge_cells(
                start_row=1,
                start_column=col_idx,
                end_row=1,
                end_column=col_idx + span - 1,
            )
        scell = ws.cell(row=1, column=col_idx, value="Station Load")
        scell.fill = header_fill
        scell.font = header_font
        scell.alignment = header_align

        for m in station_metric_keys:
            subcell = ws.cell(row=2, column=col_idx, value=metric_label_map[m])
            subcell.fill = header_fill
            subcell.font = header_font
            subcell.alignment = header_align
            col_idx += 1

    def format_int(value: Any) -> Any:
        if not isinstance(value, (int, float)):
            return None
        return int(round(value))

    def format_dec(value: Any) -> Any:
        if not isinstance(value, (int, float)):
            return None
        return round(float(value), 2)

    for key in sorted(rows_map.keys()):
        row_info = rows_map[key]
        per_ict = row_info.get("per_ict") or {}
        if mode_normalized == "day":
            station_mw = row_info.get("station_mw", 0.0)
            station_mvar = row_info.get("station_mvar", 0.0)
        else:
            station_mw = 0.0
            station_mvar = 0.0
            for v in per_ict.values():
                mw_val = v.get("mw")
                mvar_val = v.get("mvar")
                if isinstance(mw_val, (int, float)):
                    station_mw += mw_val
                if isinstance(mvar_val, (int, float)):
                    station_mvar += mvar_val
        station_mva = (station_mw ** 2 + station_mvar ** 2) ** 0.5 if station_mw or station_mvar else None

        ict_values: dict[str, Any] = {}
        for fid, v in per_ict.items():
            mw_val = v.get("mw")
            mvar_val = v.get("mvar")
            if isinstance(mw_val, (int, float)) and isinstance(mvar_val, (int, float)):
                ict_mva = (mw_val ** 2 + mvar_val ** 2) ** 0.5
            else:
                ict_mva = None
            ict_values[fid] = {**v, "mva": ict_mva}

        best_date = None
        best_time = None
        best_mw_val = None
        for v in ict_values.values():
            mw_val = v.get("mw")
            if isinstance(mw_val, (int, float)):
                if best_mw_val is None or mw_val > best_mw_val:
                    best_mw_val = mw_val
                    best_date = v.get("date")
                    best_time = v.get("time")
        if best_date:
            date_str = format_date(best_date)
        else:
            date_str = row_info.get("period") or ""
        time_str = best_time or ""

        excel_row: list[Any] = [date_str, time_str]
        for fid in sorted_ids:
            v = ict_values.get(fid) or {}
            for m in active_metric_keys:
                if m == "amps":
                    excel_row.append(format_int(v.get("amps")))
                elif m == "mw":
                    excel_row.append(format_int(v.get("mw")))
                elif m == "mvar":
                    excel_row.append(format_dec(v.get("mvar")))
                else:
                    excel_row.append(format_dec(v.get("mva")))
        for m in station_metric_keys:
            if m == "mva":
                excel_row.append(format_dec(station_mva))
            elif m == "mw":
                excel_row.append(format_int(station_mw))
            else:
                excel_row.append(format_dec(station_mvar))
        ws.append(excel_row)

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            try:
                cell_val = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(cell_val))
            except Exception:
                continue
        ws.column_dimensions[col_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename_parts = ["StationLoad", mode_normalized]
    if start_date:
        filename_parts.append(start_date)
    if end_date:
        filename_parts.append(end_date)
    filename = "_".join(filename_parts) + ".xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api_router.get("/admin/analytics/interruptions")
async def admin_interruptions_analytics(
    feeder_ids: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_admin: User = Depends(get_current_admin),
):
    query: dict[str, Any] = {}
    if feeder_ids:
        ids = [f for f in feeder_ids.split(",") if f]
        if ids:
            query["feeder_id"] = {"$in": ids}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["date"] = {"$gte": start_date}
    elif end_date:
        query["date"] = {"$lte": end_date}
    entries = await db.interruption_entries.find(query, {"_id": 0}).sort("date", 1).to_list(10000)
    feeders = await db.max_min_feeders.find({}, {"_id": 0}).to_list(1000)
    return {"entries": entries, "feeders": feeders}


@api_router.post("/admin/bulk-import/energy")
async def admin_bulk_import_energy(
    payload: dict,
    current_admin: User = Depends(get_current_admin),
):
    sheet_id = payload.get("sheet_id")
    entries = payload.get("entries") or []
    overwrite = bool(payload.get("overwrite"))
    year = payload.get("year")
    month = payload.get("month")
    if not sheet_id or not isinstance(entries, list):
        raise HTTPException(status_code=400, detail="sheet_id and entries are required")
    total_entries = len(entries)
    inserted = 0
    skipped_existing = 0
    validation_errors = 0
    per_month: dict[str, dict[str, int]] = {}
    errors: list[dict[str, Any]] = []
    meters = await db.energy_meters.find({"sheet_id": sheet_id}, {"_id": 0}).to_list(100)
    meter_map = {m["id"]: m for m in meters}
    for idx, e in enumerate(sorted(entries, key=lambda x: x.get("date") or "")):
        date_str = e.get("date")
        if not date_str:
            validation_errors += 1
            errors.append({"index": idx, "reason": "Missing date", "date": None, "sheet_id": sheet_id})
            continue
        month_key = date_str[:7]
        if month_key not in per_month:
            per_month[month_key] = {"inserted": 0, "skipped_existing": 0, "validation_errors": 0}
        try:
            existing = await db.energy_entries.find_one({"sheet_id": sheet_id, "date": date_str})
            if existing and not overwrite:
                skipped_existing += 1
                per_month[month_key]["skipped_existing"] += 1
                continue
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            prev_date = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
            prev_entry = await db.energy_entries.find_one({"sheet_id": sheet_id, "date": prev_date}, {"_id": 0})
            prev_finals: dict[str, float] = {}
            if prev_entry:
                for r in prev_entry.get("readings", []):
                    prev_finals[r["meter_id"]] = r["final"]
            readings = []
            total_consumption = 0.0
            for r in e.get("readings", []):
                meter = meter_map.get(r.get("meter_id"))
                if not meter:
                    continue
                initial = prev_finals.get(r["meter_id"], 0.0)
                final = float(r.get("final", 0.0))
                consumption = (final - initial) * meter["mf"]
                readings.append(
                    {
                        "meter_id": r["meter_id"],
                        "initial": initial,
                        "final": final,
                        "consumption": consumption,
                    }
                )
                total_consumption += consumption
            entry_doc = EnergyEntry(
                sheet_id=sheet_id,
                date=date_str,
                readings=[EnergyReading(**rr) for rr in readings],
                total_consumption=total_consumption,
            ).model_dump()
            if isinstance(entry_doc.get("created_at"), datetime):
                entry_doc["created_at"] = entry_doc["created_at"].isoformat()
            if isinstance(entry_doc.get("updated_at"), datetime):
                entry_doc["updated_at"] = entry_doc["updated_at"].isoformat()
            if existing and overwrite:
                await db.energy_entries.replace_one({"_id": existing["_id"]}, entry_doc)
            else:
                await db.energy_entries.insert_one(entry_doc)
            inserted += 1
            per_month[month_key]["inserted"] += 1
        except Exception as exc:
            validation_errors += 1
            per_month[month_key]["validation_errors"] += 1
            errors.append(
                {"index": idx, "reason": str(exc), "date": date_str, "sheet_id": sheet_id}
            )
    per_month_rows = []
    for key, stats in sorted(per_month.items()):
        per_month_rows.append(
            {
                "month": key,
                "inserted": stats["inserted"],
                "skipped_existing": stats["skipped_existing"],
                "validation_errors": stats["validation_errors"],
            }
        )
    return {
        "module": "Energy",
        "year": year,
        "month": month,
        "overwrite": overwrite,
        "total_entries": total_entries,
        "inserted": inserted,
        "skipped_existing": skipped_existing,
        "validation_errors": validation_errors,
        "per_month": per_month_rows,
        "errors": errors,
    }


@api_router.post("/admin/bulk-import/check/energy")
async def admin_bulk_import_check_energy(
    payload: dict,
    current_admin: User = Depends(get_current_admin),
):
    sheet_id = payload.get("sheet_id")
    year = payload.get("year")
    month = payload.get("month")
    if not sheet_id:
        raise HTTPException(status_code=400, detail="sheet_id is required")
    query: dict[str, Any] = {"sheet_id": sheet_id}
    if isinstance(year, int):
        y = year
        if isinstance(month, int):
            start_date = f"{y:04d}-{month:02d}-01"
            end_date = f"{y:04d}-{month:02d}-31"
        else:
            start_date = f"{y:04d}-01-01"
            end_date = f"{y:04d}-12-31"
        query["date"] = {"$gte": start_date, "$lte": end_date}
    count = await db.energy_entries.count_documents(query)
    return {
        "module": "Energy",
        "sheet_id": sheet_id,
        "year": year,
        "month": month,
        "has_existing": count > 0,
        "total_existing": int(count),
    }


@api_router.post("/admin/bulk-import/line-losses")
async def admin_bulk_import_line_losses(
    payload: dict,
    current_admin: User = Depends(get_current_admin),
):
    feeder_id = payload.get("feeder_id")
    entries = payload.get("entries") or []
    overwrite = bool(payload.get("overwrite"))
    year = payload.get("year")
    month = payload.get("month")
    if not feeder_id or not isinstance(entries, list):
        raise HTTPException(status_code=400, detail="feeder_id and entries are required")
    feeder = await db.feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    total_entries = len(entries)
    inserted = 0
    skipped_existing = 0
    validation_errors = 0
    per_month: dict[str, dict[str, int]] = {}
    errors: list[dict[str, Any]] = []
    for idx, e in enumerate(sorted(entries, key=lambda x: x.get("date") or "")):
        date_str = e.get("date")
        if not date_str:
            validation_errors += 1
            errors.append({"index": idx, "reason": "Missing date", "date": None, "feeder_id": feeder_id})
            continue
        month_key = date_str[:7]
        if month_key not in per_month:
            per_month[month_key] = {"inserted": 0, "skipped_existing": 0, "validation_errors": 0}
        try:
            existing = await db.entries.find_one({"feeder_id": feeder_id, "date": date_str})
            if existing and not overwrite:
                skipped_existing += 1
                per_month[month_key]["skipped_existing"] += 1
                continue
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            prev_date = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
            prev_entry = await db.entries.find_one({"feeder_id": feeder_id, "date": prev_date}, {"_id": 0})
            end1_import_initial = prev_entry["end1_import_final"] if prev_entry else 0
            end1_export_initial = prev_entry["end1_export_final"] if prev_entry else 0
            end2_import_initial = prev_entry["end2_import_final"] if prev_entry else 0
            end2_export_initial = prev_entry["end2_export_final"] if prev_entry else 0
            end1_import_final = float(e.get("end1_import_final", 0) or 0)
            end1_export_final = float(e.get("end1_export_final", 0) or 0)
            end2_import_final = float(e.get("end2_import_final", 0) or 0)
            end2_export_final = float(e.get("end2_export_final", 0) or 0)
            end1_import_consumption = (
                end1_import_final - end1_import_initial
            ) * feeder["end1_import_mf"]
            end1_export_consumption = (
                end1_export_final - end1_export_initial
            ) * feeder["end1_export_mf"]
            end2_import_consumption = (
                end2_import_final - end2_import_initial
            ) * feeder["end2_import_mf"]
            end2_export_consumption = (
                end2_export_final - end2_export_initial
            ) * feeder["end2_export_mf"]
            total_import = end1_import_consumption + end2_import_consumption
            loss_percent = (
                0
                if total_import == 0
                else (
                    end1_import_consumption
                    - end1_export_consumption
                    + end2_import_consumption
                    - end2_export_consumption
                )
                / total_import
                * 100
            )
            entry_obj = DailyEntry(
                feeder_id=feeder_id,
                date=date_str,
                end1_import_initial=end1_import_initial,
                end1_import_final=end1_import_final,
                end1_export_initial=end1_export_initial,
                end1_export_final=end1_export_final,
                end2_import_initial=end2_import_initial,
                end2_import_final=end2_import_final,
                end2_export_initial=end2_export_initial,
                end2_export_final=end2_export_final,
                end1_import_consumption=end1_import_consumption,
                end1_export_consumption=end1_export_consumption,
                end2_import_consumption=end2_import_consumption,
                end2_export_consumption=end2_export_consumption,
                loss_percent=loss_percent,
            ).model_dump()
            created_at = entry_obj.get("created_at")
            updated_at = entry_obj.get("updated_at")
            if isinstance(created_at, datetime):
                entry_obj["created_at"] = created_at.isoformat()
            if isinstance(updated_at, datetime):
                entry_obj["updated_at"] = updated_at.isoformat()
            if existing and overwrite:
                await db.entries.replace_one({"_id": existing["_id"]}, entry_obj)
            else:
                await db.entries.insert_one(entry_obj)
            inserted += 1
            per_month[month_key]["inserted"] += 1
        except Exception as exc:
            validation_errors += 1
            per_month[month_key]["validation_errors"] += 1
            errors.append(
                {"index": idx, "reason": str(exc), "date": date_str, "feeder_id": feeder_id}
            )
    per_month_rows = []
    for key, stats in sorted(per_month.items()):
        per_month_rows.append(
            {
                "month": key,
                "inserted": stats["inserted"],
                "skipped_existing": stats["skipped_existing"],
                "validation_errors": stats["validation_errors"],
            }
        )
    return {
        "module": "Line Losses",
        "year": year,
        "month": month,
        "overwrite": overwrite,
        "total_entries": total_entries,
        "inserted": inserted,
        "skipped_existing": skipped_existing,
        "validation_errors": validation_errors,
        "per_month": per_month_rows,
        "errors": errors,
    }


@api_router.post("/admin/bulk-import/check/line-losses")
async def admin_bulk_import_check_line_losses(
    payload: dict,
    current_admin: User = Depends(get_current_admin),
):
    feeder_id = payload.get("feeder_id")
    year = payload.get("year")
    month = payload.get("month")
    if not feeder_id:
        raise HTTPException(status_code=400, detail="feeder_id is required")
    query: dict[str, Any] = {"feeder_id": feeder_id}
    if isinstance(year, int):
        y = year
        if isinstance(month, int):
            start_date = f"{y:04d}-{month:02d}-01"
            end_date = f"{y:04d}-{month:02d}-31"
        else:
            start_date = f"{y:04d}-01-01"
            end_date = f"{y:04d}-12-31"
        query["date"] = {"$gte": start_date, "$lte": end_date}
    count = await db.entries.count_documents(query)
    return {
        "module": "Line Losses",
        "feeder_id": feeder_id,
        "year": year,
        "month": month,
        "has_existing": count > 0,
        "total_existing": int(count),
    }


@api_router.post("/admin/bulk-import/max-min")
async def admin_bulk_import_max_min(
    payload: dict,
    current_admin: User = Depends(get_current_admin),
):
    feeder_id = payload.get("feeder_id")
    entries = payload.get("entries") or []
    overwrite = bool(payload.get("overwrite"))
    year = payload.get("year")
    month = payload.get("month")
    if not feeder_id or not isinstance(entries, list):
        raise HTTPException(status_code=400, detail="feeder_id and entries are required")
    feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
    if not feeder:
        raise HTTPException(status_code=404, detail="Feeder not found")
    total_entries = len(entries)
    inserted = 0
    skipped_existing = 0
    validation_errors = 0
    per_month: dict[str, dict[str, int]] = {}
    errors: list[dict[str, Any]] = []
    for idx, e in enumerate(sorted(entries, key=lambda x: x.get("date") or "")):
        date_str = e.get("date")
        if not date_str:
            validation_errors += 1
            errors.append({"index": idx, "reason": "Missing date", "date": None, "feeder_id": feeder_id})
            continue
        month_key = date_str[:7]
        if month_key not in per_month:
            per_month[month_key] = {"inserted": 0, "skipped_existing": 0, "validation_errors": 0}
        data = e.get("data") or {}
        try:
            existing = await db.max_min_entries.find_one({"feeder_id": feeder_id, "date": date_str})
            if existing and not overwrite:
                skipped_existing += 1
                per_month[month_key]["skipped_existing"] += 1
                continue
            if existing and overwrite:
                await db.max_min_entries.update_one(
                    {"_id": existing["_id"]},
                    {
                        "$set": {
                            "data": data,
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                        }
                    },
                )
            else:
                entry_obj = MaxMinEntry(feeder_id=feeder_id, date=date_str, data=data)
                doc = entry_obj.model_dump()
                if isinstance(doc.get("created_at"), datetime):
                    doc["created_at"] = doc["created_at"].isoformat()
                if isinstance(doc.get("updated_at"), datetime):
                    doc["updated_at"] = doc["updated_at"].isoformat()
                await db.max_min_entries.insert_one(doc)
            inserted += 1
            per_month[month_key]["inserted"] += 1
        except Exception as exc:
            validation_errors += 1
            per_month[month_key]["validation_errors"] += 1
            errors.append(
                {"index": idx, "reason": str(exc), "date": date_str, "feeder_id": feeder_id}
            )
    per_month_rows = []
    for key, stats in sorted(per_month.items()):
        per_month_rows.append(
            {
                "month": key,
                "inserted": stats["inserted"],
                "skipped_existing": stats["skipped_existing"],
                "validation_errors": stats["validation_errors"],
            }
        )
    return {
        "module": "Max-Min",
        "year": year,
        "month": month,
        "overwrite": overwrite,
        "total_entries": total_entries,
        "inserted": inserted,
        "skipped_existing": skipped_existing,
        "validation_errors": validation_errors,
        "per_month": per_month_rows,
        "errors": errors,
    }


@api_router.post("/admin/bulk-import/check/max-min")
async def admin_bulk_import_check_max_min(
    payload: dict,
    current_admin: User = Depends(get_current_admin),
):
    feeder_id = payload.get("feeder_id")
    year = payload.get("year")
    month = payload.get("month")
    if not feeder_id:
        raise HTTPException(status_code=400, detail="feeder_id is required")
    query: dict[str, Any] = {"feeder_id": feeder_id}
    if isinstance(year, int):
        y = year
        if isinstance(month, int):
            start_date = f"{y:04d}-{month:02d}-01"
            end_date = f"{y:04d}-{month:02d}-31"
        else:
            start_date = f"{y:04d}-01-01"
            end_date = f"{y:04d}-12-31"
        query["date"] = {"$gte": start_date, "$lte": end_date}
    count = await db.max_min_entries.count_documents(query)
    return {
        "module": "Max-Min",
        "feeder_id": feeder_id,
        "year": year,
        "month": month,
        "has_existing": count > 0,
        "total_existing": int(count),
    }


@api_router.post("/admin/bulk-import/interruptions")
async def admin_bulk_import_interruptions(
    payload: dict,
    current_admin: User = Depends(get_current_admin),
):
    entries = payload.get("entries") or []
    overwrite = bool(payload.get("overwrite"))
    year = payload.get("year")
    month = payload.get("month")
    if not isinstance(entries, list):
        raise HTTPException(status_code=400, detail="entries are required")
    total_entries = len(entries)
    inserted = 0
    skipped_existing = 0
    validation_errors = 0
    per_month: dict[str, dict[str, int]] = {}
    errors: list[dict[str, Any]] = []
    for idx, e in enumerate(
        sorted(entries, key=lambda x: (x.get("feeder_id") or "", x.get("date") or "", x.get("start_time") or ""))
    ):
        feeder_id = e.get("feeder_id")
        date_str = e.get("date")
        start_time = e.get("start_time")
        if not feeder_id or not date_str or not start_time:
            validation_errors += 1
            errors.append(
                {
                    "index": idx,
                    "reason": "Missing feeder_id, date, or start_time",
                    "date": date_str,
                    "feeder_id": feeder_id,
                }
            )
            continue
        month_key = date_str[:7]
        if month_key not in per_month:
            per_month[month_key] = {"inserted": 0, "skipped_existing": 0, "validation_errors": 0}
        try:
            feeder = await db.max_min_feeders.find_one({"id": feeder_id}, {"_id": 0})
            if not feeder:
                validation_errors += 1
                per_month[month_key]["validation_errors"] += 1
                errors.append(
                    {"index": idx, "reason": "Feeder not found", "date": date_str, "feeder_id": feeder_id}
                )
                continue
            if feeder.get("type") not in [
                "feeder_400kv",
                "feeder_220kv",
                "ict_feeder",
                "reactor_feeder",
                "bay_feeder",
            ]:
                validation_errors += 1
                per_month[month_key]["validation_errors"] += 1
                errors.append(
                    {
                        "index": idx,
                        "reason": "Interruptions not supported for feeder type",
                        "date": date_str,
                        "feeder_id": feeder_id,
                    }
                )
                continue
            existing = await db.interruption_entries.find_one(
                {"feeder_id": feeder_id, "date": date_str, "data.start_time": start_time},
                {"_id": 0},
            )
            if existing and not overwrite:
                skipped_existing += 1
                per_month[month_key]["skipped_existing"] += 1
                continue
            data = {
                "start_time": start_time,
                "end_time": e.get("end_time"),
                "end_date": e.get("end_date") or date_str,
                "duration_minutes": e.get("duration_minutes"),
                "description": e.get("description"),
                "cause_of_interruption": e.get("cause_of_interruption"),
                "relay_indications_lc_work": e.get("relay_indications_lc_work"),
                "breakdown_declared": e.get("breakdown_declared"),
                "fault_identified_during_patrolling": e.get(
                    "fault_identified_during_patrolling"
                ),
                "fault_location": e.get("fault_location"),
                "remarks": e.get("remarks"),
                "action_taken": e.get("action_taken"),
            }
            entry_obj = InterruptionEntry(feeder_id=feeder_id, date=date_str, data=data).model_dump()
            created_at = entry_obj.get("created_at")
            updated_at = entry_obj.get("updated_at")
            if isinstance(created_at, datetime):
                entry_obj["created_at"] = created_at.isoformat()
            if isinstance(updated_at, datetime):
                entry_obj["updated_at"] = updated_at.isoformat()
            if existing and overwrite:
                await db.interruption_entries.replace_one(
                    {
                        "feeder_id": feeder_id,
                        "date": date_str,
                        "data.start_time": start_time,
                    },
                    entry_obj,
                )
            else:
                await db.interruption_entries.insert_one(entry_obj)
            inserted += 1
            per_month[month_key]["inserted"] += 1
        except Exception as exc:
            validation_errors += 1
            per_month[month_key]["validation_errors"] += 1
            errors.append(
                {"index": idx, "reason": str(exc), "date": date_str, "feeder_id": feeder_id}
            )
    per_month_rows = []
    for key, stats in sorted(per_month.items()):
        per_month_rows.append(
            {
                "month": key,
                "inserted": stats["inserted"],
                "skipped_existing": stats["skipped_existing"],
                "validation_errors": stats["validation_errors"],
            }
        )
    return {
        "module": "Interruptions",
        "year": year,
        "month": month,
        "overwrite": overwrite,
        "total_entries": total_entries,
        "inserted": inserted,
        "skipped_existing": skipped_existing,
        "validation_errors": validation_errors,
        "per_month": per_month_rows,
        "errors": errors,
    }


@api_router.post("/admin/bulk-import/check/interruptions")
async def admin_bulk_import_check_interruptions(
    payload: dict,
    current_admin: User = Depends(get_current_admin),
):
    feeder_ids = payload.get("feeder_ids") or []
    year = payload.get("year")
    month = payload.get("month")
    if not feeder_ids or not isinstance(feeder_ids, list):
        raise HTTPException(status_code=400, detail="feeder_ids are required")
    query: dict[str, Any] = {"feeder_id": {"$in": feeder_ids}}
    if isinstance(year, int):
        y = year
        if isinstance(month, int):
            start_date = f"{y:04d}-{month:02d}-01"
            end_date = f"{y:04d}-{month:02d}-31"
        else:
            start_date = f"{y:04d}-01-01"
            end_date = f"{y:04d}-12-31"
        query["date"] = {"$gte": start_date, "$lte": end_date}
    count = await db.interruption_entries.count_documents(query)
    return {
        "module": "Interruptions",
        "feeder_ids": feeder_ids,
        "year": year,
        "month": month,
        "has_existing": count > 0,
        "total_existing": int(count),
    }


@api_router.get("/daily-status")
async def check_daily_status(current_user: User = Depends(get_current_user)) -> dict:
    try:
        today = datetime.now().date()
        today_str = today.strftime("%Y-%m-%d")
        
        # 1. Line Losses
        feeders = await db.feeders.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
        feeder_ids = [f["id"] for f in feeders]
        feeder_map = {f["id"]: f["name"] for f in feeders}

        line_losses_entries = await db.entries.find(
            {
                "date": today_str,
                "feeder_id": {"$in": feeder_ids},
            },
            {"feeder_id": 1},
        ).to_list(None)

        entered_feeder_ids = {e["feeder_id"] for e in line_losses_entries}

        latest_line_entry = await db.entries.find_one(
            {
                "feeder_id": {"$in": feeder_ids},
                "date": {"$lte": today_str},
            },
            {"date": 1, "_id": 0},
            sort=[("date", -1)],
        )
        latest_line_date = latest_line_entry["date"] if latest_line_entry and latest_line_entry.get("date") else None

        # Group missing feeders by category so that when ALL feeders in a
        # category are pending we show a single aggregated label instead of
        # listing every feeder name.
        def _line_category(name: str) -> str:
            upper = (name or "").upper()
            if upper.startswith("400 KV ") or upper.startswith("400KV "):
                return "400KV"
            if upper.startswith("220 KV ") or upper.startswith("220KV "):
                return "220KV"
            if "ICT" in upper:
                return "ICT"
            return "OTHER"

        line_cat_feeders: dict[str, list[str]] = {"400KV": [], "220KV": [], "ICT": [], "OTHER": []}
        for f in feeders:
            cat = _line_category(f["name"])
            if cat not in line_cat_feeders:
                line_cat_feeders[cat] = []
            line_cat_feeders[cat].append(f["id"])

        missing_line_labels: list[str] = []
        for cat, ids in line_cat_feeders.items():
            if not ids:
                continue
            missing_ids = [fid for fid in ids if fid not in entered_feeder_ids]
            if not missing_ids:
                continue
            if len(missing_ids) == len(ids):
                if cat == "400KV":
                    missing_line_labels.append("All 400KV Feeders")
                    continue
                if cat == "220KV":
                    missing_line_labels.append("All 220KV Feeders")
                    continue
                if cat == "ICT":
                    missing_line_labels.append("All ICT’s")
                    continue
            # Partial category or OTHER: list individual names
            missing_line_labels.extend(feeder_map[fid] for fid in missing_ids)

        line_losses_status = {
            "complete": len(missing_line_labels) == 0,
            "missing_feeders": missing_line_labels,
            "missing_dates": [],
            "latest_entry_date": latest_line_date,
        }
        
        # If no entries for today, check last 2 days
        if len(entered_feeder_ids) == 0:
             # Check yesterday
             yesterday = today - timedelta(days=1)
             yesterday_str = yesterday.strftime("%Y-%m-%d")
             yesterday_count = await db.entries.count_documents({"date": yesterday_str})
             
             if yesterday_count == 0:
                 # Missing today and yesterday
                 line_losses_status["missing_dates"] = [yesterday_str, today_str]

        # 2. Energy Consumption
        sheets = await db.energy_sheets.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
        sheet_ids = [s["id"] for s in sheets]
        sheet_map = {s["id"]: s["name"] for s in sheets}

        energy_entries = await db.energy_entries.find(
            {
                "date": today_str,
                "sheet_id": {"$in": sheet_ids},
            },
            {"sheet_id": 1},
        ).to_list(None)

        entered_sheet_ids = {e["sheet_id"] for e in energy_entries}

        latest_energy_entry = await db.energy_entries.find_one(
            {
                "sheet_id": {"$in": sheet_ids},
                "date": {"$lte": today_str},
            },
            {"date": 1, "_id": 0},
            sort=[("date", -1)],
        )
        latest_energy_date = (
            latest_energy_entry["date"]
            if latest_energy_entry and latest_energy_entry.get("date")
            else None
        )

        # Group ICT sheets so that when all ICT sheets are pending we show
        # "All ICT’s" instead of listing ICT-1..4 individually.
        ict_sheet_ids = [s["id"] for s in sheets if (s["name"] or "").upper().startswith("ICT-")]
        other_sheet_ids = [s["id"] for s in sheets if s["id"] not in ict_sheet_ids]

        missing_sheet_labels: list[str] = []

        # ICT group
        if ict_sheet_ids:
            ict_missing_ids = [sid for sid in ict_sheet_ids if sid not in entered_sheet_ids]
            if ict_missing_ids:
                if len(ict_missing_ids) == len(ict_sheet_ids):
                    missing_sheet_labels.append("All ICT’s")
                else:
                    missing_sheet_labels.extend(sheet_map[sid] for sid in ict_missing_ids)

        # Other sheets (e.g. 33KV) always listed individually
        for sid in other_sheet_ids:
            if sid not in entered_sheet_ids:
                missing_sheet_labels.append(sheet_map[sid])

        energy_status = {
            "complete": len(missing_sheet_labels) == 0,
            "missing_sheets": missing_sheet_labels,
            "missing_dates": [],
            "latest_entry_date": latest_energy_date,
        }
        
        if len(entered_sheet_ids) == 0:
             yesterday = today - timedelta(days=1)
             yesterday_str = yesterday.strftime("%Y-%m-%d")
             yesterday_count = await db.energy_entries.count_documents({"date": yesterday_str})
             
             if yesterday_count == 0:
                 energy_status["missing_dates"] = [yesterday_str, today_str]

        # 3. Max-Min Data
        mm_feeders = await db.max_min_feeders.find(
            {},
            {"_id": 0, "id": 1, "name": 1, "type": 1},
        ).to_list(None)

        # Exclude Bus + Station (legacy), 125MVAR Bus Reactor and all bay feeders
        # from the reminder logic, since Max–Min data is not entered for these.
        def _include_mm_feeder(f: dict) -> bool:
            t = f.get("type")
            name = f.get("name") or ""
            if t == "bay_feeder":
                return False
            if name in ["Bus + Station", "125MVAR Bus Reactor"]:
                return False
            return True

        mm_feeders = [f for f in mm_feeders if _include_mm_feeder(f)]

        mm_feeder_ids = [f["id"] for f in mm_feeders]
        mm_feeder_map = {f["id"]: f["name"] for f in mm_feeders}

        mm_entries = await db.max_min_entries.find(
            {
                "date": today_str,
                "feeder_id": {"$in": mm_feeder_ids},
            },
            {"feeder_id": 1},
        ).to_list(None)

        entered_mm_ids = {e["feeder_id"] for e in mm_entries}

        latest_mm_entry = await db.max_min_entries.find_one(
            {
                "feeder_id": {"$in": mm_feeder_ids},
                "date": {"$lte": today_str},
            },
            {"date": 1, "_id": 0},
            sort=[("date", -1)],
        )
        latest_mm_date = (
            latest_mm_entry["date"]
            if latest_mm_entry and latest_mm_entry.get("date")
            else None
        )

        # Group Max–Min feeders by type (400KV, 220KV, ICT) and aggregate when
        # all feeders in a category are pending.
        def _mm_category(feeder: dict) -> str:
            t = feeder.get("type")
            if t == "feeder_400kv":
                return "400KV"
            if t == "feeder_220kv":
                return "220KV"
            if t == "ict_feeder":
                return "ICT"
            return "OTHER"

        mm_cat_feeders: dict[str, list[str]] = {"400KV": [], "220KV": [], "ICT": [], "OTHER": []}
        for f in mm_feeders:
            cat = _mm_category(f)
            if cat not in mm_cat_feeders:
                mm_cat_feeders[cat] = []
            mm_cat_feeders[cat].append(f["id"])

        missing_mm_labels: list[str] = []
        for cat, ids in mm_cat_feeders.items():
            if not ids:
                continue
            missing_ids = [fid for fid in ids if fid not in entered_mm_ids]
            if not missing_ids:
                continue
            if len(missing_ids) == len(ids):
                if cat == "400KV":
                    missing_mm_labels.append("All 400KV Feeders")
                    continue
                if cat == "220KV":
                    missing_mm_labels.append("All 220KV Feeders")
                    continue
                if cat == "ICT":
                    missing_mm_labels.append("All ICT’s")
                    continue
            # Partial category or OTHER: list individual names
            missing_mm_labels.extend(mm_feeder_map[fid] for fid in missing_ids)

        max_min_status = {
            "complete": len(missing_mm_labels) == 0,
            "missing_feeders": missing_mm_labels,
            "missing_dates": [],
            "latest_entry_date": latest_mm_date,
        }
        
        if len(entered_mm_ids) == 0:
             yesterday = today - timedelta(days=1)
             yesterday_str = yesterday.strftime("%Y-%m-%d")
             yesterday_count = await db.max_min_entries.count_documents({"date": yesterday_str})
             
             if yesterday_count == 0:
                 max_min_status["missing_dates"] = [yesterday_str, today_str]

        return {
            "line_losses": line_losses_status,
            "energy_consumption": energy_status,
            "max_min": max_min_status
        }

    except Exception as e:
        print(f"Error checking daily status: {e}")
        raise HTTPException(status_code=500, detail=str(e))

app.include_router(api_router)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
